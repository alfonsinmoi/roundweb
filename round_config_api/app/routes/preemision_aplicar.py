"""POST /api/cuotas/preemision/<mes>/aplicar-correcciones-excel

Lee el Excel de validación pre-emisión que el usuario ha editado, aplica
los cambios propuestos (categoría / cuota / periodicidad) y deja la pre-
emisión lista para volver a validarse desde la UI.

Reglas (consensuadas con el usuario, jun 2026):
- ALL-OR-NOTHING: si CUALQUIER fila tiene un valor inválido, no se aplica
  NADA y se devuelve la lista de errores.
- Categoría → upsert en cliente_categoria (solo BD local, no NoofitPro).
- Cuota / periodicidad → replace en round.subscription de Odoo (cierra la
  activa, crea la nueva). Si el manager no tiene Odoo cuotas activo, se
  rechaza con feature_not_enabled (el decorador @require_feature lo cubre).
- El cliente_idnoofit (col A) es la clave; el Excel ya solo contiene los
  clientes coherentes (lo demás está en 'INCOHERENCIAS').
"""
from __future__ import annotations

from io import BytesIO
import datetime as dt
import logging

from flask import Blueprint, request, jsonify, g
from openpyxl import load_workbook

from ..db import get_conn
from ..auth import auth_required, require_permission
from ..odoo_guard import require_feature
from ..odoo_cuotas import OdooCuotas
from ..audit_log import log_action, actor_from_request
from .preemision_validar import PERIODICIDADES_VALIDAS

bp = Blueprint('preemision_aplicar', __name__)
log = logging.getLogger('preemision_aplicar')


def _odoo():
    return OdooCuotas()


def _company_id():
    """Resuelve company_id Odoo para el manager actual (mismo patrón que
    subscriptions.py).  Si OdooCuotas no expone helper, lo derivamos por
    su identidad cargada al construirse."""
    o = _odoo()
    cid = getattr(o, 'company_id', None)
    if cid:
        return cid
    # Fallback: leer del manager
    try:
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                SELECT odoo_company_id FROM manager_config
                WHERE id_manager=%s
            """, (str(g.id_manager),))
            r = cur.fetchone()
            if r and r.get('odoo_company_id'):
                return int(r['odoo_company_id'])
    except Exception:
        pass
    return None


def _norm(v):
    """Normaliza una celda: None → '', strip, str."""
    if v is None:
        return ''
    s = str(v).strip()
    return s


def _read_sheet(ws, cli_col_name):
    """Lee una hoja (OK o INCOHERENCIAS) y devuelve (header_idx, rows).
    cli_col_name es el header donde está el id de cliente — varía entre
    'Cód. cliente' (OK) e 'idnoofit' (INCOHERENCIAS)."""
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return None, []
    header_idx = {}
    for i, h in enumerate(header):
        if h:
            header_idx[str(h).strip()] = i
    # Si falta la columna del cliente o las 3 de corrección, la hoja
    # no es relevante (puede ser un Excel viejo). Devolvemos vacío.
    requeridas = [cli_col_name, 'Nueva categoría', 'Nueva cuota',
                  'Nueva periodicidad']
    if any(r not in header_idx for r in requeridas):
        return None, []
    rows = []
    for row in rows_iter:
        if not row:
            continue
        first = _norm(row[0] if len(row) > 0 else '')
        if first.upper() == 'TOTAL':
            continue
        rows.append(row)
    return header_idx, rows


def _read_xlsx_correcciones(file_storage):
    """Carga el .xlsx subido y devuelve lista de (origen, header_idx, rows)
    leyendo tanto 'OK' como 'INCOHERENCIAS'. Si NINGUNA hoja válida tiene
    cabeceras de corrección, lanza ValueError."""
    wb = load_workbook(filename=BytesIO(file_storage.read()),
                       data_only=True, read_only=True)
    fuentes = []
    if 'OK' in wb.sheetnames:
        hi, rs = _read_sheet(wb['OK'], 'Cód. cliente')
        if hi is not None:
            fuentes.append(('OK', 'Cód. cliente', hi, rs))
    if 'INCOHERENCIAS' in wb.sheetnames:
        hi, rs = _read_sheet(wb['INCOHERENCIAS'], 'idnoofit')
        if hi is not None:
            fuentes.append(('INCOHERENCIAS', 'idnoofit', hi, rs))
    if not fuentes:
        raise ValueError('El Excel no contiene hojas "OK" o "INCOHERENCIAS" '
                         'con las columnas "Nueva categoría/cuota/periodicidad". '
                         'Descarga uno nuevo desde "Validar antes de emitir".')
    return fuentes


@bp.route('/<mes>/aplicar-correcciones-excel', methods=['POST'])
@auth_required
@require_feature('cuotas')
@require_permission('cuotas_clientes.reemplazar')
def aplicar_correcciones_excel(mes):
    """Aplica las correcciones del Excel editado a la BD local + Odoo.
    Valida TODO antes de tocar nada; si una fila falla, devuelve errores
    y no aplica nada (regla all-or-nothing consensuada con usuario)."""
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'falta_fichero',
                        'msg': 'Adjunta el .xlsx en el campo "file"'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'ok': False, 'error': 'extension_invalida',
                        'msg': 'El fichero debe ser .xlsx'}), 400

    try:
        fuentes = _read_xlsx_correcciones(f)
    except Exception as e:
        return jsonify({'ok': False, 'error': 'xlsx_invalido',
                        'msg': str(e)}), 400

    # ── Cargar catálogos del manager para validar valores ──
    id_manager = str(g.id_manager)
    id_trainer = str(g.id_trainer) if getattr(g, 'id_trainer', None) else None
    try:
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                SELECT id, nombre FROM categoria
                WHERE id_manager=%s AND activa=TRUE
            """, (id_manager,))
            # get_conn() devuelve filas como dict (row_factory=dict_row).
            categorias_map = {r['nombre']: r['id'] for r in cur.fetchall()}

            if id_trainer:
                cur.execute("""
                    SELECT DISTINCT codigo, id, scope, id_trainer
                    FROM cuota
                    WHERE id_manager=%s
                      AND (scope='plantilla_manager'
                           OR (scope='trainer' AND id_trainer=%s))
                """, (id_manager, id_trainer))
            else:
                cur.execute("""
                    SELECT DISTINCT codigo, id, scope, id_trainer
                    FROM cuota
                    WHERE id_manager=%s
                """, (id_manager,))
            # Si hay códigos duplicados (plantilla + trainer override), nos
            # quedamos con el del trainer. Si no, con la plantilla.
            cuotas_por_codigo = {}
            for r in cur.fetchall():
                cod, cid, scope, tr = r['codigo'], r['id'], r['scope'], r['id_trainer']
                cur_pick = cuotas_por_codigo.get(cod)
                if cur_pick is None:
                    cuotas_por_codigo[cod] = (cid, scope, tr)
                elif scope == 'trainer' and cur_pick[1] != 'trainer':
                    cuotas_por_codigo[cod] = (cid, scope, tr)
    except Exception as e:
        log.exception('cargar_catalogos')
        return jsonify({'ok': False, 'error': str(e)}), 500

    # ── Validación all-or-nothing ──
    cambios = []          # lista de dicts con {cli, set_cat, set_cuota, set_per}
    errores = []
    cli_visto = set()     # evita duplicar si el mismo id aparece en OK e INCOHERENCIAS
    for origen, cli_col_name, header_idx, rows in fuentes:
        col_cli = header_idx[cli_col_name]
        col_cat = header_idx['Nueva categoría']
        col_cuo = header_idx['Nueva cuota']
        col_per = header_idx['Nueva periodicidad']
        for n, row in enumerate(rows, 2):   # n = nº fila Excel (1=cabecera)
            cli = _norm(row[col_cli])
            if not cli:
                continue
            nueva_cat = _norm(row[col_cat]) if col_cat < len(row) else ''
            nueva_cuo = _norm(row[col_cuo]) if col_cuo < len(row) else ''
            nueva_per = _norm(row[col_per]) if col_per < len(row) else ''

            if not (nueva_cat or nueva_cuo or nueva_per):
                continue   # fila sin cambios

            if cli in cli_visto:
                # Mismo cliente con cambios en dos hojas → preferimos la
                # primera (OK) y avisamos.
                errores.append(f'{origen} fila {n}: cliente {cli} ya tenía '
                               f'cambios en otra hoja; deja solo una entrada.')
                continue
            cli_visto.add(cli)

            cambio = {'cli': cli, 'fila': n, 'origen': origen,
                      'set_cat': None, 'set_cuota': None, 'set_per': None}

            if nueva_cat:
                cat_id = categorias_map.get(nueva_cat)
                if not cat_id:
                    errores.append(f'{origen} fila {n} ({cli}): categoría '
                                   f'"{nueva_cat}" no existe o no está activa.')
                    continue
                cambio['set_cat'] = (nueva_cat, cat_id)

            if nueva_cuo:
                pick = cuotas_por_codigo.get(nueva_cuo)
                if not pick:
                    errores.append(f'{origen} fila {n} ({cli}): cuota '
                                   f'"{nueva_cuo}" no existe en el catálogo '
                                   f'del manager.')
                    continue
                cambio['set_cuota'] = (nueva_cuo, pick[0])  # (codigo, id)

            if nueva_per:
                if nueva_per.lower() not in PERIODICIDADES_VALIDAS:
                    errores.append(
                        f'{origen} fila {n} ({cli}): periodicidad "{nueva_per}" '
                        f'inválida (válidas: {", ".join(PERIODICIDADES_VALIDAS)}).')
                    continue
                cambio['set_per'] = nueva_per.lower()

            cambios.append(cambio)

    if errores:
        return jsonify({'ok': False, 'error': 'validacion_fallida',
                        'errores': errores,
                        'msg': 'No se aplicó NINGÚN cambio. Corrige el '
                               'Excel y vuelve a subirlo.'}), 422

    if not cambios:
        return jsonify({'ok': True, 'aplicados': 0,
                        'msg': 'El Excel no contiene cambios pendientes.'})

    # ── Aplicación (fase 1: BD local; fase 2: Odoo) ──
    aplicados_cat = 0
    aplicados_sub = 0
    fallos = []      # cambios parciales aplicados pero algún paso falló

    try:
        with get_conn() as conn:
            conn.autocommit = False
            try:
                with conn.cursor() as cur:
                    for ch in cambios:
                        cli = ch['cli']
                        if ch['set_cat']:
                            cat_nombre, cat_id = ch['set_cat']
                            cur.execute("""
                                INSERT INTO cliente_categoria
                                  (id_manager, cliente_idnoofit, categoria_id, updated_at)
                                VALUES (%s, %s, %s, NOW())
                                ON CONFLICT (id_manager, cliente_idnoofit)
                                DO UPDATE SET categoria_id=EXCLUDED.categoria_id,
                                              updated_at=NOW()
                            """, (id_manager, cli, cat_id))
                            aplicados_cat += 1
                conn.commit()
            except Exception:
                conn.rollback()
                raise
    except Exception as e:
        log.exception('aplicar_categorias')
        return jsonify({'ok': False, 'error': 'fallo_categorias',
                        'msg': str(e)}), 500

    # ── Cambios de cuota/periodicidad en Odoo ──
    # Si Odoo falla en mitad, lo registramos pero NO deshacemos categorías
    # (no es transaccional cross-system). Devolvemos detalle.
    necesita_odoo = any(ch['set_cuota'] or ch['set_per'] for ch in cambios)
    if necesita_odoo:
        try:
            o = _odoo()
            company_id = _company_id()
            today_iso = str(dt.date.today())
            for ch in cambios:
                if not (ch['set_cuota'] or ch['set_per']):
                    continue
                cli = ch['cli']
                try:
                    # Localizar partner por id_noofit
                    pids = o._call('res.partner', 'search',
                        [('id_noofit', '=', str(cli))], limit=1)
                    if not pids:
                        fallos.append(f'{cli}: partner Odoo no encontrado.')
                        continue
                    pid = pids[0]
                    # Buscar subscription activa
                    subs = o._call('round.subscription', 'search_read',
                        [('partner_id', '=', pid),
                         ('company_id', '=', company_id),
                         ('estado', '=', 'activa')],
                        ['id', 'cuota_id', 'periodicidad', 'forma_pago',
                         'descuentos_activos_ids'],
                        order='fecha_inicio DESC', limit=1)
                    if not subs:
                        fallos.append(f'{cli}: sin subscription activa en Odoo.')
                        continue
                    old = subs[0]
                    sid_old = old['id']
                    new_cuota_id = (ch['set_cuota'][1] if ch['set_cuota']
                                    else (old['cuota_id'][0] if old.get('cuota_id') else None))
                    new_per = ch['set_per'] or old.get('periodicidad') or 'mensual'
                    if not new_cuota_id:
                        fallos.append(f'{cli}: subscripción sin cuota_id (no se puede reemplazar).')
                        continue
                    # Cerrar la actual + crear la nueva
                    o._call('round.subscription', 'write', [sid_old],
                        {'fecha_fin': today_iso, 'estado': 'cancelada'})
                    new_vals = {
                        'partner_id': pid,
                        'cuota_id': int(new_cuota_id),
                        'periodicidad': new_per,
                        'forma_pago': old.get('forma_pago') or 'sepa',
                        'fecha_inicio': today_iso,
                        'estado': 'activa',
                        'company_id': company_id,
                    }
                    if old.get('descuentos_activos_ids'):
                        new_vals['descuentos_activos_ids'] = [
                            (6, 0, list(old['descuentos_activos_ids']))]
                    new_sid = o._call('round.subscription', 'create', new_vals)
                    aplicados_sub += 1
                    log_action(actor_from_request(),
                               entidad='subscription', entidad_id=sid_old,
                               accion='replace_por_excel',
                               resumen=(f'Reemplazo desde Excel pre-emisión {mes} '
                                        f'(cliente {cli}) — old={sid_old} new={new_sid}'),
                               cambios={'cli': cli, 'mes': mes,
                                        'old_id': sid_old, 'new_id': new_sid,
                                        'set_cuota': ch['set_cuota'],
                                        'set_per': ch['set_per']})
                except Exception as e:
                    log.warning(f'replace cli={cli}: {e}')
                    fallos.append(f'{cli}: {e}')
        except Exception as e:
            log.exception('aplicar_subscriptions')
            return jsonify({
                'ok': False,
                'error': 'odoo_error_global',
                'msg': str(e),
                'aplicados_categoria': aplicados_cat,
                'aplicados_subscription': aplicados_sub,
                'fallos': fallos,
            }), 500

    # Auditoría global
    log_action(actor_from_request(),
               entidad='preemision', entidad_id=mes,
               accion='aplicar_correcciones_excel',
               resumen=(f'Aplicadas {aplicados_cat} categoría(s) + '
                        f'{aplicados_sub} subscription(s) — pre-emisión {mes}'),
               cambios={'mes': mes,
                        'aplicados_categoria': aplicados_cat,
                        'aplicados_subscription': aplicados_sub,
                        'fallos': fallos})

    return jsonify({
        'ok': True,
        'mes': mes,
        'aplicados_categoria': aplicados_cat,
        'aplicados_subscription': aplicados_sub,
        'fallos': fallos,
        'msg': ('Cambios aplicados. Re-validando para mostrar el nuevo Excel...'
                if not fallos else
                'Algunos cambios no se pudieron aplicar (ver fallos).'),
    })
