"""Wizard de facturación trimestral.

Endpoints:
  GET  /api/cuotas/facturacion-trimestre/<YYYY-Tn>            preview
  GET  /api/cuotas/facturacion-trimestre/<YYYY-Tn>/excel      Excel descarga
  POST /api/cuotas/facturacion-trimestre/<YYYY-Tn>/facturar   crea facturas
       body = {recibo_ids: [...]}
"""
import datetime as dt
import logging
from io import BytesIO
from collections import defaultdict
from flask import Blueprint, request, jsonify, g, send_file

from ..auth import auth_required
from ..odoo_guard import require_feature
from ..db import get_conn
from ..audit_log import log_action, actor_from_request

bp = Blueprint('facturacion_trimestre', __name__)
log = logging.getLogger(__name__)


def _odoo(id_manager=None):
    from ..odoo_alta import OdooAlta
    o = OdooAlta(id_manager=id_manager); o._connect()
    return o


def _company_id():
    from .. import config as appconfig
    return getattr(appconfig, 'ODOO_COMPANY', 3) or 3


def _trimestre_a_meses(trim):
    """trim='2026-T2' → ['2026-04','2026-05','2026-06']"""
    y, t = trim.split('-T')
    y = int(y); t = int(t)
    base = (t - 1) * 3 + 1
    return [f'{y}-{m:02d}' for m in (base, base + 1, base + 2)]


@bp.route('/<trim>', methods=['GET'])
@auth_required
@require_feature('cuotas')
def preview(trim):
    """Lista los recibos del trimestre (cobrados pendientes de facturar +
    los ya facturados a título informativo)."""
    meses = _trimestre_a_meses(trim)
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT id, cliente_idnoofit, cliente_nombre, cuota_codigo,
                   importe_base, importe_iva, importe_total,
                   metodo_pago, estado, periodo, fecha_emision, fecha_pago,
                   account_move_id, account_move_ref
              FROM recibo
             WHERE id_manager=%s AND periodo = ANY(%s)
               AND estado IN ('pagado', 'facturado', 'impagado', 'devuelto')
             ORDER BY estado, periodo, cliente_nombre
        """, (str(g.id_manager), meses))
        rows = cur.fetchall()

    # Stats
    pagados_pte = [r for r in rows if r['estado'] == 'pagado' and not r['account_move_id']]
    facturados = [r for r in rows if r['estado'] == 'facturado']
    impagados = [r for r in rows if r['estado'] == 'impagado']

    total_pte = sum(float(r['importe_total'] or 0) for r in pagados_pte)
    total_fac = sum(float(r['importe_total'] or 0) for r in facturados)
    total_imp = sum(float(r['importe_total'] or 0) for r in impagados)

    return jsonify({
        'ok': True,
        'trimestre': trim,
        'meses': meses,
        'pagados_pendientes_facturar': len(pagados_pte),
        'ya_facturados': len(facturados),
        'impagados': len(impagados),
        'importe_pendiente': total_pte,
        'importe_facturado': total_fac,
        'importe_impagado': total_imp,
        'recibos': rows,
    })


@bp.route('/<trim>/excel', methods=['GET'])
@auth_required
@require_feature('cuotas')
def preview_excel(trim):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    meses = _trimestre_a_meses(trim)
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT id, cliente_idnoofit, cliente_nombre, cuota_codigo,
                   importe_base, importe_iva, importe_total,
                   metodo_pago, estado, periodo, fecha_emision, fecha_pago,
                   account_move_id, account_move_ref
              FROM recibo
             WHERE id_manager=%s AND periodo = ANY(%s)
             ORDER BY estado, periodo, cliente_nombre
        """, (str(g.id_manager), meses))
        rows = cur.fetchall()

    wb = Workbook(); ws = wb.active; ws.title = f'Trimestre {trim}'
    headers = [('id', 8), ('Cliente', 30), ('idnoofit', 11), ('Cuota', 30),
               ('Periodo', 9), ('Método', 10), ('Estado', 12),
               ('Base', 10), ('IVA', 10), ('Total', 10),
               ('Fecha pago', 12), ('Factura Odoo', 14), ('Marcar (X = facturar)', 18)]
    fill_h = PatternFill('solid', fgColor='2DD4A8')
    font_h = Font(bold=True, color='FFFFFF', size=11)
    border = Border(*[Side(style='thin', color='CCCCCC')]*4)
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.fill = fill_h; c.font = font_h
        c.alignment = Alignment(horizontal='center'); c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'

    # Resaltar la columna "Marcar"
    fill_marca = PatternFill('solid', fgColor='FFF7CC')
    for i, r in enumerate(rows, 2):
        vals = [
            r['id'], r['cliente_nombre'] or '', r['cliente_idnoofit'],
            r['cuota_codigo'] or '', r['periodo'], r['metodo_pago'], r['estado'],
            float(r['importe_base'] or 0), float(r['importe_iva'] or 0), float(r['importe_total'] or 0),
            str(r['fecha_pago'])[:10] if r['fecha_pago'] else '',
            r['account_move_ref'] or '',
            '',
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v); c.border = border
            if j == 13: c.fill = fill_marca

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'facturacion_trimestre_{trim}.xlsx')


@bp.route('/<trim>/facturar', methods=['POST'])
@auth_required
@require_feature('cuotas')
def facturar(trim):
    """Crea account.move out_invoice agrupando por cliente.
    body = {recibo_ids: [...], agrupar_por_cliente: bool=True}
    """
    d = request.get_json() or {}
    recibo_ids = d.get('recibo_ids') or []
    agrupar = d.get('agrupar_por_cliente', True)
    if not recibo_ids:
        return jsonify({'ok': False, 'error': 'recibo_ids_required'}), 400

    o = _odoo(g.id_manager)
    company_id = o.company_id      # company del MANAGER (no el env fijo)

    # Buscar tax 21% S
    tax_ids = o._call('account.tax', 'search',
        [('company_id', '=', company_id), ('amount', '=', 21.0),
         ('type_tax_use', '=', 'sale'), ('name', 'like', '21% S')], limit=1)
    if not tax_ids:
        tax_ids = o._call('account.tax', 'search',
            [('company_id', '=', company_id), ('amount', '=', 21.0),
             ('type_tax_use', '=', 'sale')], limit=1)
    tax_id = tax_ids[0] if tax_ids else None

    # La analítica se resuelve por el TRAINER de cada cliente dentro del bucle
    # (antes estaba hardcodeada a "Round Málaga Centro" → mezclaba Añoreta).

    # Sale journal
    sale_ids = o._call('account.journal', 'search',
        [('company_id', '=', company_id), ('type', '=', 'sale')], limit=1)
    sale_id = sale_ids[0] if sale_ids else None

    # Cargar recibos
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT id, cliente_idnoofit, cliente_nombre, cuota_codigo,
                   importe_base, importe_iva, importe_total, periodo,
                   estado, account_move_id, account_payment_id
              FROM recibo
             WHERE id_manager=%s AND id = ANY(%s)
               AND estado IN ('pagado','impagado','devuelto')
               AND account_move_id IS NULL
        """, (str(g.id_manager), recibo_ids))
        recibos = cur.fetchall()

    if not recibos:
        return jsonify({'ok': False, 'error': 'no_facturables (todos ya tienen asiento)'}), 400

    # Agrupar
    if agrupar:
        por_cliente = defaultdict(list)
        for r in recibos:
            por_cliente[r['cliente_idnoofit']].append(r)
    else:
        por_cliente = {f'{r["cliente_idnoofit"]}-{r["id"]}': [r] for r in recibos}

    facturas_creadas = 0
    errores = 0
    detalle = []

    for cli, lista_recibos in por_cliente.items():
        # Resolver partner Odoo (uno por agrupación)
        cli_idnoofit = lista_recibos[0]['cliente_idnoofit']
        partner_ids = o._call('res.partner', 'search',
            [('id_noofit', '=', str(cli_idnoofit))], limit=1)
        if not partner_ids:
            errores += 1
            log.warning(f'partner no encontrado: {cli_idnoofit}')
            continue
        pid = partner_ids[0]
        # Analítica del TRAINER del cliente (no hardcode) → cada factura con la
        # analítica de su propio centro, sin mezclar Añoreta/Málaga.
        analytic_id = o._resolve_analytic_for_partner(pid)

        # Líneas de factura: 1 por recibo
        line_ids = []
        ref_recibos = []
        for r in lista_recibos:
            line = {
                'name': f'{r["cuota_codigo"] or "Cuota"} · {r["periodo"]}',
                'quantity': 1,
                'price_unit': float(r['importe_base']),  # base sin IVA
            }
            if tax_id: line['tax_ids'] = [(6, 0, [tax_id])]
            if analytic_id: line['analytic_distribution'] = {str(analytic_id): 100.0}
            line_ids.append((0, 0, line))
            ref_recibos.append(str(r['id']))

        invoice_vals = {
            'partner_id': pid,
            'move_type': 'out_invoice',
            'ref': f'TRIM-{trim}-cli{cli_idnoofit}',
            'invoice_date': str(dt.date.today()),
            'company_id': company_id,
            'narration': f'Factura agrupada trimestre {trim} · Recibos: {",".join(ref_recibos)}',
            'invoice_line_ids': line_ids,
        }
        if sale_id: invoice_vals['journal_id'] = sale_id

        try:
            inv_id = o._call('account.move', 'create', invoice_vals)
            o._call('account.move', 'action_post', [inv_id])
            # C3 (junio 2026) — netting "cuando se facture": estos recibos ya
            # están pagados (payment a cuenta del cliente creado por
            # marcar_pagado). Al emitir la factura, reconciliamos cada payment
            # contra ella para que la factura quede pagada y el crédito a
            # cuenta del cliente se netee. Un fallo de reconcile NO aborta la
            # facturación (la factura ya existe; se puede reconciliar luego).
            reconciliados = 0
            try:
                from ..odoo_pos_sync import _reconcile
                for rr in lista_recibos:
                    pay_id = rr.get('account_payment_id')
                    if not pay_id:
                        continue
                    try:
                        _reconcile(o, inv_id, pay_id, company_id)
                        reconciliados += 1
                    except Exception as e:
                        log.warning(f'facturacion_trimestre: reconcile payment '
                                    f'{pay_id} (recibo {rr["id"]}) ↔ factura '
                                    f'{inv_id} falló: {e}')
            except Exception:
                log.exception('facturacion_trimestre: import/loop reconcile')
            # Vincular recibos
            inv_data = o._call('account.move', 'read', [inv_id], ['name'])[0]
            with get_conn() as conn, conn.cursor() as cur:
                # Marca el asiento en TODOS; estado='facturado' solo si estaba
                # 'pagado' (los impagados/devueltos conservan su estado para no
                # romper el seguimiento de impago/recobro, pero ya con asiento).
                cur.execute("""
                    UPDATE recibo
                       SET account_move_id=%s, account_move_ref=%s,
                           fecha_facturacion=NOW(),
                           estado = CASE WHEN estado='pagado' THEN 'facturado'
                                         ELSE estado END
                     WHERE id_manager=%s AND id = ANY(%s)
                """, (inv_id, inv_data['name'], str(g.id_manager),
                      [r['id'] for r in lista_recibos]))
            facturas_creadas += 1
            detalle.append({
                'invoice_id': inv_id, 'name': inv_data['name'],
                'cliente_idnoofit': cli_idnoofit,
                'recibos': len(lista_recibos),
                'pagos_reconciliados': reconciliados,
                'importe_total': sum(float(r['importe_total'] or 0) for r in lista_recibos),
            })
        except Exception as e:
            errores += 1
            log.exception(f'facturar agrup {cli}')

    log_action(actor_from_request(), entidad='facturacion_trimestre',
               entidad_id=trim, accion='facturar',
               resumen=f'{facturas_creadas} facturas · {errores} errores')

    return jsonify({
        'ok': True, 'trimestre': trim,
        'facturas_creadas': facturas_creadas, 'errores': errores,
        'detalle': detalle,
    })
