"""Validación previa antes de emitir recibos del mes.

Comprueba coherencia entre:
  - round.subscription activa en Odoo (cuota asignada)
  - cuota del cliente Round (categoría + cuota)
  - forma_pago_cliente activa
  - cliente NoofitPro (enabled=true)
  - importe esperado vs precio del catálogo cuota

Devuelve:
  - coherentes:    [cliente]
  - incoherencias: [{tipo, detalle, cliente, propuesta}]

Tipos de incoherencia:
  - sin_subscripcion         → cliente con categoría que tiene_cuota=true sin sub Odoo
  - sin_forma_pago           → cliente con sub activa pero sin forma de pago
  - cliente_inactivo_noofit  → cliente NF enabled=false con sub activa
  - importe_inconsistente    → precio sub no coincide con catálogo
  - varias_subs_misma_cuota  → bug: 2 subs activas para misma cuota
  - cliente_sin_sub          → cliente con categoría que tiene_cuota=true sin sub Odoo
  - fp_sin_sub               → cliente con forma_pago activa pero sin sub Odoo
                               (alta abortada — IBAN guardado pero cuota nunca creada)
  - sub_sin_categoria        → sub Odoo activa pero cliente sin categoría en BD
                               (cobra a oscuras — el banner de nuevos no lo marca)
"""
import logging
from io import BytesIO
from collections import defaultdict
from flask import Blueprint, request, jsonify, g, send_file

from ..auth import auth_required
from ..db import get_conn

bp = Blueprint('preemision_validar', __name__)
log = logging.getLogger(__name__)


# Etiquetas amigables para los códigos de incoherencia. Se usan en el Excel
# para que el gestor entienda cada caso sin tener que conocer el código
# técnico. Si añades un nuevo tipo en la validación, añade su etiqueta aquí.
TIPO_LABELS = {
    'sub_no_toca_este_mes':
        'Cuota no toca este mes (trimestral/anual fuera de ciclo)',
    'baja_programada_efectiva':
        'Baja programada en vigor (cliente dado de baja este mes)',
    'inactivo_temporal':
        'En pausa / baja temporal (no se cobra durante la pausa)',
    'fp_sin_sub':
        'Forma de pago activa pero sin suscripción Odoo',
    'cliente_inactivo_nf':
        'Cliente archivado en NoofitPro (inactivo)',
    'cliente_desvinculado':
        'Cliente desvinculado del centro (no aparece en NoofitPro)',
    'cliente_inactivo_odoo':
        'Cliente Odoo desactivado',
    'cliente_sin_sub':
        'Cliente con categoría de pago pero sin suscripción activa',
    'sub_sin_categoria':
        'Suscripción sin categoría asignada al cliente',
    'sub_sin_cuota':
        'Suscripción sin cuota asignada',
    'varias_subs_misma_cuota':
        'Tiene varias suscripciones de la misma cuota (duplicado)',
    'cat_sin_cuota_con_sub':
        'Categoría no debería tener cuota pero tiene suscripción activa',
    'ya_cubierto_post_mes':
        'Ya tiene pago vigente hasta un mes posterior (no se le emite)',
    'recibo_ya_existe_mes':
        'Ya tiene un recibo emitido este mes (no se duplica)',
    'sin_forma_pago':
        'Sin forma de pago configurada',
    'sepa_sin_iban':
        'Forma de pago SEPA sin IBAN registrado',
    'sepa_iban_invalido':
        'IBAN del cliente matemáticamente inválido (banco rechazaría)',
    'tarjeta_sin_token':
        'Forma de pago tarjeta sin token registrado',
    'fecha_fin_pasada_activa':
        'Suscripción activa pero con fecha fin en el pasado',
    'importe_invalido':
        'Cuota sin precio configurado para la periodicidad',
    'importe_inconsistente':
        'Importe de la cuota inconsistente con la configuración',
}


def _label_tipo(t):
    """Devuelve la etiqueta amigable, cayendo al código si no está mapeado."""
    return TIPO_LABELS.get(t, t)


# Periodicidades válidas para los dropdowns del Excel y para la validación
# del endpoint de aplicar correcciones. Coincide con config.PERIODICIDADES.
PERIODICIDADES_VALIDAS = ['mensual', 'bimensual', 'trimestral', 'semestral', 'anual']


def _ultimo_recibo_pagado_por_cliente(id_manager):
    """Devuelve {cliente_idnoofit: {'id', 'fecha_desde', 'fecha_hasta'}}
    con el último recibo `pagado` de cada cliente del manager. Un solo
    query (DISTINCT ON) para no degradar la generación del Excel."""
    try:
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT ON (cliente_idnoofit)
                       cliente_idnoofit, id, fecha_desde, fecha_hasta,
                       importe_total, cuota_codigo, cuota_descripcion
                FROM recibo
                WHERE id_manager = %s AND estado = 'pagado'
                ORDER BY cliente_idnoofit,
                         COALESCE(fecha_desde, '1970-01-01'::date) DESC,
                         id DESC
            """, (str(id_manager),))
            out = {}
            # get_conn() usa row_factory=dict_row → cada fila es dict.
            for row in cur.fetchall():
                cid = row['cliente_idnoofit']
                # Concepto: descripción si existe, si no el código de cuota.
                concepto = (row.get('cuota_descripcion')
                            or row.get('cuota_codigo') or '')
                out[str(cid)] = {
                    'id':            row['id'],
                    'fecha_desde':   row['fecha_desde'],
                    'fecha_hasta':   row['fecha_hasta'],
                    'importe_total': row.get('importe_total'),
                    'concepto':      concepto,
                }
            log.info(f'_ultimo_recibo_pagado_por_cliente: {len(out)} clientes')
            return out
    except Exception:
        log.exception('_ultimo_recibo_pagado_por_cliente')
        return {}


def _categorias_nombres(id_manager):
    """Lista de nombres de categorías activas del manager (para dropdown)."""
    try:
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                SELECT nombre FROM categoria
                WHERE id_manager = %s AND activa = TRUE
                ORDER BY nombre
            """, (str(id_manager),))
            out = [r['nombre'] for r in cur.fetchall()]
            log.info(f'_categorias_nombres: {len(out)} categorías')
            return out
    except Exception:
        log.exception('_categorias_nombres')
        return []


def _cuotas_codigos(id_manager, id_trainer=None):
    """Códigos de cuota visibles para el manager / trainer (para dropdown).
    Incluye plantillas del manager + cuotas específicas del trainer si
    se especifica."""
    try:
        with get_conn() as conn, conn.cursor() as cur:
            if id_trainer:
                cur.execute("""
                    SELECT DISTINCT codigo FROM cuota
                    WHERE id_manager = %s
                      AND (scope = 'plantilla_manager'
                           OR (scope = 'trainer' AND id_trainer = %s))
                    ORDER BY codigo
                """, (str(id_manager), str(id_trainer)))
            else:
                cur.execute("""
                    SELECT DISTINCT codigo FROM cuota
                    WHERE id_manager = %s
                    ORDER BY codigo
                """, (str(id_manager),))
            out = [r['codigo'] for r in cur.fetchall()]
            log.info(f'_cuotas_codigos: {len(out)} cuotas')
            return out
    except Exception:
        log.exception('_cuotas_codigos')
        return []


def _odoo():
    from ..odoo_alta import OdooAlta
    o = OdooAlta(); o._connect()
    return o


def _company_id():
    from .. import config as appconfig
    return getattr(appconfig, 'ODOO_COMPANY', 3) or 3


def _validar_emision(id_manager, mes, id_trainer=None):
    # Antes de validar, recalculamos los descuentos automáticos del manager
    # (familiares + varias_cuotas) para que la pre-emisión refleje el estado
    # actualizado al instante (cualquier cliente que haya cambiado de cuota,
    # entrado/salido de familia, etc. desde el último cron de las 03:15 será
    # detectado AHORA). Se ejecuta antes del bloque real de validación.
    try:
        from ..cron_descuentos_auto import recalcular_descuentos_auto
        recalcular_descuentos_auto(id_manager)
    except Exception as _e:
        log.warning(f'recalcular_descuentos_auto pre-validación: {_e}')
    return _validar_emision_inner(id_manager, mes, id_trainer)


def _validar_emision_inner(id_manager, mes, id_trainer=None):
    """Devuelve (coherentes, incoherencias). NO escribe nada.

    Casos detectados:
      - varias_subs_misma_cuota   2+ subs activas misma cuota
      - sin_forma_pago            sub activa pero sin forma_pago en BD
      - sepa_sin_iban             forma_pago=sepa pero IBAN vacío
      - sepa_iban_invalido        IBAN no pasa mod97 o DC español (mayo 2026)
      - tarjeta_sin_token         forma_pago=tarjeta_token sin card_token
      - importe_invalido          cuota sin precio para periodicidad
      - sub_sin_cuota             sub sin cuota_id asignada
      - fecha_fin_pasada_activa   sub estado=activa con fecha_fin pasada
      - cliente_inactivo_odoo     partner.active=False con sub activa
      - cliente_sin_sub           cliente con categoría tiene_cuota=true sin sub Odoo
      - fp_sin_sub                forma_pago activa pero sin sub Odoo (mayo 2026)
      - sub_sin_categoria         sub Odoo pero sin categoría BD (mayo 2026)
    """
    import datetime as dt
    o = _odoo()
    company_id = _company_id()
    hoy = dt.date.today()

    subs = o._call('round.subscription', 'search_read',
        [('estado', '=', 'activa'), ('company_id', '=', company_id)],
        ['id', 'partner_id', 'cuota_id', 'periodicidad', 'forma_pago',
         'fecha_inicio', 'fecha_fin'])

    cuotas = o._call('round.cuota.catalogo', 'search_read', [],
        ['id', 'codigo', 'descripcion',
         'precio_mensual', 'precio_trimestral', 'precio_semestral', 'precio_anual'])
    cuotas_by_id = {c['id']: c for c in cuotas}

    partner_ids = list(set(s['partner_id'][0] for s in subs if s.get('partner_id')))
    partners = o._call('res.partner', 'read', partner_ids,
        ['id', 'name', 'id_noofit', 'vat', 'email', 'active'])
    partners_by_id = {p['id']: p for p in partners}

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT cliente_idnoofit, forma_pago, iban, card_token, fecha_inicio
              FROM forma_pago_cliente
             WHERE id_manager = %s AND estado = 'activa'
        """, (str(id_manager),))
        fp_by_idnoofit = {r['cliente_idnoofit']: r for r in cur.fetchall()}

    # cliente_cache: estado real del cliente en NoofitPro (enabled) y trainer.
    #   enabled=False → archivado/inactivo en NF
    #   no aparece    → desvinculado del manager (NF ya no lo expone)
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT id::text AS id, id_trainer::text AS id_trainer, enabled,
                   raw_data->>'name' AS nombre, raw_data->>'surname' AS apellido
              FROM cliente_cache WHERE id_manager=%s
        """, (str(id_manager),))
        cache_rows = cur.fetchall()
    cache_idnoofit_enabled = {r['id']: bool(r['enabled']) for r in cache_rows}
    cache_idnoofit_trainer = {r['id']: r['id_trainer'] for r in cache_rows}
    cache_idnoofit_nombre  = {r['id']: r.get('nombre') for r in cache_rows}
    cache_idnoofit_apellido = {r['id']: r.get('apellido') for r in cache_rows}

    # Bajas programadas con fecha_baja <= día 1 del mes que vamos a emitir.
    # Un cliente con baja efectiva el día 1 (o antes) NO debería emitir recibo
    # del mes — independientemente de si el cron diario ya las ejecutó o no.
    target_y, target_m = map(int, mes.split('-'))
    primer_dia_mes = dt.date(target_y, target_m, 1)
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT cliente_idnoofit, fecha_baja, motivo
              FROM cliente_baja_programada
             WHERE id_manager=%s AND fecha_baja <= %s
        """, (str(id_manager), primer_dia_mes))
        bajas_pendientes_mes = {str(r['cliente_idnoofit']): r for r in cur.fetchall()}

    # Inactividad TEMPORAL (pausa) cuya ventana SOLAPA el mes a emitir → no se
    # emite cuota (regla "no cobrar ningún mes que la pausa toque"). Espejo del
    # guard de preemision_v2.generar (auditoría #26): el validador NO lo tenía,
    # así que un cliente en pausa aparecía falsamente como "a emitir" aunque la
    # emisión real lo salta. Overlap: inicio <= último día Y fin >= primer día.
    import calendar as _cal_v
    ultimo_dia_mes_v = dt.date(target_y, target_m, _cal_v.monthrange(target_y, target_m)[1])
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT cliente_idnoofit, fecha_inicio, fecha_fin, motivo
              FROM cliente_inactivo_temporal
             WHERE id_manager=%s AND estado <> 'cancelada'
               AND fecha_inicio <= %s AND fecha_fin >= %s
        """, (str(id_manager), ultimo_dia_mes_v, primer_dia_mes))
        inactivos_temporal_mes = {str(r['cliente_idnoofit']): r for r in cur.fetchall()}

    # Idempotencia: clientes que ya tienen un recibo del MISMO mes en BD
    # (de cualquier origen y estado no anulado). La emisión los salta — la
    # validación los marca como aviso para no aparecer falsamente como OK.
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT cliente_idnoofit, id, origen, estado
              FROM recibo
             WHERE id_manager=%s AND periodo=%s
               AND estado IN ('emitido','pagado','facturado','impagado')
        """, (str(id_manager), mes))
        ya_existen_mes = {str(r['cliente_idnoofit']): r for r in cur.fetchall()}

    # Filtro `_toca_emitir`: mensual emite cada mes, trimestral cada 3, etc.
    # Mirror exacto de preemision_v2._toca_emitir para evitar diferir.
    def _toca_emitir_local(sub, mes_str):
        # Jun 2026 — espejo de preemision_v2._toca_emitir: el gate real es la
        # COBERTURA (fecha_hasta), no el ciclo n%step desde fecha_inicio (que
        # era un artefacto del import masivo). Aquí solo "ha empezado"; los
        # cubiertos los marca ya_cubierto_post_mes por su fecha_hasta.
        fi = sub.get('fecha_inicio')
        if not fi: return False
        if isinstance(fi, str):
            fi = dt.datetime.fromisoformat(fi[:10]).date()
        fi_mes = fi.year * 12 + fi.month
        ty, tm = map(int, mes_str.split('-'))
        target_mes = ty * 12 + tm
        return target_mes >= fi_mes

    # Categorías de cliente (para detectar Trabajador/Invitado/Wellhub con sub)
    cat_by_idnoofit = {}
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT cc.cliente_idnoofit, cat.id AS cat_id, cat.nombre AS cat_nombre,
                   cat.tiene_cuota
              FROM cliente_categoria cc
              JOIN categoria cat ON cat.id = cc.categoria_id
             WHERE cc.id_manager = %s
        """, (str(id_manager),))
        for r in cur.fetchall():
            cat_by_idnoofit[r['cliente_idnoofit']] = {
                'id': r['cat_id'], 'nombre': r['cat_nombre'],
                'tiene_cuota': r['tiene_cuota'],
            }

    # Lookup GP: DNI normalizado → {codigo, dni, nombre}
    # Para enriquecer hoja OK con DNI y código GestPlus.
    import os, json as _json, re as _re, unicodedata as _ud
    def _norm_dni(d):
        s = ''.join(c for c in _ud.normalize('NFD', str(d or ''))
                    if _ud.category(c) != 'Mn')
        return _re.sub(r'[^A-Z0-9]', '', s.upper())
    gp_by_dni, gp_by_email = {}, {}
    try:
        gp_path = os.getenv('GP_DUMP_PATH', '/opt/round_config_api/gestplus_dump_LATEST.json')
        if not os.path.exists(gp_path):
            # Fallback al dump fechado más reciente
            import glob
            cands = sorted(glob.glob('/opt/round_config_api/gestplus_dump_*.json'))
            if cands: gp_path = cands[-1]
        if os.path.exists(gp_path):
            _gp = _json.load(open(gp_path, 'r', encoding='utf-8'))
            for c in (_gp.get('altas') or []) + (_gp.get('bajas_recientes_12m') or []):
                d = _norm_dni(c.get('dni') or c.get('dniContr'))
                if d and len(d) >= 7:
                    gp_by_dni.setdefault(d, c)
                e = (c.get('email') or '').strip().lower()
                if e and '@' in e:
                    gp_by_email.setdefault(e, c)
    except Exception as e:
        log.warning(f'load GP dump for codes: {e}')

    def _gp_codigo(partner):
        d = _norm_dni(partner.get('vat'))
        gp_c = gp_by_dni.get(d) if d and len(d) >= 7 else None
        if not gp_c:
            email = (partner.get('email') or '').strip().lower()
            if email: gp_c = gp_by_email.get(email)
        if not gp_c: return ''
        cod = gp_c.get('codigo') or ''
        # Quitar padding de ceros
        s = _re.sub(r'[^0-9]', '', str(cod))
        return str(int(s)) if s else ''

    coherentes = []
    incoherencias = []
    # Agrupamos subs por id_noofit (NO por partner_id) para evitar duplicados
    # cuando un mismo cliente NoofitPro tiene 2+ partners en Odoo (caso real
    # cuando un partner se dio de alta dos veces por error). Elegimos un
    # partner "canónico" — el de menor id — para representar al cliente.
    subs_by_partner = defaultdict(list)   # partner_id canónico → subs combinadas
    canonico_por_idn = {}                 # idnoofit → partner_id canónico
    duplicados_partners = []              # informativo
    for s in subs:
        if not s.get('partner_id'): continue
        pid = s['partner_id'][0]
        partner = partners_by_id.get(pid, {})
        idn = partner.get('id_noofit')
        if idn:
            existente = canonico_por_idn.get(idn)
            if existente is None:
                canonico_por_idn[idn] = pid
                subs_by_partner[pid].append(s)
            else:
                # Mismo cliente con otro partner — sumar subs al canónico
                # (con menor id) y registrar el duplicado para warning.
                if pid < existente:
                    # Cambiar canónico al nuevo pid (menor)
                    subs_by_partner[pid] = subs_by_partner.pop(existente) + [s]
                    canonico_por_idn[idn] = pid
                    duplicados_partners.append((idn, existente, pid))
                else:
                    subs_by_partner[existente].append(s)
                    duplicados_partners.append((idn, pid, existente))
        else:
            # Sin idnoofit — agrupar por partner_id directo
            subs_by_partner[pid].append(s)
    if duplicados_partners:
        log.info(f'preemision_validar: detectados {len(duplicados_partners)} partners '
                 f'Odoo duplicados con mismo id_noofit (consolidados al canónico)')

    # Clientes con un recibo PAGADO/emitido cuya cobertura (`fecha_hasta`)
    # LLEGA hasta este mes (fecha_hasta >= primer día del mes). Si está
    # cubierto no se emite otro. Jun 2026: boundary `>=` (no `>` sobre el
    # último día) para que la decisión sea por fin de cobertura, no por ciclo
    # de fecha_inicio. Un recibo que cubre hasta 30/06 cubre todo junio (no
    # re-emitir) pero NO julio (30/06 < 01/07 → en julio toca).
    target_y, target_m = map(int, mes.split('-'))
    primer_dia_cobertura = dt.date(target_y, target_m, 1)
    if target_m == 12:
        ultimo_dia_mes = dt.date(target_y, 12, 31)
    else:
        ultimo_dia_mes = (dt.date(target_y, target_m + 1, 1)
                          - dt.timedelta(days=1))
    ya_cubiertos_post_mes = {}    # idnoofit → fecha_hasta (más lejana)
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT cliente_idnoofit, MAX(fecha_hasta) AS hasta
              FROM recibo
             WHERE id_manager = %s
               AND estado IN ('pagado','emitido','facturado')
               AND fecha_hasta IS NOT NULL
               AND fecha_hasta >= %s
               AND importe_total > 0
             GROUP BY cliente_idnoofit
        """, (str(id_manager), primer_dia_cobertura))
        for r in cur.fetchall():
            ya_cubiertos_post_mes[r['cliente_idnoofit']] = r['hasta']

    # Imports diferidos para aplicar descuentos + modificaciones del cliente,
    # en simetría con preemision_v2.generar.
    from ..descuentos_apply import (
        calcular_precio_con_descuentos, aplicar_descuentos_familiares,
        get_descuentos_familiares_activos,
        aplicar_descuentos_varias_cuotas_auto,
        get_descuentos_varias_cuotas_activos,
    )
    from ..modificaciones_apply import (
        get_modificaciones_activas_mes, aplicar_modif_a_cuota,
        aplicar_modif_globales,
    )

    # Pre-cómputo familias + cuotas activas por cliente (para descuento familiares)
    cuotas_activas_by_idnoofit = defaultdict(set)
    for s in subs:
        pid = s['partner_id'][0] if s.get('partner_id') else None
        if not pid: continue
        partner = partners_by_id.get(pid, {})
        idn = partner.get('id_noofit')
        if not idn: continue
        cuota = cuotas_by_id.get(s['cuota_id'][0]) if s.get('cuota_id') else None
        if cuota and cuota.get('codigo'):
            cuotas_activas_by_idnoofit[idn].add(cuota['codigo'])

    familia_por_cliente = {}
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT m.cliente_idnoofit AS yo, ARRAY(
                SELECT cliente_idnoofit FROM familia_miembro mm
                 WHERE mm.familia_id = m.familia_id
            ) AS miembros
              FROM familia_miembro m
             WHERE m.id_manager = %s
        """, (str(id_manager),))
        for r in cur.fetchall():
            familia_por_cliente[r['yo']] = r.get('miembros') or []
    descuentos_familiares = get_descuentos_familiares_activos(id_manager)
    descuentos_varias = get_descuentos_varias_cuotas_activos(id_manager)

    # Helper: precio aplicado a una sub SIN descuentos ni modificaciones.
    def _precio_sub(s):
        if not s.get('cuota_id'): return 0.0
        cuota = cuotas_by_id.get(s['cuota_id'][0])
        if not cuota: return 0.0
        return float(cuota.get(f'precio_{s.get("periodicidad","mensual")}') or 0)

    # Helper: precio sub TRAS descuentos del catálogo + descuentos familiares
    # automáticos + modificaciones puntuales.
    def _precio_sub_con_ajustes(s, idnoofit, cuotas_activas_codigos, modifs_mes):
        if not s.get('cuota_id'): return 0.0
        cuota = cuotas_by_id.get(s['cuota_id'][0])
        if not cuota: return 0.0
        p_normal = float(cuota.get(f'precio_{s.get("periodicidad","mensual")}') or 0)
        # 1) Descuentos asignados al cliente (porcentaje/importe únicamente
        #    — varias_cuotas se aplica automático en el paso 2a desde may 2026)
        _tr = cache_idnoofit_trainer.get(idnoofit)   # scope por trainer (#28)
        p_tras_desc, _info_d = calcular_precio_con_descuentos(
            id_manager, idnoofit, cuota.get('codigo'), p_normal,
            cuotas_activas_codigos=cuotas_activas_codigos, id_trainer_cliente=_tr)
        # 2a) Descuento AUTOMÁTICO "dos cuotas" (sin asignación previa)
        p_tras_varias, _info_v = aplicar_descuentos_varias_cuotas_auto(
            id_manager, idnoofit, cuota.get('codigo'), p_tras_desc,
            cuotas_activas_codigos=cuotas_activas_codigos,
            descuentos_varias=descuentos_varias, id_trainer_cliente=_tr)
        # 2b) Descuento automático por familia (≥2 miembros con la cuota)
        miembros_fam = familia_por_cliente.get(idnoofit) or []
        cuotas_por_familiar = {
            idn: cuotas_activas_by_idnoofit.get(idn, set())
            for idn in miembros_fam
        }
        p_tras_fam, _info_f = aplicar_descuentos_familiares(
            id_manager, idnoofit, cuota.get('codigo'), p_tras_varias,
            cuotas_por_familiar, descuentos_familiares=descuentos_familiares,
            id_trainer_cliente=_tr)
        # 3) Modificaciones por cuota — match por código
        p_final, _info_m, _ids = aplicar_modif_a_cuota(modifs_mes, cuota['id'],
            p_tras_fam, cuota_codigo=cuota.get('codigo'))
        return p_final

    for pid, plist in subs_by_partner.items():
        partner = partners_by_id.get(pid, {})
        idnoofit = partner.get('id_noofit') or ''
        nombre = partner.get('name') or '?'
        cliente_info = {
            'partner_id': pid, 'idnoofit': idnoofit, 'nombre': nombre,
            'email': partner.get('email') or '',
            'dni': partner.get('vat') or '',
            'codigo_gp': _gp_codigo(partner),
        }
        problemas = []

        # cliente_inactivo_odoo
        if partner.get('active') is False:
            problemas.append({
                'tipo': 'cliente_inactivo_odoo',
                'detalle': 'Partner Odoo inactivo (active=False) con sub activa',
                'propuesta': 'Reactivar el partner en Odoo o cancelar las subs.',
            })

        # ya_cubierto_post_mes: cliente con un recibo pagado/emitido cuya
        # fecha_hasta cubre más allá del mes de emisión (típico trimestral
        # o anual ya cobrado). NO se le emite otro y aparece como aviso.
        cubierto_hasta = ya_cubiertos_post_mes.get(idnoofit)
        if cubierto_hasta:
            problemas.append({
                'tipo': 'ya_cubierto_post_mes',
                'detalle': (f'Tiene un recibo pagado cuya cobertura llega '
                            f'hasta {cubierto_hasta.isoformat()} '
                            f'(posterior al fin de mes {ultimo_dia_mes.isoformat()}). '
                            f'No se le emite otro este mes.'),
                'propuesta': ('Revisar: si la cobertura es correcta, no hacer nada '
                              '(el cron lo omite). Si fue un error, ajustar '
                              'la fecha_hasta del recibo previo.'),
            })

        # recibo_ya_existe_mes: el cliente ya tiene un recibo de este mes en
        # BD (emisión previa o manual). La generación lo salta → la
        # validación lo marca como aviso para que no aparezca como OK.
        rec_exist = ya_existen_mes.get(idnoofit)
        if rec_exist:
            problemas.append({
                'tipo': 'recibo_ya_existe_mes',
                'detalle': (f'Ya hay un recibo de {mes} (id {rec_exist["id"]}, '
                            f'origen {rec_exist["origen"]}, estado {rec_exist["estado"]}). '
                            f'La emisión no creará otro.'),
                'propuesta': 'Si el recibo previo es correcto, ignorar. Si no, borrarlo antes de emitir.',
            })

        # sub_no_toca_este_mes: TODAS las subs del cliente caen fuera del
        # calendario del mes (típico trimestral en mes intermedio). La
        # emisión lo salta sin crear recibo — la validación lo refleja
        # como informativo (no es un error real).
        subs_que_tocan = [s for s in plist if _toca_emitir_local(s, mes)]
        if not subs_que_tocan and plist:
            periodicidades = sorted({s.get('periodicidad') or '?' for s in plist})
            problemas.append({
                'tipo': 'sub_no_toca_este_mes',
                'detalle': (f'Ninguna de sus {len(plist)} sub(s) toca emitir en {mes} '
                            f'(periodicidades: {", ".join(periodicidades)}). '
                            f'Se cobra solo en los meses del ciclo desde fecha_inicio.'),
                'propuesta': 'Sin acción — no se emite este mes por diseño.',
            })

        # baja_programada_efectiva: el cliente tiene una baja con fecha
        # <= día 1 del mes que vamos a emitir. NO se emite (el cron diario
        # lo archivará si aún no lo hizo, pero la emisión ya lo respeta).
        baja = bajas_pendientes_mes.get(idnoofit)
        if baja:
            problemas.append({
                'tipo': 'baja_programada_efectiva',
                'detalle': (f'Tiene baja programada para {baja["fecha_baja"].isoformat()} '
                            f'(<= día 1 del mes {primer_dia_mes.isoformat()})'
                            + (f'. Motivo: {baja.get("motivo")}' if baja.get("motivo") else '')),
                'propuesta': ('No se le emite recibo. Si la baja es errónea, '
                              'cancélala desde la ficha del cliente.'),
            })

        # inactivo_temporal: pausa cuya ventana toca el mes → no se cobra (espejo
        # del guard de preemision_v2; auditoría #26).
        pausa = inactivos_temporal_mes.get(idnoofit)
        if pausa:
            problemas.append({
                'tipo': 'inactivo_temporal',
                'detalle': (f'En pausa temporal '
                            f'({pausa["fecha_inicio"].isoformat()} → {pausa["fecha_fin"].isoformat()}'
                            + (f', {pausa.get("motivo")}' if pausa.get("motivo") else '') + '). '
                            'No se emite recibo de los meses que toca la pausa.'),
                'propuesta': 'Sin acción — no se cobra durante la baja temporal.',
            })

        # cliente_inactivo_nf: enabled=False en NoofitPro (archivado).
        # Tiene sub activa en Odoo pero su cuenta NF está archivada → no emitir.
        if idnoofit and idnoofit in cache_idnoofit_enabled and not cache_idnoofit_enabled[idnoofit]:
            problemas.append({
                'tipo': 'cliente_inactivo_nf',
                'detalle': 'Cliente archivado en NoofitPro (enabled=False) con sub activa',
                'propuesta': ('Cancelar la suscripción Odoo. Si fue archivado por error, '
                              'reactivar primero el cliente en NoofitPro.'),
            })

        # cliente_desvinculado: no aparece en cliente_cache del manager.
        # NoofitPro ya no lo expone para ningún trainer → desvinculado o eliminado.
        if idnoofit and idnoofit not in cache_idnoofit_enabled:
            problemas.append({
                'tipo': 'cliente_desvinculado',
                'detalle': 'Cliente no aparece en NoofitPro para ningún trainer del manager',
                'propuesta': ('Verificar en NoofitPro. Si quedó huérfano, cancelar la sub. '
                              'Si fue desvinculado por error, re-vincularlo al trainer.'),
            })

        # cat_sin_cuota_con_sub: cliente con categoría Trabajador/Invitado/Wellhub
        # (tiene_cuota=false) pero con suscripción activa → no debería emitir recibos.
        cat = cat_by_idnoofit.get(idnoofit) or {}
        if cat and cat.get('tiene_cuota') is False:
            problemas.append({
                'tipo': 'cat_sin_cuota_con_sub',
                'detalle': (f'Categoría "{cat.get("nombre","?")}" no debería tener cuota, '
                            f'pero hay {len(plist)} sub(s) activa(s) en Odoo'),
                'propuesta': ('Cancelar las suscripciones (no le cobramos) o cambiar la '
                              'categoría a "Cliente" si es socio de pago.'),
            })

        # varias_subs_misma_cuota
        cuotas_count = defaultdict(int)
        for s in plist:
            if s.get('cuota_id'): cuotas_count[s['cuota_id'][0]] += 1
        for cid, n in cuotas_count.items():
            if n > 1:
                cuota = cuotas_by_id.get(cid, {})
                problemas.append({
                    'tipo': 'varias_subs_misma_cuota',
                    'detalle': f'{n} subs activas para cuota {cuota.get("codigo","?")}',
                    'propuesta': f'Cancelar las {n-1} subs sobrantes y dejar solo una activa.',
                })

        # sub_sin_cuota
        for s in plist:
            if not s.get('cuota_id'):
                problemas.append({
                    'tipo': 'sub_sin_cuota',
                    'detalle': f'Sub id={s["id"]} sin cuota_id asignada',
                    'propuesta': 'Editar la subscription y asignar una cuota del catálogo.',
                })

        # fecha_fin_pasada_activa
        for s in plist:
            ff = s.get('fecha_fin')
            if ff:
                if isinstance(ff, str): ff = dt.datetime.fromisoformat(ff[:10]).date()
                if ff < hoy:
                    problemas.append({
                        'tipo': 'fecha_fin_pasada_activa',
                        'detalle': f'Sub id={s["id"]} estado=activa con fecha_fin={ff.isoformat()} < hoy',
                        'propuesta': 'Cambiar estado a "cancelada" o quitar fecha_fin.',
                    })

        # forma de pago
        fp = fp_by_idnoofit.get(idnoofit)
        if not fp:
            problemas.append({
                'tipo': 'sin_forma_pago',
                'detalle': 'Cliente sin forma de pago activa en BD',
                'propuesta': 'Configurar forma de pago en perfil del cliente (pestaña Cuota y fechas).',
            })
        else:
            if fp['forma_pago'] == 'sepa' and not (fp.get('iban') or '').strip():
                problemas.append({
                    'tipo': 'sepa_sin_iban',
                    'detalle': 'forma_pago=sepa pero IBAN vacío',
                    'propuesta': 'Añadir IBAN al cliente o cambiar forma de pago a efectivo.',
                })
            elif fp['forma_pago'] == 'sepa':
                # Validar matemáticamente el IBAN (mod97 + CCC español). Los
                # IBANs mal escritos hacen que el banco rechace el fichero
                # SEPA entero, así que es crítico detectarlos antes de generar.
                from ..iban_validator import validar_iban
                v = validar_iban(fp.get('iban'))
                if not v['ok']:
                    problemas.append({
                        'tipo': 'sepa_iban_invalido',
                        'detalle': (f'IBAN matemáticamente inválido '
                                    f'({v["error"]}): {v["detalle"]}. '
                                    f'IBAN: {fp.get("iban")}'),
                        'propuesta': ('Pedir al cliente el IBAN correcto y '
                                      'actualizarlo en perfil → Cuota y fechas, '
                                      'o cambiar forma de pago a efectivo.'),
                    })
            if fp['forma_pago'] == 'tarjeta_token' and not (fp.get('card_token') or '').strip():
                problemas.append({
                    'tipo': 'tarjeta_sin_token',
                    'detalle': 'forma_pago=tarjeta_token pero sin card_token',
                    'propuesta': 'Tokenizar la tarjeta o cambiar forma de pago.',
                })

        # importe_invalido
        for s in plist:
            cuota = cuotas_by_id.get(s['cuota_id'][0]) if s.get('cuota_id') else None
            if not cuota: continue
            campo_precio = f'precio_{s["periodicidad"]}'
            precio_catalogo = float(cuota.get(campo_precio) or 0)
            if precio_catalogo <= 0:
                problemas.append({
                    'tipo': 'importe_invalido',
                    'detalle': f'Cuota {cuota.get("codigo")} sin precio para periodicidad {s["periodicidad"]}',
                    'propuesta': f'Configurar precio_{s["periodicidad"]} en catálogo cuotas.',
                })

        # importe_inconsistente: misma cuota con periodicidades distintas
        per_by_cuota = defaultdict(set)
        for s in plist:
            if s.get('cuota_id'): per_by_cuota[s['cuota_id'][0]].add(s.get('periodicidad'))
        for cid, pers in per_by_cuota.items():
            if len(pers) > 1:
                cuota = cuotas_by_id.get(cid, {})
                problemas.append({
                    'tipo': 'importe_inconsistente',
                    'detalle': f'Cuota {cuota.get("codigo")} con varias periodicidades: {sorted(pers)}',
                    'propuesta': 'Dejar solo una periodicidad activa.',
                })

        for p in problemas:
            incoherencias.append({**p, 'cliente': cliente_info})

        if not problemas:
            # Set de códigos de cuotas activas del cliente (para 'varias_cuotas')
            cuotas_activas_cli = set()
            for s in plist:
                cu = cuotas_by_id.get(s['cuota_id'][0]) if s.get('cuota_id') else None
                if cu and cu.get('codigo'):
                    cuotas_activas_cli.add(cu['codigo'])
            # Modificaciones vigentes del cliente para el mes
            modifs_mes = get_modificaciones_activas_mes(id_manager, idnoofit, mes) if idnoofit else []

            # Cuotas por miembro (para descuento familiares)
            miembros_fam = familia_por_cliente.get(idnoofit) or []
            cuotas_por_familiar = {
                idn: cuotas_activas_by_idnoofit.get(idn, set())
                for idn in miembros_fam
            }

            # Detalle por cuota: precio_normal y precio_final tras descuentos+mods
            # Importante: iterar solo las subs que TOCA emitir este mes
            # (mensual cada mes, trimestral cada 3, etc.) — espejo exacto
            # de preemision_v2.generar para que el importe coincida.
            cuotas_detalle = []
            importe_subtotal = 0.0
            for s in subs_que_tocan:
                if not s.get('cuota_id'): continue
                cuota = cuotas_by_id.get(s['cuota_id'][0])
                if not cuota: continue
                p_normal = float(cuota.get(f'precio_{s.get("periodicidad","mensual")}') or 0)
                _tr = cache_idnoofit_trainer.get(idnoofit)   # scope por trainer (#28)
                # 1) Descuentos asignados (porcentaje/importe)
                p_tras_desc, info_desc = calcular_precio_con_descuentos(
                    id_manager, idnoofit, cuota.get('codigo'), p_normal,
                    cuotas_activas_codigos=cuotas_activas_cli, id_trainer_cliente=_tr)
                # 2a) Descuento AUTOMÁTICO "dos cuotas" (sin asignación)
                p_tras_varias, info_varias = aplicar_descuentos_varias_cuotas_auto(
                    id_manager, idnoofit, cuota.get('codigo'), p_tras_desc,
                    cuotas_activas_codigos=cuotas_activas_cli,
                    descuentos_varias=descuentos_varias, id_trainer_cliente=_tr)
                # 2b) Descuento automático por familia
                p_tras_fam, info_fam = aplicar_descuentos_familiares(
                    id_manager, idnoofit, cuota.get('codigo'), p_tras_varias,
                    cuotas_por_familiar, descuentos_familiares=descuentos_familiares,
                    id_trainer_cliente=_tr)
                # 3) Modificaciones por cuota — match por código (los IDs
                #    locales no coinciden con los de Odoo)
                p_final, info_mod, _ids = aplicar_modif_a_cuota(
                    modifs_mes, cuota['id'], p_tras_fam,
                    cuota_codigo=cuota.get('codigo'))
                importe_subtotal += p_final
                desc_partes = (list(info_desc or [])
                               + list(info_varias or [])
                               + list(info_fam or []))
                cuotas_detalle.append({
                    'codigo': cuota.get('codigo') or '?',
                    'periodicidad': s.get('periodicidad'),
                    'precio_normal': round(p_normal, 2),
                    'precio_final': round(p_final, 2),
                    'descuentos': [
                        f"{x['descuento_codigo']} ({x['precio_antes']}€→{x['precio_despues']}€)"
                        for x in desc_partes
                    ],
                    # Estructurado para construir las columnas dinámicas del
                    # listado: por cada descuento, su código + importe ahorrado.
                    'descuentos_struct': [
                        {'codigo': x['descuento_codigo'],
                         'ahorro': round(float(x['precio_antes']) - float(x['precio_despues']), 2)}
                        for x in desc_partes
                    ],
                    'modificaciones': [
                        f"{m['tipo']} {m['valor']}€" + (f": {m['razon']}" if m.get('razon') else '')
                        for m in info_mod
                    ] if info_mod else [],
                    # Estructurado: el `delta` = precio_despues - precio_antes
                    # es la variación REAL que aplica la modificación al total.
                    # Para descuento/cargo_extra coincide con `valor`. Para
                    # `precio_alternativo` puede ser distinto (depende del
                    # precio que sustituye). Usamos `delta` en la suma.
                    'modificaciones_struct': [
                        {'tipo': m['tipo'], 'valor': float(m['valor'] or 0),
                         'delta': round(float(m.get('precio_despues') or 0)
                                        - float(m.get('precio_antes') or 0), 2),
                         'razon': m.get('razon') or ''}
                        for m in (info_mod or [])
                    ],
                })

            # Aplicar modificaciones globales (sin cuota_id) sobre el total
            importe_total, info_global, _ids_g = aplicar_modif_globales(
                modifs_mes, importe_subtotal)
            modif_globales_label = [
                f"{m['tipo']} {m['valor']}€" + (f": {m['razon']}" if m.get('razon') else '')
                for m in info_global
            ] if info_global else []
            modif_globales_struct = [
                {'tipo': m['tipo'], 'valor': float(m['valor'] or 0),
                 'delta': round(float(m.get('total_despues') or 0)
                                - float(m.get('total_antes') or 0), 2),
                 'razon': m.get('razon') or ''}
                for m in (info_global or [])
            ]

            forma_pago = (fp or {}).get('forma_pago', '?')
            # Periodicidad principal (la más común entre las subs del cliente)
            from collections import Counter as _Counter
            per_counts = _Counter(s.get('periodicidad') for s in subs_que_tocan if s.get('periodicidad'))
            periodicidad = per_counts.most_common(1)[0][0] if per_counts else 'mensual'
            cuotas_codigos = sorted({d['codigo'] for d in cuotas_detalle})
            coherentes.append({
                **cliente_info,    # incluye dni y codigo_gp
                'id_trainer': cache_idnoofit_trainer.get(idnoofit),
                'nombre_solo': cache_idnoofit_nombre.get(idnoofit) or '',
                'apellido':    cache_idnoofit_apellido.get(idnoofit) or '',
                'subs': len(plist),
                'cuotas': cuotas_codigos,
                'cuotas_detalle': cuotas_detalle,
                'modificaciones_globales': modif_globales_label,
                'modificaciones_globales_struct': modif_globales_struct,
                'forma_pago': forma_pago,
                'iban': (fp or {}).get('iban', ''),
                'periodicidad': periodicidad,
                'importe_total': round(importe_total, 2),
                'categoria': (cat or {}).get('nombre', ''),
            })

    # cliente_sin_sub: clientes con categoría tiene_cuota=true sin sub activa.
    # Enriquecemos con:
    #   - nombre/email desde un partner Odoo con ese id_noofit (si existe,
    #     aunque no tenga sub) — vía búsqueda batch.
    #   - fallback al `recibo.cliente_nombre` más reciente en BD.
    try:
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                SELECT cc.cliente_idnoofit, cat.nombre AS categoria_nombre
                  FROM cliente_categoria cc
                  JOIN categoria cat ON cat.id = cc.categoria_id
                 WHERE cc.id_manager=%s AND cat.tiene_cuota=TRUE AND cat.activa=TRUE
            """, (str(id_manager),))
            con_cuota_obligada = {r['cliente_idnoofit']: r['categoria_nombre'] for r in cur.fetchall()}
        idnoofits_con_sub = {partners_by_id[pid].get('id_noofit')
                             for pid in subs_by_partner if partners_by_id.get(pid, {}).get('id_noofit')}
        sin_sub_idnoofits = [idn for idn in con_cuota_obligada
                             if idn not in idnoofits_con_sub]

        # 1) Buscar partners Odoo (sin filtrar por sub) por id_noofit
        partner_lookup = {}    # idnoofit → {partner_id, name, email, active}
        if sin_sub_idnoofits:
            try:
                rows = o._call('res.partner', 'search_read',
                    [('id_noofit', 'in', sin_sub_idnoofits),
                     ('company_id', 'in', [False, company_id])],
                    ['id', 'id_noofit', 'name', 'email', 'active'])
                for r in rows:
                    if r.get('id_noofit'):
                        partner_lookup[r['id_noofit']] = r
            except Exception as e:
                log.warning(f'partner_lookup: {e}')

        # 2) Fallback: nombre desde recibo (último cliente_nombre cacheado)
        nombre_recibo = {}
        if sin_sub_idnoofits:
            with get_conn() as conn, conn.cursor() as cur:
                cur.execute("""
                    SELECT DISTINCT ON (cliente_idnoofit) cliente_idnoofit, cliente_nombre
                      FROM recibo
                     WHERE id_manager = %s AND cliente_idnoofit = ANY(%s)
                       AND cliente_nombre IS NOT NULL AND cliente_nombre <> ''
                     ORDER BY cliente_idnoofit, fecha_emision DESC
                """, (str(id_manager), sin_sub_idnoofits))
                for r in cur.fetchall():
                    nombre_recibo[r['cliente_idnoofit']] = r['cliente_nombre']

        # 3) Fallback: nombre/email desde dump NoofitPro local
        # (`/opt/round_config_api/noofit_clientes_dump.json`). Cubre clientes
        # que aún no tienen partner Odoo ni recibos pero existen en NF.
        # Además marca si están archivados (enabled=false) para sugerir baja.
        nf_lookup = {}    # idnoofit → {name, surname, email, enabled}
        try:
            import os, json as _json
            dump_path = os.getenv(
                'NOOFIT_DUMP_PATH',
                '/opt/round_config_api/noofit_clientes_dump.json',
            )
            if os.path.exists(dump_path) and sin_sub_idnoofits:
                _ids = set(sin_sub_idnoofits)
                with open(dump_path, 'r', encoding='utf-8') as f:
                    _dump = _json.load(f)
                for c in _dump.get('clientes', []):
                    idn = str(c.get('id') or '')
                    if idn in _ids:
                        nf_lookup[idn] = {
                            'name': (c.get('name') or '').strip(),
                            'surname': (c.get('surname') or '').strip(),
                            'email': (c.get('email') or '').strip(),
                            'enabled': bool(c.get('enabled')),
                        }
        except Exception as e:
            log.warning(f'nf_dump fallback: {e}')

        for idnoofit in sin_sub_idnoofits:
            cat_nombre = con_cuota_obligada[idnoofit]
            p = partner_lookup.get(idnoofit) or {}
            nf = nf_lookup.get(idnoofit) or {}
            # Prioridad: Odoo > recibo > NF dump
            nombre = (p.get('name') or '').strip()
            if not nombre and nombre_recibo.get(idnoofit):
                nombre = nombre_recibo[idnoofit]
            if not nombre and (nf.get('name') or nf.get('surname')):
                nombre = (nf.get('name', '') + ' ' + nf.get('surname', '')).strip()
            email = (p.get('email') or '').strip() or nf.get('email') or ''
            partner_id = p.get('id') or None

            # Detectar archivado en NF → sugerir baja de categoría
            propuesta = 'Asignar cuota desde el perfil del cliente.'
            detalle_extra = ''
            if nf and nf.get('enabled') is False:
                propuesta = ('Cliente archivado en NoofitPro. Quitar la categoría '
                             '"' + cat_nombre + '" o reactivarlo en NF.')
                detalle_extra = ' [archivado en NF]'
            elif p.get('active') is False:
                propuesta = ('Partner Odoo inactivo. Reactivarlo y asignar cuota, '
                             'o cambiar la categoría del cliente a una sin cuota.')

            incoherencias.append({
                'tipo': 'cliente_sin_sub',
                'detalle': (f'Categoría "{cat_nombre}" (tiene_cuota=true) '
                            f'pero sin sub activa Odoo' + detalle_extra),
                'cliente': {'partner_id': partner_id, 'idnoofit': idnoofit,
                            'nombre': nombre, 'email': email},
                'propuesta': propuesta,
            })
    except Exception as e:
        log.warning(f'cliente_sin_sub check: {e}')

    # ── fp_sin_sub: forma de pago activa pero ninguna sub Odoo activa ──────
    # Síntoma típico de alta de cliente abortada (operador rellenó IBAN/SEPA
    # pero el wizard de cuota falló a mitad). Esos clientes están a un paso
    # de cobrar pero no se les llega a emitir recibo → conviene avisar para
    # completar el alta o quitar la forma de pago.
    try:
        idnoofits_con_sub = {partners_by_id[pid].get('id_noofit')
                             for pid in subs_by_partner
                             if partners_by_id.get(pid, {}).get('id_noofit')}
        fp_sin_sub_idn = [idn for idn in fp_by_idnoofit
                          if idn and idn not in idnoofits_con_sub]
        # Nombre/email desde Odoo si existe partner; si no, cache cliente NF.
        partner_lookup_fp = {}
        nf_lookup_fp = {}
        if fp_sin_sub_idn:
            try:
                rows = o._call('res.partner', 'search_read',
                    [('id_noofit', 'in', fp_sin_sub_idn),
                     ('company_id', 'in', [False, company_id])],
                    ['id', 'id_noofit', 'name', 'email'])
                partner_lookup_fp = {r['id_noofit']: r for r in rows if r.get('id_noofit')}
            except Exception as e:
                log.warning(f'fp_sin_sub partner_lookup: {e}')
            # Fallback nombre desde cliente_cache (BD local)
            try:
                with get_conn() as conn, conn.cursor() as cur:
                    cur.execute("""SELECT id, name, surname, email, enabled
                                     FROM cliente_cache
                                    WHERE id_manager = %s AND id = ANY(%s)""",
                                (str(id_manager),
                                 [int(x) for x in fp_sin_sub_idn if str(x).isdigit()]))
                    for r in cur.fetchall():
                        nf_lookup_fp[str(r['id'])] = r
            except Exception as e:
                log.warning(f'fp_sin_sub cache lookup: {e}')
        for idn in fp_sin_sub_idn:
            fp = fp_by_idnoofit.get(idn) or {}
            cat = cat_by_idnoofit.get(idn) or {}
            p = partner_lookup_fp.get(idn) or {}
            nf = nf_lookup_fp.get(str(idn)) or {}
            nombre = (p.get('name')
                      or f"{nf.get('name','')} {nf.get('surname','')}".strip()
                      or '?')
            email = (p.get('email') or nf.get('email') or '')
            if nf.get('enabled') is False:
                propuesta = ('Cliente archivado en NoofitPro. Quitar la forma '
                             'de pago activa o reactivar al cliente.')
                extra = ' [archivado en NF]'
            elif cat and cat.get('tiene_cuota') is False:
                propuesta = (f'Categoría "{cat.get("nombre")}" no requiere cuota — '
                             'la forma de pago está colgada, conviene quitarla.')
                extra = f' [cat={cat.get("nombre")}]'
            else:
                propuesta = ('Completar la asignación de cuota desde el perfil '
                             '(Cuota y fechas → Asignar nueva cuota) o quitar la '
                             'forma de pago si el cliente ya no debe cobrar.')
                extra = ''
            incoherencias.append({
                'tipo': 'fp_sin_sub',
                'detalle': (f'Forma de pago "{fp.get("forma_pago")}" activa pero '
                            f'sin sub Odoo' + extra),
                'cliente': {'partner_id': p.get('id'), 'idnoofit': idn,
                            'nombre': nombre, 'email': email},
                'propuesta': propuesta,
            })
    except Exception as e:
        log.warning(f'fp_sin_sub check: {e}')

    # ── sub_sin_categoria: sub Odoo activa pero sin categoría asignada en BD ─
    # Cobra pero el banner "Nuevos clientes esperando cobro" no lo detecta
    # como ya atendido (porque la señal canónica es "tiene categoría"). Cuotas
    # generadas a oscuras. Indicio de cliente creado vía import o sub creada
    # manualmente en Odoo sin pasar por el wizard de alta.
    try:
        sub_sin_cat_pids = [pid for pid in subs_by_partner
                            if partners_by_id.get(pid, {}).get('id_noofit')
                            and partners_by_id[pid]['id_noofit'] not in cat_by_idnoofit]
        for pid in sub_sin_cat_pids:
            partner = partners_by_id.get(pid, {})
            idn = partner.get('id_noofit')
            cuotas_act = sorted(cuotas_activas_by_idnoofit.get(idn, set())) or ['?']
            incoherencias.append({
                'tipo': 'sub_sin_categoria',
                'detalle': (f'Sub activa ({", ".join(cuotas_act)}) pero sin '
                            'categoría asignada en BD'),
                'cliente': {
                    'partner_id': pid, 'idnoofit': idn,
                    'nombre': partner.get('name') or '?',
                    'email': partner.get('email') or '',
                },
                'propuesta': ('Asignar categoría desde el perfil del cliente '
                              '(Datos personales → Categoría). Mientras no la '
                              'tenga, el banner de "Nuevos clientes" lo seguirá '
                              'mostrando como pendiente.'),
            })
    except Exception as e:
        log.warning(f'sub_sin_categoria check: {e}')

    # ─── Recibos manuales en borrador para esta remesa ───────────────────
    # Los borradores creados desde "Recibos manuales" (estado='borrador_remesa')
    # NO vienen de subscripciones Odoo. Hay que añadirlos manualmente a los
    # coherentes para que aparezcan en el Excel/validación junto a los
    # auto-generados. Cada borrador → un coherente con `_manual=True`.
    try:
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("""
                SELECT id, cliente_idnoofit, cliente_nombre,
                       id_trainer::text AS id_trainer,
                       cuota_codigo, cuota_descripcion, periodicidad,
                       importe_total, metodo_pago, fecha_emision, notas
                  FROM recibo
                 WHERE id_manager = %s AND periodo = %s
                   AND estado = 'borrador_remesa'
                 ORDER BY created_at ASC
            """, (str(id_manager), mes))
            borradores = cur.fetchall()
    except Exception as e:
        log.warning(f'manuales_borrador query: {e}')
        borradores = []

    METODO_A_FORMA_PAGO = {
        'sepa': 'sepa',
        'tarjeta_tok': 'tarjeta_token',
        'caja_efectivo': 'efectivo',
        'caja_tpv_fisico': 'tpv',
        'caja_tpv_virtual': 'tpv',
        'enlace_pago': 'enlace_pago',
    }

    for r in borradores:
        idnoofit = str(r['cliente_idnoofit'] or '')
        partner_iban = (fp_by_idnoofit.get(idnoofit) or {}).get('iban', '')
        # Separar nombre/apellido si viene como "Nombre Apellido…"
        full = (r['cliente_nombre'] or '').strip()
        cache_nom = cache_idnoofit_nombre.get(idnoofit) or ''
        cache_ape = cache_idnoofit_apellido.get(idnoofit) or ''
        if not cache_nom and not cache_ape and full:
            parts = full.split(' ', 1)
            cache_nom = parts[0]
            cache_ape = parts[1] if len(parts) > 1 else ''
        coherentes.append({
            'partner_id': None,
            'idnoofit': idnoofit,
            'nombre': full or f'{cache_nom} {cache_ape}'.strip(),
            'email': '',
            'dni': '',
            'codigo_gp': '',
            'id_trainer': r.get('id_trainer') or cache_idnoofit_trainer.get(idnoofit),
            'nombre_solo': cache_nom,
            'apellido':    cache_ape,
            'subs': 0,
            'cuotas': [r['cuota_codigo']] if r.get('cuota_codigo') else ['(manual)'],
            'cuotas_detalle': [{
                'codigo': r.get('cuota_codigo') or '(manual)',
                'descripcion': r.get('cuota_descripcion') or '',
                # `precio_normal` es el campo que consume el Excel para la
                # columna "Cuota: <cod>". Para un manual usamos el importe
                # total que el operador haya puesto.
                'precio_normal': float(r.get('importe_total') or 0),
                'precio': float(r.get('importe_total') or 0),
                'periodicidad': r.get('periodicidad') or 'mensual',
                'descuentos_struct': [],
                'modificaciones_struct': [],
            }],
            'modificaciones_globales': [],
            'modificaciones_globales_struct': [],
            'forma_pago': METODO_A_FORMA_PAGO.get(r.get('metodo_pago')) or r.get('metodo_pago') or '?',
            'iban': partner_iban,
            'periodicidad': r.get('periodicidad') or 'mensual',
            'importe_total': round(float(r.get('importe_total') or 0), 2),
            'categoria': '',
            # Marcadores propios — el Excel los usa para destacarlos.
            '_manual': True,
            '_recibo_bd_id': r['id'],
            '_notas': r.get('notas') or '',
        })

    # ─── Scope por trainer (auditoría #22) ────────────────────────────────
    # Si se valida operando COMO un trainer concreto, se filtran coherentes e
    # incoherencias a SUS clientes (por el trainer real del cliente en
    # cliente_cache). El manager sin scope ve todo. Se filtra ANTES del
    # resumen para que los totales por trainer cuadren con lo que se emitirá.
    if id_trainer:
        _scope = str(id_trainer)
        coherentes = [c for c in coherentes
                      if cache_idnoofit_trainer.get(str(c.get('idnoofit') or '')) == _scope]
        incoherencias = [i for i in incoherencias
                         if cache_idnoofit_trainer.get(
                             str((i.get('cliente') or {}).get('idnoofit') or '')) == _scope]

    # ─── Resumen agregado de lo que se va a emitir ───────────────────────
    resumen = _resumir_emision(coherentes)

    return coherentes, incoherencias, resumen


def _historico_mensual(id_manager, mes_objetivo, n_meses=13):
    """Extrae desglose mensual de recibos ya emitidos (tabla `recibo`).

    Devuelve dict ordenado: {YYYY-MM: {total_recibos, total_importe,
                                       por_forma_pago, por_periodicidad,
                                       por_cuota}}
    Solo incluye meses ANTERIORES a mes_objetivo (no incluye mes_objetivo).

    Mapeo metodo_pago BD → forma_pago canónica:
      sepa → sepa | caja_* → efectivo | tarjeta_tok → tarjeta_token
      enlace_pago → enlace_pago
    """
    import datetime as dt
    from collections import defaultdict
    # Calcular ventana
    y, m = mes_objetivo.split('-')
    fecha_obj = dt.date(int(y), int(m), 1)
    meses = []
    for i in range(n_meses, 0, -1):
        # Restar i meses
        mm = fecha_obj.month - i; yy = fecha_obj.year
        while mm <= 0: mm += 12; yy -= 1
        meses.append(f'{yy}-{mm:02d}')

    if not meses: return {}

    METODO_MAP = {
        'sepa': 'sepa',
        'caja_efectivo': 'efectivo',
        'caja_tpv_fisico': 'efectivo',
        'caja_tpv_virtual': 'efectivo',
        'tarjeta_tok': 'tarjeta_token',
        'enlace_pago': 'enlace_pago',
    }

    def _familia_cuota(codigo):
        """Mapea códigos GestPlus antiguos a las familias de cuota actuales.
          RT LX/MJ * → RT 2 dias  (antes 2 sesiones/semana, ahora 1 sub)
          RT 1D     → RT 1D
          I MYGYM * → I MYGYM
          otros     → tal cual
        """
        if not codigo: return '?'
        c = codigo.strip()
        cu = c.upper()
        if cu.startswith('RT 1D'): return 'RT 1D'
        if cu.startswith('RT LX') or cu.startswith('RT MJ'): return 'RT 2 dias'
        if cu.startswith('I MYGYM'): return 'I MYGYM'
        return c

    out = {m: {
        'total_recibos': 0, 'total_importe': 0.0,
        'por_forma_pago': defaultdict(lambda: {'n': 0, 'importe': 0.0}),
        'por_periodicidad': defaultdict(lambda: {'n': 0, 'importe': 0.0}),
        'por_cuota': defaultdict(lambda: {'n': 0, 'importe': 0.0}),
    } for m in meses}

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT periodo, metodo_pago, periodicidad, cuota_codigo,
                   COUNT(*) AS n, COALESCE(SUM(importe_total), 0) AS total
              FROM recibo
             WHERE id_manager = %s AND periodo = ANY(%s)
               AND estado IN ('emitido', 'pagado', 'facturado')
             GROUP BY periodo, metodo_pago, periodicidad, cuota_codigo
        """, (str(id_manager), meses))
        for r in cur.fetchall():
            mes = r['periodo']
            if mes not in out: continue
            n = int(r['n']); total = float(r['total'] or 0)
            out[mes]['total_recibos'] += n
            out[mes]['total_importe'] += total
            fp = METODO_MAP.get(r['metodo_pago'], r['metodo_pago'] or '?')
            out[mes]['por_forma_pago'][fp]['n'] += n
            out[mes]['por_forma_pago'][fp]['importe'] += total
            per = r['periodicidad'] or 'mensual'
            out[mes]['por_periodicidad'][per]['n'] += n
            out[mes]['por_periodicidad'][per]['importe'] += total
            cod = _familia_cuota(r['cuota_codigo'])
            out[mes]['por_cuota'][cod]['n'] += n
            out[mes]['por_cuota'][cod]['importe'] += total

    # Convertir defaultdict a dict y redondear; filtrar meses sin datos
    final = {}
    for mes in meses:  # respeta orden cronológico
        if out[mes]['total_recibos'] == 0:
            continue   # no incluimos meses sin recibos
        out[mes]['total_importe'] = round(out[mes]['total_importe'], 2)
        for k in ('por_forma_pago', 'por_periodicidad', 'por_cuota'):
            d = out[mes][k]
            for kk in d:
                d[kk]['importe'] = round(d[kk]['importe'], 2)
            out[mes][k] = dict(d)
        final[mes] = out[mes]
    return final


def _resumir_emision(coherentes):
    """Calcula totales agregados para los recibos a emitir.

    Devuelve un dict con:
      - total_recibos / total_importe
      - por_forma_pago: {sepa: {n, importe}, efectivo:..., tarjeta_token:...}
      - por_cuota: {codigo: {n, importe}}
      - por_periodicidad: {mensual: {n, importe}, ...}
    """
    total_recibos = len(coherentes)
    total_importe = round(sum(c.get('importe_total', 0) for c in coherentes), 2)

    por_forma_pago = defaultdict(lambda: {'n': 0, 'importe': 0.0})
    por_periodicidad = defaultdict(lambda: {'n': 0, 'importe': 0.0})
    por_cuota = defaultdict(lambda: {'n': 0, 'importe': 0.0})
    por_trainer = defaultdict(lambda: {'n': 0, 'importe': 0.0})

    for c in coherentes:
        imp = c.get('importe_total', 0) or 0
        fp = c.get('forma_pago', '?')
        per = c.get('periodicidad', 'mensual')
        tr = str(c.get('id_trainer') or 'sin_trainer')
        por_forma_pago[fp]['n'] += 1
        por_forma_pago[fp]['importe'] += imp
        por_periodicidad[per]['n'] += 1
        por_periodicidad[per]['importe'] += imp
        por_trainer[tr]['n'] += 1
        por_trainer[tr]['importe'] += imp
        # Por cuota: usar el precio_final real por cuota (con descuentos+mods).
        # Si no hay detalle (compatibilidad), repartir el total a partes iguales.
        detalle = c.get('cuotas_detalle') or []
        if detalle:
            for d in detalle:
                cod = d.get('codigo') or '?'
                por_cuota[cod]['n'] += 1
                por_cuota[cod]['importe'] += float(d.get('precio_final') or 0)
        else:
            for cod in (c.get('cuotas') or ['?']):
                por_cuota[cod]['n'] += 1
                por_cuota[cod]['importe'] += imp / max(1, len(c.get('cuotas') or [1]))

    # Redondear importes
    for d in (por_forma_pago, por_periodicidad, por_cuota, por_trainer):
        for k in d:
            d[k]['importe'] = round(d[k]['importe'], 2)

    return {
        'total_recibos': total_recibos,
        'total_importe': total_importe,
        'por_trainer': dict(por_trainer),
        'por_forma_pago': dict(por_forma_pago),
        'por_periodicidad': dict(por_periodicidad),
        'por_cuota': dict(por_cuota),
    }


@bp.route('/<mes>/validar', methods=['GET'])
@auth_required
def validar(mes):
    try:
        _tr = str(g.id_trainer) if getattr(g, 'id_trainer', None) else None
        coherentes, incoherencias, resumen = _validar_emision(g.id_manager, mes, _tr)
        # Stats por tipo
        from collections import Counter
        por_tipo = dict(Counter(i['tipo'] for i in incoherencias))
        try:
            historico = _historico_mensual(g.id_manager, mes, n_meses=13)
        except Exception as e:
            log.warning(f'historico_mensual: {e}')
            historico = {}
        return jsonify({
            'ok': True,
            'mes': mes,
            'coherentes': len(coherentes),
            'incoherencias': len(incoherencias),
            'por_tipo': por_tipo,
            'resumen_emision': resumen,
            'historico': historico,
            'detalle': incoherencias[:200],   # cap
        })
    except Exception as e:
        log.exception('validar')
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/<mes>/validar/excel', methods=['GET'])
@auth_required
def validar_excel(mes):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        _tr = str(g.id_trainer) if getattr(g, 'id_trainer', None) else None
        coherentes, incoherencias, resumen = _validar_emision(g.id_manager, mes, _tr)

        wb = Workbook(); wb.remove(wb.active)
        border = Border(*[Side(style='thin', color='CCCCCC')] * 4)
        font_h = Font(bold=True, color='FFFFFF', size=11)
        font_subh = Font(bold=True, size=11, color='1F2937')
        fill_total = PatternFill('solid', fgColor='DBEAFE')

        # ─── RESUMEN ─────────────────────────────────────────────────────
        ws = wb.create_sheet('RESUMEN')
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14

        ws.cell(1, 1, f'Validación pre-emisión — {mes}').font = Font(bold=True, size=14)

        # Totales generales
        ws.cell(3, 1, 'Total recibos a emitir:').font = font_subh
        ws.cell(3, 2, resumen.get('total_recibos', 0)).font = Font(bold=True)
        ws.cell(3, 2).fill = fill_total
        ws.cell(4, 1, 'Importe total (€):').font = font_subh
        ws.cell(4, 2, resumen.get('total_importe', 0)).font = Font(bold=True)
        ws.cell(4, 2).fill = fill_total
        ws.cell(4, 2).number_format = '#,##0.00 €'

        ws.cell(5, 1, 'Incoherencias detectadas:').font = Font(bold=True, color='C00000')
        ws.cell(5, 2, len(incoherencias))

        # Por forma de pago
        r = 7
        ws.cell(r, 1, 'POR FORMA DE PAGO').font = Font(bold=True, color='FFFFFF')
        ws.cell(r, 1).fill = PatternFill('solid', fgColor='2563EB')
        ws.cell(r, 2, '# Recibos').font = Font(bold=True, color='FFFFFF')
        ws.cell(r, 2).fill = PatternFill('solid', fgColor='2563EB')
        ws.cell(r, 3, 'Importe €').font = Font(bold=True, color='FFFFFF')
        ws.cell(r, 3).fill = PatternFill('solid', fgColor='2563EB')
        for col in (1, 2, 3): ws.cell(r, col).alignment = Alignment(horizontal='center')
        r += 1
        for fp, d in sorted(resumen.get('por_forma_pago', {}).items(),
                            key=lambda x: -x[1]['importe']):
            ws.cell(r, 1, fp)
            ws.cell(r, 2, d['n'])
            ws.cell(r, 3, d['importe']).number_format = '#,##0.00 €'
            r += 1

        # Por periodicidad
        r += 1
        ws.cell(r, 1, 'POR PERIODICIDAD').font = Font(bold=True, color='FFFFFF')
        ws.cell(r, 1).fill = PatternFill('solid', fgColor='7C3AED')
        ws.cell(r, 2, '# Recibos').font = Font(bold=True, color='FFFFFF')
        ws.cell(r, 2).fill = PatternFill('solid', fgColor='7C3AED')
        ws.cell(r, 3, 'Importe €').font = Font(bold=True, color='FFFFFF')
        ws.cell(r, 3).fill = PatternFill('solid', fgColor='7C3AED')
        for col in (1, 2, 3): ws.cell(r, col).alignment = Alignment(horizontal='center')
        r += 1
        for per, d in sorted(resumen.get('por_periodicidad', {}).items(),
                             key=lambda x: -x[1]['importe']):
            ws.cell(r, 1, per)
            ws.cell(r, 2, d['n'])
            ws.cell(r, 3, d['importe']).number_format = '#,##0.00 €'
            r += 1

        # Por cuota
        r += 1
        ws.cell(r, 1, 'POR CUOTA').font = Font(bold=True, color='FFFFFF')
        ws.cell(r, 1).fill = PatternFill('solid', fgColor='059669')
        ws.cell(r, 2, '# Recibos').font = Font(bold=True, color='FFFFFF')
        ws.cell(r, 2).fill = PatternFill('solid', fgColor='059669')
        ws.cell(r, 3, 'Importe €').font = Font(bold=True, color='FFFFFF')
        ws.cell(r, 3).fill = PatternFill('solid', fgColor='059669')
        for col in (1, 2, 3): ws.cell(r, col).alignment = Alignment(horizontal='center')
        r += 1
        for cod, d in sorted(resumen.get('por_cuota', {}).items(),
                             key=lambda x: -x[1]['importe']):
            ws.cell(r, 1, cod)
            ws.cell(r, 2, d['n'])
            ws.cell(r, 3, d['importe']).number_format = '#,##0.00 €'
            r += 1

        # ─── COMPARATIVA HISTÓRICA ───────────────────────────────────────
        try:
            historico = _historico_mensual(g.id_manager, mes, n_meses=13)
        except Exception as e:
            log.warning(f'historico_mensual: {e}')
            historico = {}

        if historico:
            meses_orden = list(historico.keys()) + [mes]
            # Recopilar todas las claves a comparar
            forma_pagos_all = set()
            periodicidades_all = set()
            cuotas_all = set()
            for mh in historico.values():
                forma_pagos_all.update(mh['por_forma_pago'].keys())
                periodicidades_all.update(mh['por_periodicidad'].keys())
                cuotas_all.update(mh['por_cuota'].keys())
            forma_pagos_all.update(resumen.get('por_forma_pago', {}).keys())
            periodicidades_all.update(resumen.get('por_periodicidad', {}).keys())
            cuotas_all.update(resumen.get('por_cuota', {}).keys())

            r += 2
            ws.cell(r, 1, f'COMPARATIVA HISTÓRICA — últimos {len(historico)} meses + previsión {mes}').font = Font(bold=True, color='FFFFFF', size=12)
            ws.cell(r, 1).fill = PatternFill('solid', fgColor='0F766E')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(meses_orden) + 2)
            r += 1

            # Cabeceras: Concepto | mes1 | mes2 | ... | mes_obj* | Δ %
            fill_thead = PatternFill('solid', fgColor='115E59')
            ws.cell(r, 1, 'Concepto').font = font_h
            ws.cell(r, 1).fill = fill_thead
            ws.cell(r, 1).border = border
            for col_idx, mh in enumerate(meses_orden, start=2):
                label = mh + (' *' if mh == mes else '')
                c = ws.cell(r, col_idx, label)
                c.font = font_h; c.fill = fill_thead; c.border = border
                c.alignment = Alignment(horizontal='center')
            delta_col = len(meses_orden) + 2
            c = ws.cell(r, delta_col, 'Δ % vs mes ant.')
            c.font = font_h; c.fill = fill_thead; c.border = border
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(delta_col)].width = 16

            for col_idx in range(2, delta_col):
                ws.column_dimensions[get_column_letter(col_idx)].width = 13

            r += 1

            def _val_mes(mh_key, modo, cat=None):
                """Devuelve nº/importe de una categoría en un mes dado.
                modo: 'n_total', 'imp_total', 'n_fp', 'imp_fp', etc."""
                if mh_key == mes:
                    src = resumen
                else:
                    src = historico.get(mh_key, {})
                if modo == 'n_total': return src.get('total_recibos', 0)
                if modo == 'imp_total': return src.get('total_importe', 0)
                if modo == 'n_fp': return src.get('por_forma_pago', {}).get(cat, {}).get('n', 0)
                if modo == 'imp_fp': return src.get('por_forma_pago', {}).get(cat, {}).get('importe', 0)
                if modo == 'n_per': return src.get('por_periodicidad', {}).get(cat, {}).get('n', 0)
                if modo == 'imp_per': return src.get('por_periodicidad', {}).get(cat, {}).get('importe', 0)
                if modo == 'n_cuota': return src.get('por_cuota', {}).get(cat, {}).get('n', 0)
                if modo == 'imp_cuota': return src.get('por_cuota', {}).get(cat, {}).get('importe', 0)
                return 0

            def _add_row(concepto, modo, cat=None, is_imp=False, fill_row=None,
                         bold_label=False):
                nonlocal r
                cell = ws.cell(r, 1, concepto)
                cell.border = border
                if bold_label: cell.font = Font(bold=True)
                if fill_row: cell.fill = fill_row
                vals = []
                for col_idx, mh_key in enumerate(meses_orden, start=2):
                    v = _val_mes(mh_key, modo, cat)
                    vals.append(v)
                    c = ws.cell(r, col_idx, v)
                    c.border = border
                    if is_imp: c.number_format = '#,##0.00 €'
                    if fill_row: c.fill = fill_row
                # Δ%
                if len(vals) >= 2 and vals[-2]:
                    delta = (vals[-1] - vals[-2]) / vals[-2] * 100
                    c = ws.cell(r, delta_col, delta / 100)
                    c.number_format = '+0.0%;-0.0%;0.0%'
                    c.border = border
                    if delta < -5:
                        c.font = Font(color='C00000', bold=True)
                    elif delta > 5:
                        c.font = Font(color='006400', bold=True)
                    if fill_row: c.fill = fill_row
                else:
                    c = ws.cell(r, delta_col, '—')
                    c.alignment = Alignment(horizontal='center')
                    c.border = border
                    if fill_row: c.fill = fill_row
                r += 1

            # Totales
            fill_total_row = PatternFill('solid', fgColor='DBEAFE')
            _add_row('TOTAL # recibos', 'n_total', bold_label=True, fill_row=fill_total_row)
            _add_row('TOTAL importe (€)', 'imp_total', is_imp=True, bold_label=True, fill_row=fill_total_row)

            # Por forma de pago
            r += 1
            ws.cell(r, 1, '— Por forma de pago —').font = Font(italic=True, color='2563EB')
            r += 1
            for fp in sorted(forma_pagos_all):
                _add_row(f'  {fp} — # recibos', 'n_fp', cat=fp)
                _add_row(f'  {fp} — importe', 'imp_fp', cat=fp, is_imp=True)

            # Por periodicidad
            r += 1
            ws.cell(r, 1, '— Por periodicidad —').font = Font(italic=True, color='7C3AED')
            r += 1
            for per in sorted(periodicidades_all):
                _add_row(f'  {per} — # recibos', 'n_per', cat=per)
                _add_row(f'  {per} — importe', 'imp_per', cat=per, is_imp=True)

            # Por cuota
            r += 1
            ws.cell(r, 1, '— Por cuota —').font = Font(italic=True, color='059669')
            r += 1
            for cod in sorted(cuotas_all):
                _add_row(f'  {cod} — # recibos', 'n_cuota', cat=cod)
                _add_row(f'  {cod} — importe', 'imp_cuota', cat=cod, is_imp=True)

        # Tipos de incoherencia (con etiqueta amigable para el gestor)
        if incoherencias:
            r += 1
            # Cabecera (3 columnas: motivo | # casos | código técnico)
            ws.cell(r, 1, 'MOTIVOS DE EXCLUSIÓN').font = Font(bold=True, color='FFFFFF')
            ws.cell(r, 1).fill = PatternFill('solid', fgColor='DC2626')
            ws.cell(r, 2, '# Casos').font = Font(bold=True, color='FFFFFF')
            ws.cell(r, 2).fill = PatternFill('solid', fgColor='DC2626')
            ws.cell(r, 3, 'Código técnico').font = Font(bold=True, color='FFFFFF')
            ws.cell(r, 3).fill = PatternFill('solid', fgColor='DC2626')
            for col in (1, 2, 3): ws.cell(r, col).alignment = Alignment(horizontal='center')
            # Ensanchar col A para el texto descriptivo
            ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width or 38, 55)
            r += 1
            from collections import Counter
            tipos = Counter(i['tipo'] for i in incoherencias)
            for t, n in tipos.most_common():
                ws.cell(r, 1, _label_tipo(t))
                ws.cell(r, 2, n)
                ws.cell(r, 3, t).font = Font(color='6B7280', size=10)
                r += 1

        # ─── INCOHERENCIAS ──────────────────────────────────────────────
        if incoherencias:
            ws = wb.create_sheet('INCOHERENCIAS')
            heads = [('Motivo', 50), ('Código técnico', 22),
                     ('Cliente', 30), ('idnoofit', 11), ('Email', 26),
                     ('Detalle', 50), ('Propuesta de solución', 60)]
            heads_inc_correc = [
                ('Último recibo pagado',    20),
                ('Importe último pagado',   16),
                ('Concepto último pagado',  30),
                ('Fecha inicio pago',       14),
                ('Fecha fin pago',          14),
                ('Nueva categoría',         18),
                ('Nueva cuota',             18),
                ('Nueva periodicidad',      18),
            ]
            n_inc = len(heads)
            all_heads_inc = heads + heads_inc_correc
            fill_h          = PatternFill('solid', fgColor='F87171')
            fill_hist_inc   = PatternFill('solid', fgColor='0EA5E9')
            fill_corr_inc   = PatternFill('solid', fgColor='EC4899')
            for col, (h, w) in enumerate(all_heads_inc, 1):
                c = ws.cell(1, col, h); c.font = font_h
                c.alignment = Alignment(horizontal='center', wrap_text=True)
                c.border = border
                ws.column_dimensions[get_column_letter(col)].width = w
                if col <= n_inc:
                    c.fill = fill_h
                elif col <= n_inc + 5:
                    c.fill = fill_hist_inc
                else:
                    c.fill = fill_corr_inc
            ws.row_dimensions[1].height = 26
            ws.freeze_panes = 'A2'

            # Necesitamos los mismos catálogos que en OK para el histórico
            # y los dropdowns. Si la hoja OK no se generó, los cargamos aquí.
            ult_map_inc = _ultimo_recibo_pagado_por_cliente(g.id_manager)
            cats_inc    = _categorias_nombres(g.id_manager)
            cuotas_inc  = _cuotas_codigos(g.id_manager, _tr)

            col_h_id      = n_inc + 1
            col_h_importe = n_inc + 2
            col_h_concep  = n_inc + 3
            col_h_desde   = n_inc + 4
            col_h_hasta   = n_inc + 5
            col_c_cat     = n_inc + 6
            col_c_cuota   = n_inc + 7
            col_c_per     = n_inc + 8

            for i, inc in enumerate(incoherencias, 2):
                cli = inc['cliente']
                vals = [_label_tipo(inc['tipo']), inc['tipo'],
                        cli['nombre'], cli['idnoofit'], cli['email'],
                        inc['detalle'], inc['propuesta']]
                for j, v in enumerate(vals, 1):
                    ws.cell(i, j, v).border = border
                # Histórico de pago (lectura)
                ult = ult_map_inc.get(str(cli.get('idnoofit') or ''))
                if ult:
                    ws.cell(i, col_h_id, ult.get('id') or '').border = border
                    imp = ult.get('importe_total')
                    if imp is not None:
                        cell = ws.cell(i, col_h_importe, float(imp))
                        cell.number_format = '#,##0.00 €'
                        cell.border = border
                    else:
                        ws.cell(i, col_h_importe, '').border = border
                    ws.cell(i, col_h_concep, ult.get('concepto') or '').border = border
                    fd = ult.get('fecha_desde'); fh = ult.get('fecha_hasta')
                    ws.cell(i, col_h_desde, fd.isoformat() if fd else '').border = border
                    ws.cell(i, col_h_hasta, fh.isoformat() if fh else '').border = border
                else:
                    for col_h in (col_h_id, col_h_importe, col_h_concep,
                                  col_h_desde, col_h_hasta):
                        ws.cell(i, col_h, '').border = border
                # Correcciones (editable)
                for col_corr in (col_c_cat, col_c_cuota, col_c_per):
                    cell = ws.cell(i, col_corr, '')
                    cell.border = border
                    cell.fill = PatternFill('solid', fgColor='FCE7F3')

            # DataValidation para las 3 columnas editables de INCOHERENCIAS
            from openpyxl.worksheet.datavalidation import DataValidation
            def _dv_for_inc(values):
                if not values:
                    return None
                joined = ','.join(v.replace(',', ' ') for v in values)
                if len(joined) > 250:
                    return None
                dv = DataValidation(type='list', formula1=f'"{joined}"',
                                    allow_blank=True, showDropDown=False)
                dv.error = 'Valor no permitido'
                dv.errorTitle = 'Selecciona uno de la lista'
                return dv
            n_inc_rows = len(incoherencias)
            if n_inc_rows > 0:
                for dv_vals, dv_col in (
                    (cats_inc,   col_c_cat),
                    (cuotas_inc, col_c_cuota),
                    (PERIODICIDADES_VALIDAS, col_c_per),
                ):
                    dv = _dv_for_inc(dv_vals)
                    if dv is not None:
                        ws.add_data_validation(dv)
                        dv.add(f'{get_column_letter(dv_col)}2:'
                               f'{get_column_letter(dv_col)}{n_inc_rows + 1}')

            ws.auto_filter.ref = ws.dimensions

        # ─── OK (recibos a emitir, con columnas dinámicas) ──────────────
        # Estructura: código cliente, nombre, apellido, categoría, forma pago,
        # periodicidad, [columna por cada CUOTA del catálogo del manager con
        # el precio del cliente en ella], [columna por cada DESCUENTO del
        # catálogo con el importe ahorrado por ese cliente], modificación €,
        # nota modificación, importe total €.
        ws = wb.create_sheet('OK')

        # Cuotas del catálogo del manager (de Odoo: round.cuota.catalogo).
        # Filtrar a las que están realmente en uso por los coherentes.
        codigos_cuota_orden = sorted({d['codigo'] for c in coherentes
                                      for d in (c.get('cuotas_detalle') or [])
                                      if d.get('codigo')})

        # Descuentos en uso por los coherentes (catálogo local).
        codigos_desc_orden = sorted({x['codigo'] for c in coherentes
                                      for d in (c.get('cuotas_detalle') or [])
                                      for x in (d.get('descuentos_struct') or [])
                                      if x.get('codigo')})

        # Cabecera
        heads_fijas = [
            ('Cód. cliente', 12),
            ('Nombre', 18),
            ('Apellido', 22),
            ('Categoría', 14),
            ('Forma pago', 14),
            ('Periodicidad', 12),
        ]
        heads_modif = [
            ('Modificación €', 14),
            ('Nota modificación', 40),
            ('IMPORTE TOTAL €', 16),
        ]
        # Bloque "Histórico de pago" + "Correcciones a aplicar".
        # Las 3 últimas columnas las EDITA el usuario (vacías por defecto)
        # con dropdowns; el endpoint /aplicar-correcciones-excel las lee.
        heads_correccion = [
            ('Último recibo pagado',    20),
            ('Importe último pagado',   16),
            ('Concepto último pagado',  30),
            ('Fecha inicio pago',       14),
            ('Fecha fin pago',          14),
            ('Nueva categoría',         18),
            ('Nueva cuota',             18),
            ('Nueva periodicidad',      18),
        ]
        # Construir lista completa de columnas
        all_heads = list(heads_fijas)
        for cod in codigos_cuota_orden:
            all_heads.append((f'Cuota: {cod}', 14))
        for cod in codigos_desc_orden:
            all_heads.append((f'Desc: {cod}', 14))
        all_heads.extend(heads_modif)
        all_heads.extend(heads_correccion)

        fill_fijas      = PatternFill('solid', fgColor='2DD4A8')
        fill_cuotas     = PatternFill('solid', fgColor='059669')
        fill_desc       = PatternFill('solid', fgColor='F59E0B')
        fill_modif      = PatternFill('solid', fgColor='7C3AED')
        fill_historico  = PatternFill('solid', fgColor='0EA5E9')   # azul: lectura
        fill_correccion = PatternFill('solid', fgColor='EC4899')   # rosa: editable

        # Límites por bloque (1-based, inclusivo en el extremo derecho)
        lim_fijas      = len(heads_fijas)
        lim_cuotas     = lim_fijas + len(codigos_cuota_orden)
        lim_desc       = lim_cuotas + len(codigos_desc_orden)
        lim_modif      = lim_desc + len(heads_modif)            # +3 modif
        lim_historico  = lim_modif + 5                          # +5 lectura
        lim_correccion = lim_historico + 3                      # +3 editable

        for col, (h, w) in enumerate(all_heads, 1):
            c = ws.cell(1, col, h); c.font = font_h
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            c.border = border
            ws.column_dimensions[get_column_letter(col)].width = w
            # Color por bloque
            if col <= lim_fijas:
                c.fill = fill_fijas
            elif col <= lim_cuotas:
                c.fill = fill_cuotas
            elif col <= lim_desc:
                c.fill = fill_desc
            elif col <= lim_modif:
                c.fill = fill_modif
            elif col <= lim_historico:
                c.fill = fill_historico
            else:
                c.fill = fill_correccion
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = ws.cell(2, len(heads_fijas) + 1).coordinate

        col_modif_eur  = lim_desc + 1
        col_modif_nota = col_modif_eur + 1
        col_total      = col_modif_eur + 2
        # Columnas histórico (lectura)
        col_hist_id      = col_total + 1
        col_hist_importe = col_total + 2
        col_hist_concep  = col_total + 3
        col_hist_desde   = col_total + 4
        col_hist_hasta   = col_total + 5
        # Columnas corrección (editables)
        col_corr_cat   = col_total + 6
        col_corr_cuota = col_total + 7
        col_corr_per   = col_total + 8
        col_ultima     = col_corr_per

        # Datos auxiliares para histórico de pago y dropdowns de corrección.
        # Un solo query por tabla — no degradamos la generación del Excel.
        ultimo_recibo_map = _ultimo_recibo_pagado_por_cliente(g.id_manager)
        categorias_lista  = _categorias_nombres(g.id_manager)
        cuotas_lista      = _cuotas_codigos(g.id_manager, _tr)

        # Fill amber para destacar recibos manuales (borrador_remesa) que no
        # vienen del flujo auto-generado de Odoo.
        fill_manual = PatternFill('solid', fgColor='FEF3C7')

        for i, c in enumerate(coherentes, 2):
            es_manual = bool(c.get('_manual'))
            # Bloque 1: identidad + datos básicos
            ws.cell(i, 1, c.get('idnoofit') or '').border = border
            ws.cell(i, 2, c.get('nombre_solo') or '').border = border
            ws.cell(i, 3, c.get('apellido') or '').border = border
            ws.cell(i, 4, 'MANUAL' if es_manual else (c.get('categoria') or '')).border = border
            ws.cell(i, 5, c.get('forma_pago') or '').border = border
            ws.cell(i, 6, c.get('periodicidad') or '').border = border

            # Bloque 2: una columna por cada cuota del catálogo → precio de
            # CONFIGURACIÓN (precio_normal). El descuento aplicado va en su
            # columna aparte. Así la fila suma: precio_cuota − descuento +
            # modificación = importe total.
            precios_por_cuota = {d['codigo']: float(d.get('precio_normal') or 0)
                                 for d in (c.get('cuotas_detalle') or [])
                                 if d.get('codigo')}
            for k, cod in enumerate(codigos_cuota_orden):
                col_idx = len(heads_fijas) + 1 + k
                if cod in precios_por_cuota:
                    cell = ws.cell(i, col_idx, precios_por_cuota[cod])
                    cell.number_format = '#,##0.00 €'
                else:
                    cell = ws.cell(i, col_idx, '')
                cell.border = border

            # Bloque 3: una columna por cada descuento → importe ahorrado
            ahorro_por_desc = {}
            for d in (c.get('cuotas_detalle') or []):
                for x in (d.get('descuentos_struct') or []):
                    cod = x.get('codigo')
                    if cod:
                        ahorro_por_desc[cod] = ahorro_por_desc.get(cod, 0.0) + float(x.get('ahorro') or 0)
            for k, cod in enumerate(codigos_desc_orden):
                col_idx = len(heads_fijas) + len(codigos_cuota_orden) + 1 + k
                val = ahorro_por_desc.get(cod)
                if val:
                    cell = ws.cell(i, col_idx, -round(val, 2))  # negativo: es un descuento
                    cell.number_format = '#,##0.00 €'
                    cell.font = Font(color='006400')
                else:
                    cell = ws.cell(i, col_idx, '')
                cell.border = border

            # Bloque 4: modificación €  + nota
            # Usamos `delta` (precio_despues − precio_antes) que es el efecto
            # REAL en el total. Para precio_alternativo `delta` puede diferir
            # del `valor` crudo (sustituye precio en vez de sumar/restar).
            mod_total = 0.0
            mod_notas = []
            for d in (c.get('cuotas_detalle') or []):
                for m in (d.get('modificaciones_struct') or []):
                    delta = float(m.get('delta') or 0)
                    mod_total += delta
                    tipo = m.get('tipo') or ''
                    if tipo == 'precio_alternativo':
                        n = (f"{d['codigo']}: precio_alternativo "
                             f"{m.get('valor'):.2f}€ (delta {'+' if delta >= 0 else '−'}{abs(delta):.2f}€)")
                    else:
                        n = f"{d['codigo']}: {'+' if delta >= 0 else '−'}{abs(delta):.2f}€"
                    if m.get('razon'): n += f' ({m["razon"]})'
                    mod_notas.append(n)
            for m in (c.get('modificaciones_globales_struct') or []):
                delta = float(m.get('delta') or 0)
                mod_total += delta
                tipo = m.get('tipo') or ''
                if tipo == 'precio_alternativo':
                    n = (f"global: precio_alternativo {m.get('valor'):.2f}€ "
                         f"(delta {'+' if delta >= 0 else '−'}{abs(delta):.2f}€)")
                else:
                    n = f"global: {'+' if delta >= 0 else '−'}{abs(delta):.2f}€"
                if m.get('razon'): n += f' ({m["razon"]})'
                mod_notas.append(n)
            if mod_total != 0:
                cell = ws.cell(i, col_modif_eur, round(mod_total, 2))
                cell.number_format = '#,##0.00 €'
                cell.font = Font(color='C00000' if mod_total < 0 else '0066CC', bold=True)
            else:
                ws.cell(i, col_modif_eur, '')
            ws.cell(i, col_modif_eur).border = border
            nota_cell = ws.cell(i, col_modif_nota, ' · '.join(mod_notas))
            nota_cell.border = border
            nota_cell.alignment = Alignment(wrap_text=False, horizontal='left')

            # Importe total
            tot = ws.cell(i, col_total, float(c.get('importe_total') or 0))
            tot.number_format = '#,##0.00 €'
            tot.font = Font(bold=True)
            tot.border = border

            # Bloque 5: histórico de pago (lectura) — último recibo pagado.
            ult = ultimo_recibo_map.get(str(c.get('idnoofit') or ''))
            if ult:
                ws.cell(i, col_hist_id, ult.get('id') or '').border = border
                imp = ult.get('importe_total')
                if imp is not None:
                    cell = ws.cell(i, col_hist_importe, float(imp))
                    cell.number_format = '#,##0.00 €'
                    cell.border = border
                else:
                    ws.cell(i, col_hist_importe, '').border = border
                ws.cell(i, col_hist_concep, ult.get('concepto') or '').border = border
                fd = ult.get('fecha_desde'); fh = ult.get('fecha_hasta')
                ws.cell(i, col_hist_desde, fd.isoformat() if fd else '').border = border
                ws.cell(i, col_hist_hasta, fh.isoformat() if fh else '').border = border
            else:
                for col_h in (col_hist_id, col_hist_importe, col_hist_concep,
                              col_hist_desde, col_hist_hasta):
                    ws.cell(i, col_h, '').border = border

            # Bloque 6: correcciones a aplicar (editable, vacío por defecto).
            # El usuario rellena en Excel y luego sube vía endpoint
            # /aplicar-correcciones-excel. DataValidation se añade tras el loop.
            for col_corr in (col_corr_cat, col_corr_cuota, col_corr_per):
                cell = ws.cell(i, col_corr, '')
                cell.border = border
                cell.fill = PatternFill('solid', fgColor='FCE7F3')   # rosa claro

            # Si es manual, pintar toda la fila en amber para que destaque
            # visualmente entre los auto-generados.
            if es_manual:
                for col_i in range(1, col_total + 1):
                    cell = ws.cell(i, col_i)
                    cell.fill = fill_manual
                # Anotamos la nota del operador en la columna de modificación nota
                if c.get('_notas'):
                    existing = ws.cell(i, col_modif_nota).value or ''
                    extra = f'[MANUAL] {c["_notas"]}'
                    ws.cell(i, col_modif_nota,
                            f'{existing} · {extra}' if existing else extra)

        # Dropdowns (DataValidation) en las 3 columnas editables.
        # openpyxl exige que la lista quepa en 255 caracteres (comma-string)
        # o referenciarla desde un rango con celdas. Usamos string si cabe;
        # si no, fallback a sin validación (el endpoint sigue validando).
        from openpyxl.worksheet.datavalidation import DataValidation

        def _dv_for(values):
            if not values:
                return None
            joined = ','.join(v.replace(',', ' ') for v in values)
            if len(joined) > 250:
                return None  # supera el límite Excel/openpyxl
            dv = DataValidation(type='list', formula1=f'"{joined}"',
                                allow_blank=True, showDropDown=False)
            dv.error = 'Valor no permitido'
            dv.errorTitle = 'Selecciona uno de la lista'
            return dv

        n_rows_data = len(coherentes)
        if n_rows_data > 0:
            rng_cat   = f'{get_column_letter(col_corr_cat)}2:{get_column_letter(col_corr_cat)}{n_rows_data + 1}'
            rng_cuo   = f'{get_column_letter(col_corr_cuota)}2:{get_column_letter(col_corr_cuota)}{n_rows_data + 1}'
            rng_per   = f'{get_column_letter(col_corr_per)}2:{get_column_letter(col_corr_per)}{n_rows_data + 1}'

            dv_cat = _dv_for(categorias_lista)
            if dv_cat is not None:
                ws.add_data_validation(dv_cat)
                dv_cat.add(rng_cat)
            dv_cuo = _dv_for(cuotas_lista)
            if dv_cuo is not None:
                ws.add_data_validation(dv_cuo)
                dv_cuo.add(rng_cuo)
            dv_per = _dv_for(PERIODICIDADES_VALIDAS)
            if dv_per is not None:
                ws.add_data_validation(dv_per)
                dv_per.add(rng_per)

        # Fila TOTAL
        total_row = len(coherentes) + 2
        lab = ws.cell(total_row, 1, 'TOTAL')
        lab.font = Font(bold=True); lab.fill = fill_total
        # Sumar cada columna de cuota
        for k, cod in enumerate(codigos_cuota_orden):
            col_idx = len(heads_fijas) + 1 + k
            col_letter = get_column_letter(col_idx)
            t = ws.cell(total_row, col_idx,
                        f'=SUM({col_letter}2:{col_letter}{total_row - 1})')
            t.number_format = '#,##0.00 €'
            t.font = Font(bold=True); t.fill = fill_total
        for k, cod in enumerate(codigos_desc_orden):
            col_idx = len(heads_fijas) + len(codigos_cuota_orden) + 1 + k
            col_letter = get_column_letter(col_idx)
            t = ws.cell(total_row, col_idx,
                        f'=SUM({col_letter}2:{col_letter}{total_row - 1})')
            t.number_format = '#,##0.00 €'
            t.font = Font(bold=True); t.fill = fill_total
        # Modificación total
        col_letter = get_column_letter(col_modif_eur)
        t = ws.cell(total_row, col_modif_eur,
                    f'=SUM({col_letter}2:{col_letter}{total_row - 1})')
        t.number_format = '#,##0.00 €'; t.font = Font(bold=True); t.fill = fill_total
        # Importe total
        col_letter = get_column_letter(col_total)
        t = ws.cell(total_row, col_total,
                    f'=SUM({col_letter}2:{col_letter}{total_row - 1})')
        t.number_format = '#,##0.00 €'; t.font = Font(bold=True); t.fill = fill_total

        ws.auto_filter.ref = f'A1:{get_column_letter(col_ultima)}{total_row - 1}'

        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'validacion_emision_{mes}.xlsx')
    except Exception as e:
        log.exception('validar_excel')
        return jsonify({'ok': False, 'error': str(e)}), 500
