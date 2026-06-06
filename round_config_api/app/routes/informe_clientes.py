"""Informe agregado de clientes (listado completo con toda la info clave).

Endpoints:
  GET /api/informes/clientes          → JSON con todas las filas
  GET /api/informes/clientes/excel    → mismo informe en Excel

Columnas:
  Nombre, Apellidos, Teléfono, DNI,
  Forma de pago (SEPA / efectivo / tarjeta_token / enlace_pago / —),
  Alta/Baja (Activo / Inactivo),
  Categoría,
  Con curso / Sin curso (tiene suscripción Odoo activa o no),
  Descuentos asignados (códigos + tipo),
  Tiene modificación vigente (sí/no).

Cubre todos los clientes del manager (enabled true o false), filtrable por
trainer y por categoría.
"""
import datetime as dt
import logging
from io import BytesIO
from flask import Blueprint, request, jsonify, g, send_file

from functools import wraps
from ..auth import auth_required
from ..auth_usuario import usuario_web_required
from ..db import get_conn

bp = Blueprint('informe_clientes', __name__)
log = logging.getLogger(__name__)


def either_auth(fn):
    """Acepta tanto X-Round-Token (manager) como JWT Bearer (usuario_web)."""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            return usuario_web_required(fn)(*args, **kwargs)
        return auth_required(fn)(*args, **kwargs)
    return wrapper


def _odoo():
    from ..odoo_alta import OdooAlta
    o = OdooAlta(); o._connect()
    return o


def _company_id():
    from .. import config as appconfig
    return getattr(appconfig, 'ODOO_COMPANY', 3) or 3


def _build_rows(id_manager, id_trainer_filtro=None, categoria_filtro=None,
                solo_activos=False):
    """Devuelve lista de dicts con la info agregada de cada cliente."""
    hoy = dt.date.today()
    where = ['cc.id_manager = %s']
    vals = [str(id_manager)]
    if id_trainer_filtro:
        where.append('cc.id_trainer = %s'); vals.append(int(id_trainer_filtro))
    if solo_activos:
        where.append('cc.enabled = TRUE')
    if categoria_filtro:
        where.append('cat.nombre = %s'); vals.append(categoria_filtro)

    sql = f"""
        SELECT
          cc.id, cc.name, cc.surname, cc.email, cc.enabled, cc.id_trainer,
          cc.raw_data->>'cellPhone' AS telefono,
          cc.raw_data->>'dni' AS dni,
          cat.nombre AS categoria_nombre,
          fp.forma_pago AS forma_pago,
          (
            SELECT string_agg(d.codigo || ' (' || d.tipo || ')', ', '
                              ORDER BY d.codigo)
              FROM descuento_asignacion da
              JOIN descuento d ON d.id = da.descuento_id
             WHERE da.id_manager = cc.id_manager
               AND da.cliente_idnoofit = cc.id::text
               AND da.estado = 'activa'
               AND d.active = TRUE
          ) AS descuentos_codigos,
          (
            SELECT COUNT(*) > 0
              FROM modificacion m
             WHERE m.id_manager = cc.id_manager
               AND m.cliente_idnoofit = cc.id::text
               AND m.estado = 'activa'
               AND m.fecha_desde <= %s
               AND (m.fecha_hasta IS NULL OR m.fecha_hasta >= %s)
          ) AS tiene_modificacion,
          (
            SELECT string_agg(m.tipo, ', ')
              FROM modificacion m
             WHERE m.id_manager = cc.id_manager
               AND m.cliente_idnoofit = cc.id::text
               AND m.estado = 'activa'
               AND m.fecha_desde <= %s
               AND (m.fecha_hasta IS NULL OR m.fecha_hasta >= %s)
          ) AS modificaciones_tipos
          FROM cliente_cache cc
          LEFT JOIN cliente_categoria ca
                 ON ca.id_manager = cc.id_manager
                AND ca.cliente_idnoofit = cc.id::text
          LEFT JOIN categoria cat ON cat.id = ca.categoria_id
          LEFT JOIN forma_pago_cliente fp
                 ON fp.id_manager = cc.id_manager
                AND fp.cliente_idnoofit = cc.id::text
                AND fp.estado = 'activa'
         WHERE {' AND '.join(where)}
         ORDER BY cc.enabled DESC,
                  COALESCE(cc.surname, '') ASC,
                  COALESCE(cc.name, '') ASC
    """
    # Las dos subqueries usan hoy/hoy → añadimos 2x al final
    vals_expanded = list(vals)
    vals_expanded.insert(len(vals), hoy)
    vals_expanded.insert(len(vals)+1, hoy)
    vals_expanded.insert(len(vals)+2, hoy)
    vals_expanded.insert(len(vals)+3, hoy)
    # Rebuild en orden correcto del SQL (4 placeholders extras tras los where):
    # SQL placeholders order: where (vals) → mod_tiene (2 hoy) → mod_tipos (2 hoy)
    # Pero por cómo lo construí en string, los 4 %s extras van en la cláusula
    # SELECT (subqueries) ANTES del WHERE. psycopg los toma en orden de aparición
    # en la query string. Así que el orden real es: [4 hoy] + vals
    final_vals = [hoy, hoy, hoy, hoy] + vals

    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(sql, final_vals)
        rows = cur.fetchall()

    # Datos de subscriptions activas en Odoo:
    #   idnoofits_con_sub → set (columna "con curso")
    #   sub_info_by_idn   → {idn: [{periodicidad, fecha_proximo_pago}, ...]}
    idnoofits_con_sub = set()
    sub_info_by_idn = {}
    PERIODO_MESES = {'mensual': 1, 'trimestral': 3, 'semestral': 6, 'anual': 12}
    try:
        from dateutil.relativedelta import relativedelta
        o = _odoo()
        company_id = _company_id()
        subs = o._call('round.subscription', 'search_read',
            [('estado', '=', 'activa'), ('company_id', '=', company_id)],
            ['id', 'partner_id', 'periodicidad', 'fecha_inicio'])
        sub_ids = [s['id'] for s in subs]
        # Última factura por subscription para calcular próximo pago
        last_inv_by_sub = {}
        if sub_ids:
            invs = o._call('account.move', 'search_read',
                [('round_subscription_id', 'in', sub_ids),
                 ('move_type', '=', 'out_invoice'),
                 ('state', 'in', ['posted', 'draft'])],
                ['round_subscription_id', 'invoice_date'])
            for i in invs:
                sid = i.get('round_subscription_id')
                sid = sid[0] if isinstance(sid, (list, tuple)) else sid
                d = i.get('invoice_date')
                if sid and d and (last_inv_by_sub.get(sid, '') < d):
                    last_inv_by_sub[sid] = d

        sub_info_by_partner = {}
        for s in subs:
            if not s.get('partner_id'):
                continue
            pid = s['partner_id'][0]
            per = s.get('periodicidad') or 'mensual'
            last_d_str = last_inv_by_sub.get(s['id']) or s.get('fecha_inicio')
            fecha_proximo = None
            if last_d_str:
                try:
                    ld = dt.date.fromisoformat(last_d_str)
                    fecha_proximo = ld + relativedelta(months=PERIODO_MESES.get(per, 1))
                except Exception:
                    fecha_proximo = None
            sub_info_by_partner.setdefault(pid, []).append({
                'periodicidad': per,
                'fecha_proximo_pago': fecha_proximo,
            })

        pids = list(sub_info_by_partner.keys())
        if pids:
            partners = o._call('res.partner', 'read', pids, ['id', 'id_noofit'])
            for p in partners:
                idn = p.get('id_noofit')
                if not idn:
                    continue
                idnoofits_con_sub.add(idn)
                sub_info_by_idn[str(idn)] = sub_info_by_partner.get(p['id'], [])
    except Exception as e:
        log.warning(f'informe_clientes odoo lookup: {e}')

    # Enriquecer cada fila
    out = []
    for r in rows:
        idn = str(r['id'])
        infos = sub_info_by_idn.get(idn, [])
        periodicidades = sorted({i['periodicidad'] for i in infos if i.get('periodicidad')})
        fechas_prox = sorted([i['fecha_proximo_pago'] for i in infos if i.get('fecha_proximo_pago')])
        out.append({
            'id': r['id'],
            'nombre': r['name'] or '',
            'apellidos': r['surname'] or '',
            'telefono': r.get('telefono') or '',
            'dni': r.get('dni') or '',
            'email': r.get('email') or '',
            'enabled': bool(r['enabled']),
            'estado': 'Activo' if r['enabled'] else 'Baja',
            'categoria': r.get('categoria_nombre') or '',
            'forma_pago': r.get('forma_pago') or '',
            'con_curso': idn in idnoofits_con_sub,
            'periodicidad': ', '.join(periodicidades),
            'fecha_proximo_pago': fechas_prox[0].isoformat() if fechas_prox else '',
            'descuentos': r.get('descuentos_codigos') or '',
            'tiene_modificacion': bool(r.get('tiene_modificacion')),
            'modificaciones_tipos': r.get('modificaciones_tipos') or '',
            'id_trainer': r.get('id_trainer'),
        })
    return out


@bp.route('/clientes', methods=['GET'])
@either_auth
def informe_clientes():
    """JSON con el listado. Filtros opcionales:
       ?trainer=<id>   ?categoria=<nombre>   ?solo_activos=1"""
    try:
        trainer = request.args.get('trainer') or None
        categoria = request.args.get('categoria') or None
        solo = request.args.get('solo_activos') == '1'
        rows = _build_rows(g.id_manager, trainer, categoria, solo)
        return jsonify({'ok': True, 'total': len(rows), 'rows': rows})
    except Exception as e:
        log.exception('informe_clientes')
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/clientes/excel', methods=['GET'])
@either_auth
def informe_clientes_excel():
    """Devuelve el listado en Excel con cabecera estilizada."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        trainer = request.args.get('trainer') or None
        categoria = request.args.get('categoria') or None
        solo = request.args.get('solo_activos') == '1'
        rows = _build_rows(g.id_manager, trainer, categoria, solo)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Clientes'

        # Resolver nombres de centros para mostrar bonito en la columna
        with get_conn() as conn, conn.cursor() as cur:
            cur.execute("""SELECT id_trainer, nombre_centro FROM centro_contacto
                            WHERE id_manager=%s""", (str(g.id_manager),))
            centros_map = {str(r['id_trainer']): r['nombre_centro'] for r in cur.fetchall()}

        # Cabecera
        cabecera = [
            'Nombre', 'Apellidos', 'Centro', 'Teléfono', 'DNI', 'Email',
            'Estado', 'Categoría', 'Forma de pago', 'Periodicidad', 'Próximo pago',
            'Curso', 'Descuentos', 'Modificación',
        ]
        thin = Side(style='thin', color='CCCCCC')
        for col, txt in enumerate(cabecera, 1):
            c = ws.cell(1, col, txt)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', start_color='2DD4A8')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ws.row_dimensions[1].height = 24

        # Datos
        for i, r in enumerate(rows, 2):
            ws.cell(i, 1, r['nombre'])
            ws.cell(i, 2, r['apellidos'])
            ws.cell(i, 3, centros_map.get(str(r.get('id_trainer') or ''), '') or '')
            ws.cell(i, 4, r['telefono'])
            ws.cell(i, 5, r['dni'])
            ws.cell(i, 6, r['email'])
            estado_cell = ws.cell(i, 7, r['estado'])
            if not r['enabled']:
                estado_cell.font = Font(color='C00000')
            ws.cell(i, 8, r['categoria'])
            ws.cell(i, 9, r['forma_pago'])
            ws.cell(i, 10, r.get('periodicidad') or '')
            ws.cell(i, 11, r.get('fecha_proximo_pago') or '')
            ws.cell(i, 12, 'Sí' if r['con_curso'] else 'No')
            ws.cell(i, 13, r['descuentos'])
            mod_cell = ws.cell(i, 14, r['modificaciones_tipos'] if r['tiene_modificacion'] else '')
            if r['tiene_modificacion']:
                mod_cell.font = Font(color='C46500')

        # Anchos de columna (incluye nuevas columnas Periodicidad y Próximo pago)
        for col, w in enumerate([18, 24, 20, 14, 13, 28, 10, 16, 16, 12, 12, 8, 32, 18], 1):
            ws.column_dimensions[chr(64+col)].width = w
        # Freeze cabecera + filtro auto
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        nombre = f'informe_clientes_{dt.date.today().isoformat()}.xlsx'
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=nombre)
    except Exception as e:
        log.exception('informe_clientes_excel')
        return jsonify({'ok': False, 'error': str(e)}), 500
