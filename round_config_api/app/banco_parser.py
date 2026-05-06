"""Parser de extractos bancarios.

Soporta:
  - CSV (separador detectado: , ; \t)
  - XLSX (vía openpyxl)
  - XLS (no soportado nativamente; pedimos que se exporte a XLSX o CSV)

Estrategia de detección de columnas (heurística):
  Buscamos en el header columnas con nombres típicos de bancos españoles:
    Fecha:    'fecha', 'fecha operacion', 'fecha valor', 'date', 'f. operacion'
    Concepto: 'concepto', 'descripcion', 'detalle', 'observaciones', 'descripción'
    Importe:  'importe', 'amount', 'cantidad', 'movimiento'
    Saldo:    'saldo', 'balance'
    Ref:      'referencia', 'reference', 'ref'

Si solo hay 'haber' / 'debe' (BBVA estilo), el importe es Haber-Debe.
Devuelve lista de dicts:
  {fecha (date), concepto (str), importe (float), saldo (float|None),
   ref_externa (str|None), banco (str|None)}

Si la detección falla, devuelve {} con error.
"""
import csv
import io
import logging
import re
from datetime import datetime, date
from pathlib import Path

log = logging.getLogger(__name__)

# Patrones de cabecera (case-insensitive, normalize accents)
HEADERS = {
    'fecha':    [r'^fecha$', r'^fecha\s+operaci[óo]n$', r'^fecha\s+valor$',
                 r'^date$', r'^f\.?\s*op', r'^f\.?\s*operaci[óo]n$',
                 r'^fecha\s+contable$'],
    'fecha_valor': [r'^fecha\s+valor$', r'^valor$', r'^value\s+date$'],
    'concepto': [r'^concepto$', r'^descripci[óo]n$', r'^detalle$',
                 r'^observaciones$', r'^description$', r'^memo$'],
    'importe':  [r'^importe$', r'^amount$', r'^cantidad$', r'^movimiento$',
                 r'^net$', r'^net\s+amount$'],
    'haber':    [r'^haber$', r'^abono$', r'^cr[ée]dito$', r'^credit$', r'^ingreso$'],
    'debe':     [r'^debe$', r'^cargo$', r'^d[ée]bito$', r'^debit$', r'^gasto$'],
    'saldo':    [r'^saldo$', r'^balance$', r'^saldo\s+despu[ée]s$'],
    'ref':      [r'^referencia$', r'^reference$', r'^ref$'],
}


def _norm(s: str) -> str:
    """Normaliza cadenas: lowercase + sin acentos."""
    if not s:
        return ''
    s = str(s).strip().lower()
    repl = (('á','a'),('é','e'),('í','i'),('ó','o'),('ú','u'),('ñ','n'))
    for a, b in repl:
        s = s.replace(a, b)
    return s


def _match_header(header: str) -> str:
    """Devuelve el campo (fecha|importe|...) si el header matchea algún patrón."""
    n = _norm(header)
    for field, patterns in HEADERS.items():
        for p in patterns:
            if re.match(p, n):
                return field
    return None


def _parse_date(s) -> date:
    """Parsea fecha en varios formatos comunes españoles."""
    if not s:
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    s = str(s).strip()
    if not s:
        return None
    fmts = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d.%m.%Y',
            '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S']
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except Exception:
            continue
    return None


def _parse_amount(s) -> float:
    """Parsea importe '1.234,56' o '1,234.56' o '-50,00'."""
    if s is None or s == '':
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip().replace('€', '').replace(' ', '').replace('\xa0', '')
    if not s:
        return None
    # Si tiene punto Y coma, asumimos formato europeo: punto millares, coma decimal
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        # Solo coma → decimal europeo
        s = s.replace(',', '.')
    # Quitar caracteres extraños excepto - y .
    s = re.sub(r'[^\d.\-]', '', s)
    try:
        return float(s)
    except Exception:
        return None


def _detect_columns(header_row: list) -> dict:
    """Mapea index → field. Devuelve {fecha: i, concepto: j, ...}."""
    cols = {}
    for i, h in enumerate(header_row):
        f = _match_header(h)
        if f and f not in cols:
            cols[f] = i
    return cols


def _ingredient_check(cols: dict) -> str:
    """Verifica que el mapping tenga lo mínimo. Devuelve None si OK, error si no."""
    if 'fecha' not in cols:
        return 'no_fecha'
    if 'importe' not in cols and not ('haber' in cols and 'debe' in cols):
        return 'no_importe_ni_haber_debe'
    if 'concepto' not in cols:
        return 'no_concepto'
    return None


def _row_to_dict(row: list, cols: dict) -> dict:
    """Convierte una fila (lista) a un dict de movimiento."""
    def get(f):
        idx = cols.get(f)
        if idx is None or idx >= len(row): return None
        return row[idx]

    importe = _parse_amount(get('importe'))
    if importe is None and ('haber' in cols or 'debe' in cols):
        h = _parse_amount(get('haber')) or 0
        d = _parse_amount(get('debe')) or 0
        importe = h - d if (h or d) else None

    return {
        'fecha':       _parse_date(get('fecha')),
        'fecha_valor': _parse_date(get('fecha_valor')),
        'concepto':    str(get('concepto') or '').strip(),
        'importe':     importe,
        'saldo':       _parse_amount(get('saldo')),
        'ref_externa': str(get('ref') or '').strip() or None,
    }


def _hash_dedupe(mov: dict) -> str:
    """Hash para evitar duplicar la misma línea al re-importar."""
    import hashlib
    raw = f"{mov.get('fecha')}|{mov.get('concepto','')}|{mov.get('importe')}|{mov.get('ref_externa','')}"
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()[:32]


# ── API pública ─────────────────────────────────────────────────────────────

def parse_csv(path: Path, encoding: str = 'utf-8-sig') -> dict:
    """Parsea CSV detectando delimitador y columnas. Devuelve {ok, rows, error}."""
    try:
        with open(path, 'r', encoding=encoding, errors='replace') as f:
            content = f.read()
    except Exception as e:
        return {'ok': False, 'error': f'read_csv: {e}'}
    # Detectar delimitador
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(content[:4096], delimiters=',;\t|')
    except csv.Error:
        # Fallback: ;
        dialect = csv.excel
        dialect.delimiter = ';'
    reader = csv.reader(io.StringIO(content), dialect=dialect)
    rows = list(reader)
    if not rows:
        return {'ok': False, 'error': 'archivo_vacio'}

    # Buscar la fila de cabecera: la primera con ≥2 campos donde podamos
    # detectar al menos uno de nuestros campos típicos
    header_idx = -1
    cols = {}
    for i, r in enumerate(rows[:30]):
        if len(r) < 2: continue
        c = _detect_columns(r)
        if c.get('fecha') is not None and (c.get('importe') is not None or
                                            ('haber' in c and 'debe' in c)):
            header_idx = i
            cols = c
            break
    if header_idx < 0:
        return {'ok': False, 'error': 'no_se_detectan_columnas',
                'detalle': f'header sniff failed; first row was: {rows[0][:6]}'}
    err = _ingredient_check(cols)
    if err:
        return {'ok': False, 'error': err, 'columnas_detectadas': list(cols.keys())}

    # Parsear todas las filas tras header
    out = []
    for r in rows[header_idx + 1:]:
        if not r or all(not str(x).strip() for x in r): continue
        mov = _row_to_dict(r, cols)
        if mov['fecha'] and mov['importe'] is not None and mov['concepto']:
            out.append(mov)
    return {'ok': True, 'rows': out, 'columnas_detectadas': cols, 'total': len(out)}


def parse_xlsx(path: Path) -> dict:
    """Parsea XLSX con openpyxl."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return {'ok': False, 'error': f'read_xlsx: {e}'}
    if not rows:
        return {'ok': False, 'error': 'archivo_vacio'}

    header_idx = -1
    cols = {}
    for i, r in enumerate(rows[:30]):
        if not r or len(r) < 2: continue
        clean = [str(c) if c is not None else '' for c in r]
        c = _detect_columns(clean)
        if c.get('fecha') is not None and (c.get('importe') is not None or
                                            ('haber' in c and 'debe' in c)):
            header_idx = i
            cols = c
            break
    if header_idx < 0:
        return {'ok': False, 'error': 'no_se_detectan_columnas'}
    err = _ingredient_check(cols)
    if err:
        return {'ok': False, 'error': err, 'columnas_detectadas': list(cols.keys())}

    out = []
    for r in rows[header_idx + 1:]:
        if not r or all(c is None or str(c).strip() == '' for c in r): continue
        mov = _row_to_dict(list(r), cols)
        if mov['fecha'] and mov['importe'] is not None and mov['concepto']:
            out.append(mov)
    return {'ok': True, 'rows': out, 'columnas_detectadas': cols, 'total': len(out)}


def parse_extracto(path: Path) -> dict:
    """Detecta el formato por extensión y delega al parser adecuado."""
    ext = path.suffix.lower()
    if ext == '.csv' or ext == '.txt':
        # Intentar utf-8 primero, luego latin-1
        r = parse_csv(path, encoding='utf-8-sig')
        if not r.get('ok') and r.get('error', '').startswith('no_se'):
            r = parse_csv(path, encoding='latin-1')
        return r
    if ext == '.xlsx':
        return parse_xlsx(path)
    if ext == '.xls':
        return {'ok': False,
                'error': 'formato_xls_no_soportado',
                'detalle': 'Por favor exporta el extracto a XLSX o CSV.'}
    return {'ok': False, 'error': f'extension_no_soportada:{ext}'}


def attach_dedupe_hashes(rows: list) -> list:
    """Añade hash_dedupe a cada fila para evitar duplicados al re-importar."""
    for r in rows:
        r['hash_dedupe'] = _hash_dedupe(r)
    return rows
