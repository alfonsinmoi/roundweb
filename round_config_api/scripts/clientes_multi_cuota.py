"""Detectar clientes GestPlus con MÁS DE UNA cuota activa en 2026.

Una "cuota" se identifica por el campo `codcur` de los recibos.
- Si un cliente tiene >1 codcur distinto en el año 2026 → múltiples cuotas
- Detectamos solapamientos (mismo período con dos codcur)

Salida: /tmp/clientes_multi_cuota.xlsx con hojas:
  - RESUMEN: contadores y stats
  - MULTI_CUOTA: 1 fila por cliente con resumen + lista de cuotas
  - DETALLE: 1 fila por (cliente, codcur) con su histórico
  - SOLAPAMIENTOS: clientes con 2 codcur en el mismo mes (lo más sospechoso)
"""
import json, sys, re, unicodedata
from datetime import datetime
from collections import defaultdict, Counter
sys.path.insert(0, '/opt/round_config_api')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GP = '/opt/round_config_api/gestplus_dump_2026-05-08.json'
OUT = '/tmp/clientes_multi_cuota.xlsx'
TARGET_YEAR = 2026


def parse_date(s):
    if not s: return None
    try: return datetime.fromisoformat(str(s).replace('Z', '')).date()
    except Exception:
        try: return datetime.strptime(str(s)[:10], '%Y-%m-%d').date()
        except Exception: return None


def detectar_periodicidad(fd, fh):
    if not fd or not fh: return 'desconocida'
    dias = (fh - fd).days + 1
    if 25 <= dias <= 35: return 'mensual'
    if 80 <= dias <= 100: return 'trimestral'
    if 150 <= dias <= 200: return 'semestral'
    if 300 <= dias <= 400: return 'anual'
    return f'otro({dias}d)'


# ─── Cargar dump ─────────────────────────────────────────────────────────────
gp = json.load(open(GP, 'r', encoding='utf-8'))
gp_clientes = (gp.get('altas') or []) + (gp.get('bajas_recientes_12m') or [])
print(f'Clientes GP universo: {len(gp_clientes)}')

# Agrupar recibos del 2026 por (cliente, codcur)
multi = []        # clientes con >1 cuota
detalle = []      # 1 fila por (cliente, codcur)
solapados = []    # clientes con codcur en el mismo mes

for c in gp_clientes:
    recibos_2026 = []
    for r in c.get('_recibos', []):
        fd = parse_date(r.get('fechaDesde'))
        if fd and fd.year == TARGET_YEAR:
            recibos_2026.append({**r, '_fd': fd, '_fh': parse_date(r.get('fechaHasta'))})
    if not recibos_2026: continue

    # Agrupar por codcur
    por_codcur = defaultdict(list)
    for r in recibos_2026:
        por_codcur[r.get('codcur') or '(sin código)'].append(r)
    n_cuotas = len(por_codcur)

    # Detectar solapamientos: ¿hay 2 codcur distintos en el mismo mes?
    por_mes = defaultdict(set)   # mes -> set(codcur)
    for r in recibos_2026:
        if r['_fd']:
            mes = r['_fd'].strftime('%Y-%m')
            por_mes[mes].add(r.get('codcur') or '(sin código)')
    meses_con_solape = {mes: codcurs for mes, codcurs in por_mes.items() if len(codcurs) > 1}

    if n_cuotas > 1:
        cuotas_summary = []
        for codcur, lst in por_codcur.items():
            primer = min(r['_fd'] for r in lst if r['_fd'])
            ultimo = max(r['_fd'] for r in lst if r['_fd'])
            importes = [float(r.get('importeFinal') or 0) for r in lst]
            importe_medio = sum(importes) / len(importes) if importes else 0
            periodicidades = Counter(detectar_periodicidad(r['_fd'], r['_fh']) for r in lst)
            periodicidad = periodicidades.most_common(1)[0][0]
            cuotas_summary.append({
                'codcur': codcur, 'recibos': len(lst),
                'primer_mes': primer.strftime('%Y-%m'), 'ultimo_mes': ultimo.strftime('%Y-%m'),
                'importe_medio': round(importe_medio, 2),
                'periodicidad': periodicidad,
            })
            detalle.append({
                'codcli_gp': c.get('codigo'),
                'cliente': f"{c.get('nombre','')} {c.get('apellidos','')}".strip(),
                'dni': c.get('dni') or '',
                'email': c.get('email') or '',
                'estado_cliente': 'Alta' if c.get('estado') == 1 else 'Baja',
                'codcur': codcur,
                'recibos': len(lst),
                'periodicidad': periodicidad,
                'primer_mes': primer.strftime('%Y-%m'),
                'ultimo_mes': ultimo.strftime('%Y-%m'),
                'importe_medio': round(importe_medio, 2),
                'cobrados': sum(1 for r in lst if r.get('cobrado') == 1),
                'impagados': sum(1 for r in lst if r.get('cobrado') != 1),
                'importe_total': round(sum(importes), 2),
            })
        # Ordenar cuotas por nº de recibos descendente
        cuotas_summary.sort(key=lambda x: -x['recibos'])
        multi.append({
            'codcli': c.get('codigo'),
            'cliente': f"{c.get('nombre','')} {c.get('apellidos','')}".strip(),
            'dni': c.get('dni') or '',
            'email': c.get('email') or '',
            'estado_cliente': 'Alta' if c.get('estado') == 1 else 'Baja',
            'n_cuotas': n_cuotas,
            'codcurs': ', '.join(s['codcur'] for s in cuotas_summary),
            'cuotas_summary': cuotas_summary,
            'meses_solape': len(meses_con_solape),
            'detalle_solape': '; '.join(f"{m}={'+'.join(c)}" for m, c in sorted(meses_con_solape.items())),
            'importe_total': round(sum(s['importe_medio'] * s['recibos'] for s in cuotas_summary), 2),
        })
        if meses_con_solape:
            solapados.append(multi[-1])

print(f'Clientes con >1 cuota en 2026: {len(multi)}')
print(f'Clientes con SOLAPAMIENTO (2 codcur mismo mes): {len(solapados)}')


# ─── Excel ───────────────────────────────────────────────────────────────────
wb = Workbook(); wb.remove(wb.active)
border = Border(*[Side(style='thin', color='CCCCCC')]*4)


def hsheet(name, headers, color='2DD4A8'):
    ws = wb.create_sheet(name)
    fill_h = PatternFill('solid', fgColor=color)
    font_h = Font(bold=True, color='FFFFFF', size=11)
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.fill = fill_h; c.font = font_h
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
    return ws


# RESUMEN
ws = wb.create_sheet('RESUMEN', 0)
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 60
ws.cell(1, 1, f'Clientes con >1 cuota activa en {TARGET_YEAR}').font = Font(bold=True, size=14)

def w_kv(row, k, v, note=''):
    ws.cell(row, 1, k).font = Font(bold=True)
    ws.cell(row, 2, v)
    if note: ws.cell(row, 3, note).font = Font(italic=True, color='888888')

w_kv(3, 'Clientes con >1 cuota',           len(multi))
w_kv(4, '  con solapamiento (mismo mes)',  len(solapados), '⚠️ probable error o cargos puntuales')
w_kv(5, '  cambio de plan (sin solape)',   len(multi) - len(solapados), 'pasaron de un plan a otro a mitad de año')

# Distribución por nº de cuotas
n_dist = Counter(x['n_cuotas'] for x in multi)
w_kv(7, 'DISTRIBUCIÓN', '')
ws.cell(7, 1).font = Font(bold=True)
row = 8
for n in sorted(n_dist):
    w_kv(row, f'  {n} cuotas', n_dist[n])
    row += 1


# MULTI_CUOTA (1 fila por cliente)
ws_m = hsheet('MULTI_CUOTA', [
    ('Cód GP', 9), ('Cliente', 30), ('DNI', 14), ('Email', 26),
    ('Estado', 9), ('Nº cuotas', 10), ('Códigos cuota', 30),
    ('Solape (meses)', 14), ('Detalle solape', 50),
    ('Importe total', 13),
], color='5B9CF6')
multi.sort(key=lambda x: (-x['meses_solape'], -x['n_cuotas'], x['cliente'].upper()))
for i, x in enumerate(multi, 2):
    vals = [
        x['codcli'], x['cliente'], x['dni'], x['email'],
        x['estado_cliente'], x['n_cuotas'], x['codcurs'],
        x['meses_solape'], x['detalle_solape'],
        f'{x["importe_total"]:.2f}',
    ]
    for j, v in enumerate(vals, 1):
        c = ws_m.cell(i, j, v); c.border = border
        if x['meses_solape'] > 0 and j == 8:
            c.fill = PatternFill('solid', fgColor='FFE4E4')
ws_m.auto_filter.ref = ws_m.dimensions


# DETALLE (1 fila por cuota de cada cliente)
ws_d = hsheet('DETALLE_CUOTAS', [
    ('Cód GP', 9), ('Cliente', 30), ('DNI', 14), ('Email', 24),
    ('Estado', 9), ('Codcur', 14), ('Nº recibos', 10),
    ('Periodicidad', 13), ('Primer mes', 11), ('Último mes', 11),
    ('Importe medio', 12), ('Cobrados', 9), ('Impagados', 9),
    ('Importe total', 12),
], color='2DD4A8')
detalle.sort(key=lambda x: (x['cliente'].upper(), x['codcur']))
for i, x in enumerate(detalle, 2):
    vals = [
        x['codcli_gp'], x['cliente'], x['dni'], x['email'],
        x['estado_cliente'], x['codcur'], x['recibos'],
        x['periodicidad'], x['primer_mes'], x['ultimo_mes'],
        f'{x["importe_medio"]:.2f}', x['cobrados'], x['impagados'],
        f'{x["importe_total"]:.2f}',
    ]
    for j, v in enumerate(vals, 1):
        ws_d.cell(i, j, v).border = border
ws_d.auto_filter.ref = ws_d.dimensions


# SOLAPAMIENTOS
ws_s = hsheet('SOLAPAMIENTOS', [
    ('Cód GP', 9), ('Cliente', 30), ('DNI', 14), ('Email', 26),
    ('Estado', 9), ('Nº cuotas', 10), ('Códigos cuota', 30),
    ('Meses con solape', 14), ('Detalle solape', 60),
    ('Importe total', 13),
], color='F87171')
for i, x in enumerate(solapados, 2):
    vals = [
        x['codcli'], x['cliente'], x['dni'], x['email'],
        x['estado_cliente'], x['n_cuotas'], x['codcurs'],
        x['meses_solape'], x['detalle_solape'],
        f'{x["importe_total"]:.2f}',
    ]
    for j, v in enumerate(vals, 1):
        ws_s.cell(i, j, v).border = border
ws_s.auto_filter.ref = ws_s.dimensions

wb.save(OUT)
print(f'\n→ {OUT}')
print(f'  RESUMEN | MULTI_CUOTA ({len(multi)}) | DETALLE_CUOTAS ({len(detalle)}) | SOLAPAMIENTOS ({len(solapados)})')
