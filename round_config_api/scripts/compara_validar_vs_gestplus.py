"""Compara importe total por cliente entre la validación Round (junio 2026) y
el listado GestPlus. Match por DNI normalizado."""
import re
import unicodedata as ud
from openpyxl import load_workbook
from app.routes.preemision_validar import _validar_emision

GP_PATH = '/tmp/gp.xlsx'

def norm_dni(d):
    s = ''.join(c for c in ud.normalize('NFD', str(d or ''))
                if ud.category(c) != 'Mn')
    return re.sub(r'[^A-Z0-9]', '', s.upper())

# 1) Cargar GestPlus
gp_wb = load_workbook(GP_PATH)
gp_ws = gp_wb.active
gp_heads = [gp_ws.cell(1, c).value for c in range(1, gp_ws.max_column+1)]
col_dni = gp_heads.index('dniContr') + 1
col_final = gp_heads.index('Final') + 1
col_nombre = gp_heads.index('Nombre') + 1
col_codcur = gp_heads.index('codcur') + 1
col_modif = gp_heads.index('Modif') + 1
col_desc = gp_heads.index('descu') + 1

gp_por_dni = {}
for r in range(2, gp_ws.max_row+1):
    dni = norm_dni(gp_ws.cell(r, col_dni).value)
    if not dni: continue
    importe = float(gp_ws.cell(r, col_final).value or 0)
    nombre = gp_ws.cell(r, col_nombre).value or ''
    cur = gp_ws.cell(r, col_codcur).value or ''
    modif = float(gp_ws.cell(r, col_modif).value or 0)
    desc = gp_ws.cell(r, col_desc).value or ''
    # Si un cliente tiene MÚLTIPLES recibos en GP (varias cuotas), sumamos
    if dni in gp_por_dni:
        gp_por_dni[dni]['importe'] += importe
        gp_por_dni[dni]['cursos'].append(cur)
        gp_por_dni[dni]['modif'] += modif
    else:
        gp_por_dni[dni] = {'nombre': nombre, 'importe': importe,
                            'cursos': [cur], 'modif': modif, 'desc': desc}
print(f'GP: {len(gp_por_dni)} clientes únicos (por DNI) — '
      f'{gp_ws.max_row-1} recibos totales')

# 2) Cargar coherentes de validación Round
coherentes, incoherencias, resumen = _validar_emision('17675', '2026-06')
print(f'Round: {len(coherentes)} coherentes en validación')

round_por_dni = {}
for c in coherentes:
    dni = norm_dni(c.get('dni') or '')
    if not dni: continue
    round_por_dni[dni] = c

# 3) Match + diferencias
ambos = set(gp_por_dni) & set(round_por_dni)
solo_gp = set(gp_por_dni) - set(round_por_dni)
solo_round = set(round_por_dni) - set(gp_por_dni)
print(f'\n=== Cruce por DNI ===')
print(f'  en ambos:    {len(ambos)}')
print(f'  solo GP:     {len(solo_gp)}')
print(f'  solo Round:  {len(solo_round)}')

# Diferencias de importe
print(f'\n=== Diferencias de importe (|Δ| ≥ 0.01€) en los {len(ambos)} comunes ===')
diffs = []
for dni in ambos:
    gp = gp_por_dni[dni]
    rd = round_por_dni[dni]
    gp_imp = gp['importe']
    rd_imp = float(rd.get('importe_total') or 0)
    delta = round(rd_imp - gp_imp, 2)
    if abs(delta) >= 0.01:
        diffs.append((dni, gp['nombre'], rd.get('nombre'), gp_imp, rd_imp, delta,
                      '+'.join(gp['cursos']), '+'.join(rd.get('cuotas') or [])))
diffs.sort(key=lambda x: abs(x[5]), reverse=True)
print(f'Total con diferencias: {len(diffs)}')
print(f'Total cuadran:         {len(ambos) - len(diffs)}')
print()
print(f'{"DNI":<11}  {"Nombre GP":<28}  {"GP":>8}  {"Round":>8}  {"Δ":>7}  cursos GP / cuotas Round')
for d in diffs[:60]:
    dni, nom_gp, nom_rd, gp_imp, rd_imp, delta, c_gp, c_rd = d
    print(f'{dni:<11}  {(nom_gp or "")[:28]:<28}  {gp_imp:>8.2f}  {rd_imp:>8.2f}  {delta:>+7.2f}  {c_gp} / {c_rd}')

# Solo GP (no aparecen en Round)
print(f'\n=== Solo en GP (no se va a emitir en Round): {len(solo_gp)} ===')
for dni in sorted(solo_gp)[:30]:
    gp = gp_por_dni[dni]
    print(f'  {dni:<11}  {gp["nombre"][:35]:<35}  {gp["importe"]:>7.2f}  {",".join(gp["cursos"])}')

# Solo Round (no aparecen en GP)
print(f'\n=== Solo en Round (no aparecen en GP): {len(solo_round)} ===')
for dni in sorted(solo_round)[:30]:
    rd = round_por_dni[dni]
    print(f'  {dni:<11}  {(rd.get("nombre") or "")[:35]:<35}  '
          f'{float(rd.get("importe_total") or 0):>7.2f}  {",".join(rd.get("cuotas") or [])}')
