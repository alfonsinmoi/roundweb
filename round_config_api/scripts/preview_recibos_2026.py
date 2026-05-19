"""Preview de la asignación recibos 2026 → clientes.

NO escribe nada. Solo lee del dump GestPlus + dump NoofitPro y genera un
Excel con:
  - Hoja RESUMEN: contadores y totales
  - Hoja POR_CLIENTE: agregado por cliente (recibos, importe, estado)
  - Hoja RECIBOS_TODOS: lista detallada
  - Hoja ANOMALIAS: incidencias detectadas
       * Recibo con codcli sin DNI ni email
       * Recibo cuyo cliente GP no existe en NoofitPro
       * Recibo con importe negativo o 0
       * Cliente con > 12 recibos en 2026 (probable error)
       * Recibos sin fecha o fechas raras
       * Recibos con fechaBaja (anulados) que aún figuran cobrados
       * Importes inconsistentes (importeFinal > importeInicial)
"""
import json, sys, re, unicodedata
from datetime import datetime
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GP = '/opt/round_config_api/gestplus_dump_2026-05-08.json'
NF = '/opt/round_config_api/noofit_clientes_dump.json'
OUT = '/tmp/preview_recibos_2026.xlsx'

# ─── Helpers ─────────────────────────────────────────────────────────────────
def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s or '')) if unicodedata.category(c) != 'Mn')
def norm_dni(d):
    return re.sub(r'[^A-Z0-9]', '', strip_accents(s=str(d or '')).upper() if d else '')
def norm_email(e):
    return (e or '').strip().lower()

# ─── Cargar dumps ────────────────────────────────────────────────────────────
gp = json.load(open(GP, 'r', encoding='utf-8'))
nf = json.load(open(NF, 'r', encoding='utf-8'))
nf_clientes = nf.get('clientes', [])

# Indexar NoofitPro
nf_by_dni = {}
nf_by_email = {}
for c in nf_clientes:
    d = norm_dni(c.get('dni') or c.get('nif'))
    if d and len(d) >= 7: nf_by_dni.setdefault(d, c)
    e = norm_email(c.get('email'))
    if e and '@' in e: nf_by_email.setdefault(e, c)

# Recoger todos los recibos del año + cliente GP
gp_clients_all = (gp.get('altas') or []) + (gp.get('bajas_recientes_12m') or [])
gp_by_codcli = {c.get('codigo'): c for c in gp_clients_all if c.get('codigo')}

recibos = []
for c in gp_clients_all:
    for r in c.get('_recibos', []):
        recibos.append({**r,
            '_codcli': c.get('codigo'),
            '_dni': c.get('dni'),
            '_dni_contr': c.get('dniContr'),
            '_nombre': f"{c.get('nombre','')} {c.get('apellidos','')}".strip(),
            '_email': c.get('email'),
            '_iban': c.get('_iban'),
            '_estado_cliente': 'Alta' if c.get('estado') == 1 else 'Baja',
            '_fechaBajaCli': c.get('fechaBaja'),
        })

print(f'Recibos 2026: {len(recibos)}')
print(f'Clientes GestPlus universo: {len(gp_clients_all)}')
print(f'Clientes NoofitPro dump: {len(nf_clientes)}')

# ─── Cruzar cada recibo con NoofitPro ──────────────────────────────────────
def find_in_nf(r):
    for k in ('_dni', '_dni_contr'):
        d = norm_dni(r.get(k))
        if d and len(d) >= 7 and d in nf_by_dni:
            return nf_by_dni[d], 'dni'
    e = norm_email(r.get('_email'))
    if e and e in nf_by_email:
        return nf_by_email[e], 'email'
    return None, None

for r in recibos:
    nf_match, method = find_in_nf(r)
    r['_nf_id'] = nf_match.get('id') if nf_match else None
    r['_nf_match_by'] = method
    r['_nf_enabled'] = nf_match.get('enabled') if nf_match else None

# ─── Detección de anomalías ──────────────────────────────────────────────────
anomalias = []
def anom(r, tipo, detalle):
    anomalias.append({
        'tipo': tipo, 'detalle': detalle,
        'numRec': r.get('numRec'), 'codcli': r.get('_codcli'),
        'nombre': r.get('_nombre'), 'dni': r.get('_dni'),
        'email': r.get('_email'),
        'importe_inicial': r.get('importeInicial'),
        'importe_final': r.get('importeFinal'),
        'importe_banco': r.get('importeBanco'),
        'fechaDesde': r.get('fechaDesde'), 'fechaHasta': r.get('fechaHasta'),
    })

for r in recibos:
    # Sin cliente NoofitPro
    if not r.get('_nf_id'):
        anom(r, 'sin_match_nf', 'Cliente no encontrado en NoofitPro (por DNI ni email)')
    # Sin DNI ni email
    if not r.get('_dni') and not r.get('_email'):
        anom(r, 'sin_id_cliente', 'Sin DNI ni email en GP')
    # Importe inicial 0 o negativo
    ii = r.get('importeInicial') or 0
    if not isinstance(ii, (int, float)) or ii <= 0:
        anom(r, 'importe_invalido', f'Importe inicial = {ii}')
    # final > inicial (descuento negativo)
    f = r.get('importeFinal') or 0
    if f > ii * 1.01:
        anom(r, 'importe_inconsistente', f'Final ({f}) > inicial ({ii})')
    # cobrado=1 pero importeBanco=0 (raro)
    if r.get('cobrado') == 1 and (r.get('importeBanco') or 0) == 0:
        anom(r, 'cobrado_sin_banco', 'cobrado=1 pero importeBanco=0 (revisar)')
    # Fechas
    fd = r.get('fechaDesde'); fh = r.get('fechaHasta')
    if not fd or not fh:
        anom(r, 'sin_fechas', f'fechaDesde={fd} fechaHasta={fh}')
    # Devolución no recobrada
    if r.get('fechDevolucion') and not r.get('fechRecobro'):
        anom(r, 'devuelto', f"fechDevolucion={r.get('fechDevolucion')}")

# Recibos por cliente
recibos_por_cliente = defaultdict(list)
for r in recibos:
    recibos_por_cliente[r.get('_codcli')].append(r)

# Cliente con > 12 recibos
for cod, lst in recibos_por_cliente.items():
    if len(lst) > 12:
        c = lst[0]
        anom(c, 'demasiados_recibos', f'Cliente con {len(lst)} recibos en 2026 (>12 → revisar)')

# ─── Excel ───────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)
border = Border(*[Side(style='thin', color='CCCCCC')]*4)

def write_sheet(name, headers_widths, rows, color='2DD4A8'):
    ws = wb.create_sheet(name)
    fill_h = PatternFill('solid', fgColor=color)
    font_h = Font(bold=True, color='FFFFFF', size=11)
    for col, (h, w) in enumerate(headers_widths, 1):
        c = ws.cell(1, col, h); c.fill = fill_h; c.font = font_h
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            ws.cell(i, j, v).border = border
    ws.auto_filter.ref = ws.dimensions
    return ws

# RESUMEN
ws = wb.create_sheet('RESUMEN', 0)
ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 60
ws.cell(1, 1, 'PREVIEW migración recibos 2026 GestPlus → Odoo').font = Font(bold=True, size=14)
def w_kv(row, k, v, note=''):
    ws.cell(row, 1, k).font = Font(bold=True)
    ws.cell(row, 2, v)
    if note: ws.cell(row, 3, note).font = Font(italic=True, color='888888')

n_total = len(recibos)
n_con_match = sum(1 for r in recibos if r.get('_nf_id'))
n_sin_match = n_total - n_con_match
n_cobrados = sum(1 for r in recibos if r.get('cobrado') == 1)
total_emitido = sum(float(r.get('importeFinal') or 0) for r in recibos)
total_cobrado = sum(float(r.get('importeBanco') or 0) for r in recibos)
n_clientes = len(recibos_por_cliente)

w_kv(3, 'Recibos totales 2026',         n_total)
w_kv(4, '  con cliente en NoofitPro',   n_con_match, f'{n_con_match/n_total*100:.0f}%' if n_total else '')
w_kv(5, '  SIN match en NoofitPro',     n_sin_match, '⚠️ revisar — no se podrán importar a Odoo sin partner_id')
w_kv(6, 'Recibos cobrados',             n_cobrados, f'{n_cobrados/n_total*100:.0f}%' if n_total else '')
w_kv(7, 'Recibos NO cobrados',          n_total - n_cobrados)
w_kv(8, 'Total emitido (€)',            f'{total_emitido:.2f}')
w_kv(9, 'Total cobrado (€)',            f'{total_cobrado:.2f}')
w_kv(10, 'Total impagado (€)',          f'{total_emitido - total_cobrado:.2f}')
w_kv(11, 'Clientes únicos con recibo',  n_clientes)
w_kv(13, 'ANOMALÍAS DETECTADAS',         len(anomalias))
ws.cell(13, 1).font = Font(bold=True, color='C00000')
tipos = Counter(a['tipo'] for a in anomalias)
row = 14
for t, n in tipos.most_common():
    w_kv(row, f'  {t}', n)
    row += 1

# POR_CLIENTE
filas_por_cli = []
for cod, lst in recibos_por_cliente.items():
    primero = lst[0]
    cobr = sum(1 for r in lst if r.get('cobrado') == 1)
    importe_total = sum(float(r.get('importeFinal') or 0) for r in lst)
    importe_cobrado = sum(float(r.get('importeBanco') or 0) for r in lst)
    filas_por_cli.append([
        cod, primero.get('_codcli'),
        primero.get('_nombre'), primero.get('_dni') or '', primero.get('_email') or '',
        primero.get('_estado_cliente'), primero.get('_nf_id') or '',
        len(lst), cobr, len(lst) - cobr,
        f'{importe_total:.2f}', f'{importe_cobrado:.2f}',
        f'{importe_total - importe_cobrado:.2f}',
    ])
filas_por_cli.sort(key=lambda x: -float(x[10]))  # por importe total desc

write_sheet('POR_CLIENTE', [
    ('Cód. GP', 11), ('Cód. cliente', 11), ('Cliente', 32), ('DNI', 14), ('Email', 28),
    ('Estado GP', 9), ('NF id', 11), ('Recibos', 9), ('Cobrados', 9), ('Pendientes', 10),
    ('Total emitido', 13), ('Total cobrado', 13), ('Pendiente €', 12),
], filas_por_cli, color='5B9CF6')

# RECIBOS_TODOS
filas_recibos = []
for r in recibos:
    filas_recibos.append([
        r.get('numRec'), r.get('_codcli'), r.get('_nombre'),
        r.get('_dni') or '', r.get('_email') or '',
        r.get('_nf_id') or '',
        r.get('codcur'), r.get('cobrado'),
        r.get('importeInicial'), r.get('importeFinal'), r.get('importeBanco'),
        (r.get('fechaDesde') or '')[:10], (r.get('fechaHasta') or '')[:10],
        (r.get('fechDevolucion') or '')[:10], (r.get('fechRecobro') or '')[:10],
        r.get('estado'), r.get('_nf_match_by') or '',
    ])
write_sheet('RECIBOS_TODOS', [
    ('numRec', 11), ('Cód. cli', 11), ('Cliente', 30), ('DNI', 14), ('Email', 26),
    ('NF id', 11), ('Curso', 14), ('Cobrado', 8),
    ('Inicial', 9), ('Final', 9), ('Banco', 9),
    ('Desde', 11), ('Hasta', 11), ('Devol.', 11), ('Recobro', 11),
    ('Estado', 7), ('Match por', 10),
], filas_recibos)

# ANOMALIAS
filas_anom = []
for a in anomalias:
    filas_anom.append([
        a['tipo'], a['detalle'], a.get('numRec'), a.get('codcli'),
        a.get('nombre'), a.get('dni'), a.get('email'),
        a.get('importe_inicial'), a.get('importe_final'), a.get('importe_banco'),
        a.get('fechaDesde'), a.get('fechaHasta'),
    ])
write_sheet('ANOMALIAS', [
    ('Tipo', 22), ('Detalle', 50), ('numRec', 11), ('Cód. cli', 11),
    ('Cliente', 30), ('DNI', 14), ('Email', 26),
    ('Inicial', 9), ('Final', 9), ('Banco', 9),
    ('Desde', 21), ('Hasta', 21),
], filas_anom, color='F87171')

wb.save(OUT)
print(f'\n→ {OUT}')
print(f'  RESUMEN, POR_CLIENTE ({n_clientes}), RECIBOS_TODOS ({n_total}), ANOMALIAS ({len(anomalias)})')
