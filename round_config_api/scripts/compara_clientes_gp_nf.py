"""Comparativa de datos cliente GestPlus ↔ NoofitPro ↔ Round (cuotas).

Cada fila = un cliente. Compara:
  - DNI (normalizado)
  - Nombre + Apellidos
  - Email
  - Teléfono / Móvil
  - Dirección
  - Fecha alta
  - IBAN (solo en GP — informativo)
  - Forma de pago GP (efectivo / banco / etc) — informativo
  - Cuota inferida de los recibos GP del año
  - Periodicidad inferida (mensual / trimestral / etc)
  - Cuota Round (asignada en BD round_config.cuota_cliente)

Las celdas con discrepancia se resaltan en rojo. Filas SOLO_GP y SOLO_NF
también se incluyen.

Salida: /tmp/compara_clientes_gp_nf.xlsx con hojas:
  - RESUMEN: contadores y leyenda
  - COMPARATIVA: lista completa con todas las columnas
  - SOLO_NF: clientes en NoofitPro que NO están en GestPlus
  - DISCREPANCIAS: solo filas con al menos un campo distinto
"""
import json, sys, re, unicodedata
from datetime import datetime
from collections import defaultdict, Counter
sys.path.insert(0, '/opt/round_config_api')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.db import get_conn

GP = '/opt/round_config_api/gestplus_dump_2026-05-08.json'
NF = '/opt/round_config_api/noofit_clientes_dump.json'
OUT = '/tmp/compara_clientes_gp_nf.xlsx'
ID_MANAGER = '17675'  # para cuota_cliente lookup

# ─── Helpers ────────────────────────────────────────────────────────────────
def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s or '')) if unicodedata.category(c) != 'Mn')
def norm_dni(d):
    s = strip_accents(str(d or '')).upper()
    return re.sub(r'[^A-Z0-9]', '', s)
def norm_email(e): return (e or '').strip().lower()
def norm_phone(p):
    if not p: return ''
    return re.sub(r'[^0-9]', '', str(p))[-9:]   # últimos 9 dígitos
def norm_text(s): return strip_accents(str(s or '')).upper().strip()
def norm_iban(i):
    s = re.sub(r'[^A-Z0-9]', '', str(i or '').upper())
    return s if s and not s.startswith('ES000') else ''
def parse_date(s):
    if not s: return None
    try: return datetime.fromisoformat(str(s).replace('Z', '')).date()
    except Exception:
        try: return datetime.strptime(str(s)[:10], '%Y-%m-%d').date()
        except Exception: return None

def fmt_date(d):
    return d.isoformat() if d else ''


def detectar_periodicidad(fd, fh):
    if not fd or not fh: return None
    dias = (fh - fd).days + 1
    if 25 <= dias <= 35: return 'mensual'
    if 80 <= dias <= 100: return 'trimestral'
    if 150 <= dias <= 200: return 'semestral'
    if 300 <= dias <= 400: return 'anual'
    return f'otro({dias}d)'

# Forma de pago en GP: códigos típicos
FORMA_PAGO_GP = {
    'B': 'banco', 'C': 'caja', 'D': 'domiciliado', 'E': 'efectivo',
    'T': 'tarjeta', 'V': 'transferencia',
}

# ─── Cargar dumps ─────────────────────────────────────────────────────────────
gp = json.load(open(GP, 'r', encoding='utf-8'))
nf = json.load(open(NF, 'r', encoding='utf-8'))
gp_clientes = (gp.get('altas') or []) + (gp.get('bajas_recientes_12m') or [])
nf_clientes = nf.get('clientes', [])
print(f'GP universo: {len(gp_clientes)}  ·  NF dump: {len(nf_clientes)}')

# cuota_cliente no existe en BD (la cuota la mantiene NoofitPro / Odoo).
# Comparamos solo GP vs NF + lo que infiramos de los recibos.
cuota_by_idnoofit = {}

# ─── Indexar NF por DNI / email / teléfono / nombre ──────────────────────────
nf_by_dni, nf_by_email, nf_by_phone = {}, {}, {}
for c in nf_clientes:
    d = norm_dni(c.get('dni') or c.get('nif'))
    if d and len(d) >= 7: nf_by_dni.setdefault(d, c)
    e = norm_email(c.get('email'))
    if e and '@' in e: nf_by_email.setdefault(e, c)
    p = norm_phone(c.get('cellPhone') or c.get('telefono'))
    if p and len(p) >= 7: nf_by_phone.setdefault(p, c)

# ─── Construir filas comparativa ──────────────────────────────────────────────
def find_nf(g):
    for k in ('dni', 'dniContr'):
        d = norm_dni(g.get(k))
        if d and len(d) >= 7 and d in nf_by_dni:
            return nf_by_dni[d], 'dni'
    e = norm_email(g.get('email'))
    if e and e in nf_by_email:
        return nf_by_email[e], 'email'
    p = norm_phone(g.get('movil') or g.get('telefono'))
    if p and len(p) >= 7 and p in nf_by_phone:
        return nf_by_phone[p], 'phone'
    return None, None


def cuota_inferida_de_recibos(g):
    """Devuelve (codcur_dominante, periodicidad_dominante, importe_medio)."""
    rs = g.get('_recibos') or []
    if not rs: return ('', '', '')
    codcurs = Counter(r.get('codcur') for r in rs if r.get('codcur'))
    periods = Counter()
    importes = []
    for r in rs:
        fd = parse_date(r.get('fechaDesde'))
        fh = parse_date(r.get('fechaHasta'))
        p = detectar_periodicidad(fd, fh)
        if p: periods[p] += 1
        imp = r.get('importeFinal')
        if isinstance(imp, (int, float)) and imp > 0:
            importes.append(imp)
    cur = codcurs.most_common(1)[0][0] if codcurs else ''
    per = periods.most_common(1)[0][0] if periods else ''
    imp = round(sum(importes) / len(importes), 2) if importes else ''
    return (cur, per, imp)


comparativa = []
nf_used_ids = set()

for g in gp_clientes:
    n, by = find_nf(g)
    if n: nf_used_ids.add(n.get('id'))
    cur, per, imp_medio = cuota_inferida_de_recibos(g)
    cuota_round = cuota_by_idnoofit.get(str(n.get('id')) if n else '', None)

    fila = {
        'gp_codigo': g.get('codigo'),
        'gp_dni': g.get('dni') or g.get('dniContr') or '',
        'gp_nombre': g.get('nombre') or '',
        'gp_apellidos': g.get('apellidos') or '',
        'gp_email': g.get('email') or '',
        'gp_movil': g.get('movil') or g.get('telefono') or '',
        'gp_domicilio': g.get('domicilio') or '',
        'gp_poblacion': g.get('poblacion') or '',
        'gp_estado': 'Alta' if g.get('estado') == 1 else 'Baja',
        'gp_fechaAlta': fmt_date(parse_date(g.get('fechaAlta'))),
        'gp_fechaBaja': fmt_date(parse_date(g.get('fechaBaja'))),
        'gp_iban': g.get('_iban') or '',
        'gp_forma_pago': FORMA_PAGO_GP.get(g.get('_pagador_formaPago'), g.get('_pagador_formaPago') or ''),
        'gp_cuota_codcur': cur,
        'gp_periodicidad': per,
        'gp_importe_medio': imp_medio,
        # NoofitPro
        'nf_id': n.get('id') if n else '',
        'nf_dni': (n.get('dni') or n.get('nif') or '') if n else '',
        'nf_nombre': (n.get('name') or '') if n else '',
        'nf_apellidos': (n.get('surname') or '') if n else '',
        'nf_email': (n.get('email') or '') if n else '',
        'nf_telefono': (n.get('cellPhone') or n.get('telefono') or '') if n else '',
        'nf_domicilio': (n.get('address') or '') if n else '',
        'nf_enabled': ('Sí' if (n and n.get('enabled') is not False) else ('No' if n else '')),
        'nf_match_by': by or '',
        # Round (cuota asignada)
        'round_cuota_nombre': (cuota_round['nombre'] if cuota_round else ''),
        'round_cuota_periodicidad': (cuota_round['periodicidad'] if cuota_round else ''),
        'round_cuota_importe': (cuota_round['importe'] if cuota_round else ''),
        # Diff
        '_diffs': set(),
    }
    # Detectar diferencias entre GP y NF (si hay match)
    if n:
        if norm_dni(fila['gp_dni']) != norm_dni(fila['nf_dni']):
            if fila['gp_dni'] and fila['nf_dni']:
                fila['_diffs'].add('dni')
        if norm_text(fila['gp_nombre'] + ' ' + fila['gp_apellidos']) != norm_text(fila['nf_nombre'] + ' ' + fila['nf_apellidos']):
            fila['_diffs'].add('nombre')
        if norm_email(fila['gp_email']) != norm_email(fila['nf_email']):
            if fila['gp_email'] and fila['nf_email']:
                fila['_diffs'].add('email')
        if norm_phone(fila['gp_movil']) != norm_phone(fila['nf_telefono']):
            if fila['gp_movil'] and fila['nf_telefono']:
                fila['_diffs'].add('telefono')
    if not n:
        fila['_diffs'].add('SOLO_GP')
    # Coincidencia periodicidad GP ↔ Round
    if cuota_round and per:
        if (cuota_round['periodicidad'] or '').lower() != per:
            fila['_diffs'].add('periodicidad_round')
    comparativa.append(fila)

# Clientes solo en NF (no encontrados en GP)
solo_nf = [c for c in nf_clientes if c.get('id') not in nf_used_ids]
print(f'Solo en NoofitPro (huérfanos): {len(solo_nf)}')

# ─── Excel ────────────────────────────────────────────────────────────────────
wb = Workbook(); wb.remove(wb.active)
border = Border(*[Side(style='thin', color='CCCCCC')]*4)
fill_diff = PatternFill('solid', fgColor='FFE4E4')
fill_h = PatternFill('solid', fgColor='2DD4A8')
font_h = Font(bold=True, color='FFFFFF', size=11)


def hsheet(name, headers, color='2DD4A8'):
    ws = wb.create_sheet(name)
    f = PatternFill('solid', fgColor=color)
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.fill = f; c.font = font_h
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
    return ws


# RESUMEN
ws = wb.create_sheet('RESUMEN', 0)
for col, w in zip('ABC', [38, 16, 60]): ws.column_dimensions[col].width = w
ws.cell(1, 1, 'COMPARATIVA Cliente · GestPlus ↔ NoofitPro ↔ Round').font = Font(bold=True, size=14)

def w_kv(row, k, v, note=''):
    ws.cell(row, 1, k).font = Font(bold=True)
    ws.cell(row, 2, v)
    if note: ws.cell(row, 3, note).font = Font(italic=True, color='888888')

# Stats
n_total_gp = len(gp_clientes)
n_match = sum(1 for f in comparativa if not 'SOLO_GP' in f['_diffs'])
n_solo_gp = sum(1 for f in comparativa if 'SOLO_GP' in f['_diffs'])
n_diff_dni = sum(1 for f in comparativa if 'dni' in f['_diffs'])
n_diff_nombre = sum(1 for f in comparativa if 'nombre' in f['_diffs'])
n_diff_email = sum(1 for f in comparativa if 'email' in f['_diffs'])
n_diff_tel = sum(1 for f in comparativa if 'telefono' in f['_diffs'])
n_diff_per = sum(1 for f in comparativa if 'periodicidad_round' in f['_diffs'])

w_kv(3, 'Clientes GestPlus universo',     n_total_gp)
w_kv(4, '  Encontrados en NoofitPro',     n_match)
w_kv(5, '  SOLO en GestPlus (faltan NF)', n_solo_gp, '⚠️ falta crear en NF')
w_kv(6, 'Clientes SOLO en NoofitPro',     len(solo_nf), '⚠️ huérfanos sin recibo GP')
w_kv(8, 'Discrepancias (clientes en ambos):', '')
ws.cell(8, 1).font = Font(bold=True)
w_kv(9,  '  DNI distinto',                n_diff_dni)
w_kv(10, '  Nombre/apellidos distinto',   n_diff_nombre)
w_kv(11, '  Email distinto',              n_diff_email)
w_kv(12, '  Teléfono distinto',           n_diff_tel)
w_kv(13, '  Periodicidad ≠ Round',        n_diff_per, 'cuota asignada en Round vs detectada en GP')

w_kv(15, 'Clientes con cuota Round asignada', sum(1 for f in comparativa if f['round_cuota_nombre']))
w_kv(16, 'Clientes sin cuota Round',         sum(1 for f in comparativa if f['nf_id'] and not f['round_cuota_nombre']))


# COMPARATIVA (lista completa)
HEADS = [
    ('Cód GP', 9), ('GP DNI', 12), ('GP Nombre', 18), ('GP Apellidos', 22),
    ('GP Email', 24), ('GP Móvil', 12), ('GP Domicilio', 22), ('GP Población', 14),
    ('GP Estado', 9), ('GP Alta', 11), ('GP Baja', 11),
    ('GP IBAN', 26), ('GP Forma pago', 12),
    ('GP Cuota (codcur)', 14), ('GP Periodicidad', 13), ('GP Importe medio', 13),
    ('NF id', 10), ('NF DNI', 12), ('NF Nombre', 18), ('NF Apellidos', 22),
    ('NF Email', 24), ('NF Tel', 12), ('NF Dom', 22), ('NF Activo', 9), ('Match por', 10),
    ('Round Cuota', 22), ('Round Periodicidad', 14), ('Round Importe', 12),
    ('⚠️ Diffs', 25),
]
ws_c = hsheet('COMPARATIVA', HEADS)
for i, f in enumerate(comparativa, 2):
    vals = [
        f['gp_codigo'], f['gp_dni'], f['gp_nombre'], f['gp_apellidos'],
        f['gp_email'], f['gp_movil'], f['gp_domicilio'], f['gp_poblacion'],
        f['gp_estado'], f['gp_fechaAlta'], f['gp_fechaBaja'],
        f['gp_iban'], f['gp_forma_pago'],
        f['gp_cuota_codcur'], f['gp_periodicidad'], f['gp_importe_medio'],
        f['nf_id'], f['nf_dni'], f['nf_nombre'], f['nf_apellidos'],
        f['nf_email'], f['nf_telefono'], f['nf_domicilio'], f['nf_enabled'], f['nf_match_by'],
        f['round_cuota_nombre'], f['round_cuota_periodicidad'], f['round_cuota_importe'],
        ', '.join(sorted(f['_diffs'])),
    ]
    diff_cols = {
        'dni': [2, 18],
        'nombre': [3, 4, 19, 20],
        'email': [5, 21],
        'telefono': [6, 22],
        'periodicidad_round': [15, 27],
    }
    for j, v in enumerate(vals, 1):
        c = ws_c.cell(i, j, v); c.border = border
        for diff_name, cols in diff_cols.items():
            if diff_name in f['_diffs'] and j in cols:
                c.fill = fill_diff
        if 'SOLO_GP' in f['_diffs'] and 17 <= j <= 25:
            c.fill = PatternFill('solid', fgColor='FFF7CC')
ws_c.auto_filter.ref = ws_c.dimensions


# DISCREPANCIAS (subset solo con diffs)
disc = [f for f in comparativa if f['_diffs'] and 'SOLO_GP' not in f['_diffs']]
ws_d = hsheet('DISCREPANCIAS', HEADS, color='F87171')
for i, f in enumerate(disc, 2):
    vals = [
        f['gp_codigo'], f['gp_dni'], f['gp_nombre'], f['gp_apellidos'],
        f['gp_email'], f['gp_movil'], f['gp_domicilio'], f['gp_poblacion'],
        f['gp_estado'], f['gp_fechaAlta'], f['gp_fechaBaja'],
        f['gp_iban'], f['gp_forma_pago'],
        f['gp_cuota_codcur'], f['gp_periodicidad'], f['gp_importe_medio'],
        f['nf_id'], f['nf_dni'], f['nf_nombre'], f['nf_apellidos'],
        f['nf_email'], f['nf_telefono'], f['nf_domicilio'], f['nf_enabled'], f['nf_match_by'],
        f['round_cuota_nombre'], f['round_cuota_periodicidad'], f['round_cuota_importe'],
        ', '.join(sorted(f['_diffs'])),
    ]
    diff_cols = {
        'dni': [2, 18], 'nombre': [3, 4, 19, 20], 'email': [5, 21],
        'telefono': [6, 22], 'periodicidad_round': [15, 27],
    }
    for j, v in enumerate(vals, 1):
        c = ws_d.cell(i, j, v); c.border = border
        for diff_name, cols in diff_cols.items():
            if diff_name in f['_diffs'] and j in cols:
                c.fill = fill_diff
ws_d.auto_filter.ref = ws_d.dimensions


# SOLO_NF
ws_n = hsheet('SOLO_NF', [
    ('NF id', 10), ('Nombre', 22), ('Apellidos', 24), ('DNI', 14), ('Email', 30),
    ('Teléfono', 14), ('Activo', 9), ('Cuota Round', 22),
], color='5B9CF6')
for i, c in enumerate(solo_nf, 2):
    cuota_r = cuota_by_idnoofit.get(str(c.get('id')))
    vals = [
        c.get('id'), c.get('name') or '', c.get('surname') or '',
        c.get('dni') or c.get('nif') or '', c.get('email') or '',
        c.get('cellPhone') or c.get('telefono') or '',
        'Sí' if c.get('enabled') is not False else 'No',
        cuota_r['nombre'] if cuota_r else '',
    ]
    for j, v in enumerate(vals, 1):
        ws_n.cell(i, j, v).border = border
ws_n.auto_filter.ref = ws_n.dimensions

wb.save(OUT)
print(f'\n→ {OUT}')
print(f'  RESUMEN | COMPARATIVA ({len(comparativa)}) | DISCREPANCIAS ({len(disc)}) | SOLO_NF ({len(solo_nf)})')
