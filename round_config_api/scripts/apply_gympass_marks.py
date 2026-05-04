"""Aplica las marcas del Excel gympass:
  'poner gympass en gympassid' → SET gympassId = 'gympass' para ese ID NF
  'no hacer nada'              → skip

Variables: NOOFIT_EMAIL, NOOFIT_PASS, DRY_RUN (default 0 = aplicar).
"""
import os, sys, json, hashlib, time
import requests, urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from openpyxl import load_workbook

VERIFY = False
BASE = 'https://pro.wiemspro.com/wiemspro'
APP_VERSION = '1.8.39'
APP_ID = '1'
EMAIL = os.getenv('NOOFIT_EMAIL', '')
PWD = os.getenv('NOOFIT_PASS', '')
DRY = os.getenv('DRY_RUN', '0') == '1'
INPUT = os.getenv('INPUT', 'gympass_user_edit.xlsx')

if not EMAIL or not PWD: sys.exit('NOOFIT_EMAIL/NOOFIT_PASS required')

def login():
    body = {'email': EMAIL, 'appVersion': APP_VERSION,
            'password': hashlib.md5(PWD.encode()).hexdigest().upper()}
    r = requests.post(f'{BASE}/account/loginEasy', json=body, verify=VERIFY, timeout=30)
    if r.status_code != 200: sys.exit('Login fail')
    return r.headers.get('X-CustomToken'), r.headers.get('X-TRAINER_MANAGER', '')


def hdrs(t, m):
    return {'X-CustomToken':t,'X-TRAINER_MANAGER':m,'locale':'es',
            'appVersion':APP_VERSION,'appId':APP_ID,'Content-Type':'application/json'}


def get_clientes(t, m):
    r = requests.get(f'{BASE}/api/dispositivos/getClienteSimple',
                     headers=hdrs(t, m), verify=VERIFY, timeout=60)
    return {c['id']:c for c in r.json().get('clientes', [])}


def update_gympass(t, m, cliente, valor):
    payload = [{**cliente, 'gympassId': valor, 'toSend': True}]
    r = requests.post(f'{BASE}/api/dispositivos/clientePlusv2',
                      json=payload, headers=hdrs(t, m), verify=VERIFY, timeout=30)
    try:
        d = r.json()
        return r.status_code == 200 and d.get('mensaje') == 'OK', d
    except Exception:
        return False, {'raw': r.text[:200]}


def main():
    wb = load_workbook(INPUT, data_only=True)
    ws = wb.active
    targets = []
    for r in range(2, ws.max_row + 1):
        accion = (ws.cell(r, 14).value or '').lower()
        nf_id = ws.cell(r, 8).value
        if 'poner gympass' in accion and nf_id:
            targets.append({'r': r, 'nf_id': int(nf_id),
                            'cliente': ws.cell(r, 1).value,
                            'alias': ws.cell(r, 3).value})

    print(f'Targets: {len(targets)} clientes para SET gympassId')
    print(f'DRY_RUN = {DRY}\n')
    for t in targets:
        print(f"  id={t['nf_id']:>8d}  {t['cliente'][:30]:<30s}  alias='{t['alias']}'")

    if DRY:
        print('\n(DRY_RUN — no se aplica)')
        return

    print('\nLogin & fetch clientes...')
    token, manager = login()
    clientes_by_id = get_clientes(token, manager)
    print(f'  {len(clientes_by_id)} clientes cargados')

    print('\nAplicando...')
    ok = err = 0
    log = []
    for t in targets:
        c = clientes_by_id.get(t['nf_id'])
        if not c:
            print(f"  ❌ id={t['nf_id']} no encontrado")
            err += 1; continue
        success, resp = update_gympass(token, manager, c, 'gympass')
        log.append({'id': t['nf_id'], 'ok': success, 'cliente': t['cliente']})
        if success:
            ok += 1
            print(f"  ✅ id={t['nf_id']} {t['cliente']}")
        else:
            err += 1
            print(f"  ❌ id={t['nf_id']} {t['cliente']} resp={resp}")
        time.sleep(0.1)
    print(f'\nOK: {ok} · Errores: {err}')

    with open('gympass_apply_log.json', 'w', encoding='utf-8') as f:
        json.dump({'log': log, 'ok': ok, 'err': err}, f, ensure_ascii=False, indent=2)


if __name__ == '__main__':
    main()
