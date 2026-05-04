"""Aplica las marcas del Excel comparativa_v2:

  marca='ok' + acción ARCHIVAR  → set enabled=False en NF
  marca='ok' + acción CREAR     → crear cliente en NF (con datos de GP)
  marca='borrar'                → borrar cliente en GP (clientes test)
  marca='comprobar' / 'poner...' → skip
"""
import os, sys, json, hashlib, time, re
from datetime import datetime
import requests, urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from openpyxl import load_workbook

VERIFY = False
BASE_NF = 'https://pro.wiemspro.com/wiemspro'
BASE_GP = 'https://gestplus.okmas.net'
APP_VERSION = '1.8.39'; APP_ID = '1'

NF_EMAIL = os.getenv('NOOFIT_EMAIL', '')
NF_PWD   = os.getenv('NOOFIT_PASS', '')
GP_USER  = os.getenv('GESTPLUS_USER', '')
GP_PWD   = os.getenv('GESTPLUS_PASS', '')
DRY = os.getenv('DRY_RUN', '0') == '1'
INPUT = os.getenv('INPUT', 'comparativa_v2_user.xlsx')


def login_nf():
    body = {'email': NF_EMAIL, 'appVersion': APP_VERSION,
            'password': hashlib.md5(NF_PWD.encode()).hexdigest().upper()}
    r = requests.post(f'{BASE_NF}/account/loginEasy', json=body, verify=VERIFY, timeout=30)
    if r.status_code != 200: sys.exit('NF login fail')
    return r.headers.get('X-CustomToken'), r.headers.get('X-TRAINER_MANAGER', '')


def hdrs_nf(t, m):
    return {'X-CustomToken':t,'X-TRAINER_MANAGER':m,'locale':'es',
            'appVersion':APP_VERSION,'appId':APP_ID,'Content-Type':'application/json'}


def login_gp():
    sys.path.insert(0, '/opt/round_config_api/scripts')
    from gestplus_extract import login as gp_login
    s = requests.Session()
    s.headers['User-Agent'] = 'Mozilla/5.0 RoundMig/1.0'
    os.environ['GESTPLUS_USER'] = GP_USER
    os.environ['GESTPLUS_PASS'] = GP_PWD
    gp_login(s)
    return s


def archivar_nf(cliente, t, m):
    payload = [{**cliente, 'enabled': False, 'toSend': True, 'motivoArchivado': 'baja gestplus'}]
    r = requests.post(f'{BASE_NF}/api/dispositivos/clientePlusv2',
                      json=payload, headers=hdrs_nf(t, m), verify=VERIFY, timeout=30)
    try: return r.status_code == 200 and r.json().get('mensaje') == 'OK'
    except Exception: return False


def crear_nf(g, t, m):
    """Crea cliente en NF a partir de datos de GestPlus."""
    nuevo = {
        'name': g.get('nombre') or '',
        'surname': g.get('apellidos') or '',
        'email': g.get('email') or '',
        'dni': g.get('dni') or '',
        'cellPhone': g.get('movil') or g.get('telefono') or '',
        'birthdate': (g.get('fechaNacimiento') or '')[:10] if g.get('fechaNacimiento') else None,
        'gender': 'M' if g.get('sexo') == 1 else ('F' if g.get('sexo') == 2 else None),
        'enabled': True,
        'toSend': True,
    }
    nuevo = {k:v for k, v in nuevo.items() if v not in (None, '')}
    r = requests.post(f'{BASE_NF}/api/dispositivos/clientePlusv2',
                      json=[nuevo], headers=hdrs_nf(t, m), verify=VERIFY, timeout=30)
    try:
        d = r.json()
        return r.status_code == 200 and d.get('mensaje') == 'OK', d
    except Exception:
        return False, {'raw': r.text[:200]}


def borrar_gp(s, codigo):
    """Intenta dar de baja al cliente en GestPlus.
    Probamos varios endpoints típicos."""
    cli = next((c for c in CLIENTES_GP if c.get('codigo') == codigo), None)
    if not cli: return False, 'no encontrado en dump'
    s.get(f'{BASE_GP}/GestPlus/detailCliente.action',
          params={'codClippal': cli['codclippal'], 'codigo': cli['codigo']}, timeout=20)
    # Probar endpoint de baja
    for path, params in [
        ('/GestPlus/borrarCliente.action', {'codigo': codigo, 'codClippal': cli['codclippal']}),
        ('/GestPlus/eliminarCliente.action', {'codigo': codigo, 'codClippal': cli['codclippal']}),
        ('/GestPlus/anularCliente.action', {'codigo': codigo, 'codClippal': cli['codclippal']}),
    ]:
        try:
            r = s.post(f'{BASE_GP}{path}', data=params, timeout=20)
            if r.status_code == 200 and ('json' in r.headers.get('Content-Type','')
                                         or 'success' in r.text.lower()
                                         or 'ok' in r.text.lower()):
                return True, path
        except Exception: pass
    return False, 'todos los endpoints fallaron'


def main():
    wb = load_workbook(INPUT, data_only=True)
    ws = wb.active

    # Index GP por código
    with open('gestplus_dump_2026-05-02.json', 'r', encoding='utf-8') as f:
        gp = json.load(f)
    global CLIENTES_GP
    CLIENTES_GP = gp['altas'] + gp['bajas_recientes_12m']
    gp_by_cod = {g['codigo']: g for g in CLIENTES_GP}

    # Leer marcas
    archivar = []  # (nf_id, cliente_name)
    crear = []     # (gp_codigo, gp_obj)
    borrar = []    # (gp_codigo, cliente_name)

    for r in range(2, ws.max_row + 1):
        marca = (ws.cell(r, 2).value or '').strip().lower()
        cliente = ws.cell(r, 1).value
        cod_gp = ws.cell(r, 5).value
        nf_id = ws.cell(r, 6).value
        accion = ws.cell(r, 14).value or ''
        if marca == 'ok':
            if 'ARCHIVAR' in accion and nf_id:
                archivar.append((int(nf_id), cliente))
            elif 'CREAR' in accion and cod_gp:
                g = gp_by_cod.get(str(cod_gp).strip())
                if g: crear.append((cod_gp, g))
        elif marca == 'borrar' and cod_gp:
            borrar.append((cod_gp, cliente))

    print(f'A ARCHIVAR (NF):  {len(archivar)}')
    for nf_id, c in archivar: print(f'  - id={nf_id} {c}')
    print(f'\nA CREAR (NF):     {len(crear)}')
    for cod, g in crear: print(f'  - GP {cod} {g.get("nombre")} {g.get("apellidos")} email={g.get("email")} dni={g.get("dni")}')
    print(f'\nA BORRAR (GP):    {len(borrar)}')
    for cod, c in borrar: print(f'  - GP {cod} {c}')

    if DRY:
        print('\n(DRY_RUN — no se aplica)'); return

    print('\nLogin NF...')
    t, m = login_nf()
    # Cargar clientes NF para tener datos completos al archivar
    r = requests.get(f'{BASE_NF}/api/dispositivos/getClienteSimple', headers=hdrs_nf(t, m), verify=VERIFY, timeout=60)
    nf_by_id = {c['id']: c for c in r.json().get('clientes', [])}

    log = {'archivar': [], 'crear': [], 'borrar': []}

    print('\n=== ARCHIVAR ===')
    for nf_id, cli in archivar:
        c = nf_by_id.get(nf_id)
        if not c:
            print(f'  ❌ {nf_id} {cli} no encontrado en NF'); continue
        ok = archivar_nf(c, t, m)
        log['archivar'].append({'id':nf_id,'cliente':cli,'ok':ok})
        print(f'  {"✅" if ok else "❌"} id={nf_id} {cli}')
        time.sleep(0.1)

    print('\n=== CREAR ===')
    for cod, g in crear:
        ok, resp = crear_nf(g, t, m)
        log['crear'].append({'gp_codigo':cod,'cliente':f"{g.get('nombre')} {g.get('apellidos')}",'ok':ok,'resp':resp})
        print(f'  {"✅" if ok else "❌"} GP {cod} {g.get("nombre")} {g.get("apellidos")}')
        if not ok: print(f'     resp: {resp}')
        time.sleep(0.15)

    print('\n=== BORRAR EN GP ===')
    if borrar:
        s = login_gp()
        for cod, cli in borrar:
            ok, info = borrar_gp(s, cod)
            log['borrar'].append({'gp_codigo':cod,'cliente':cli,'ok':ok,'info':info})
            print(f'  {"✅" if ok else "⚠️"} GP {cod} {cli} → {info}')
            time.sleep(0.2)

    with open('apply_v2_log.json', 'w', encoding='utf-8') as f:
        json.dump(log, f, ensure_ascii=False, indent=2, default=str)
    print('\nGuardado: apply_v2_log.json')


if __name__ == '__main__':
    main()
