"""Genera Excel comparativo Round (junio 2026) vs GestPlus, ordenado por causa.

Hojas:
  1. Resumen
  2. Diferencias de importe (agrupado por causa)
  3. Solo en GestPlus
  4. Solo en Round
  5. DNIs casi iguales (probablemente el mismo cliente con typo)
"""
import re
import sys
import unicodedata as ud
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.routes.preemision_validar import _validar_emision

GP_PATH = '/tmp/gp.xlsx'
OUT = sys.argv[1] if len(sys.argv) > 1 else '/tmp/comparativa.xlsx'


def norm_dni(d):
    s = ''.join(c for c in ud.normalize('NFD', str(d or ''))
                if ud.category(c) != 'Mn')
    return re.sub(r'[^A-Z0-9]', '', s.upper())


def solo_digitos(d):
    return re.sub(r'[^0-9]', '', str(d or ''))


# 1) Cargar GestPlus
gp_wb = load_workbook(GP_PATH)
gp_ws = gp_wb.active
gp_heads = [gp_ws.cell(1, c).value for c in range(1, gp_ws.max_column+1)]
H = {h: i+1 for i, h in enumerate(gp_heads)}

gp_por_dni = {}
for r in range(2, gp_ws.max_row+1):
    dni = norm_dni(gp_ws.cell(r, H['dniContr']).value)
    if not dni: continue
    importe = float(gp_ws.cell(r, H['Final']).value or 0)
    nombre = gp_ws.cell(r, H['Nombre']).value or ''
    cur = gp_ws.cell(r, H['codcur']).value or ''
    modif = float(gp_ws.cell(r, H['Modif']).value or 0)
    desc = gp_ws.cell(r, H['descu']).value or ''
    if dni in gp_por_dni:
        gp_por_dni[dni]['importe'] += importe
        gp_por_dni[dni]['cursos'].append(cur)
        gp_por_dni[dni]['modif'] += modif
        if desc and desc not in gp_por_dni[dni]['descs']:
            gp_por_dni[dni]['descs'].append(desc)
    else:
        gp_por_dni[dni] = {'nombre': nombre, 'importe': importe,
                            'cursos': [cur], 'modif': modif,
                            'descs': [desc] if desc else []}

# 2) Validación Round
coherentes, _, _ = _validar_emision('17675', '2026-06')
round_por_dni = {}
for c in coherentes:
    dni = norm_dni(c.get('dni') or '')
    if not dni: continue
    round_por_dni[dni] = c

# 3) Cruce
ambos = set(gp_por_dni) & set(round_por_dni)
solo_gp = set(gp_por_dni) - set(round_por_dni)
solo_round = set(round_por_dni) - set(gp_por_dni)

# 4) DNIs casi iguales (matching por dígitos)
def _digitos(d): return solo_digitos(d)
gp_por_dig = {}
for d in solo_gp:
    digs = _digitos(d)
    if len(digs) >= 7:
        gp_por_dig.setdefault(digs, []).append(d)
round_por_dig = {}
for d in solo_round:
    digs = _digitos(d)
    if len(digs) >= 7:
        round_por_dig.setdefault(digs, []).append(d)
casi_iguales = []
for digs, gps in gp_por_dig.items():
    if digs in round_por_dig:
        casi_iguales.append((digs, gps, round_por_dig[digs]))
# o por substring
for d_gp in list(solo_gp):
    digs_gp = _digitos(d_gp)
    if len(digs_gp) < 7: continue
    for d_rd in list(solo_round):
        digs_rd = _digitos(d_rd)
        if len(digs_rd) < 7: continue
        if digs_gp == digs_rd: continue  # ya cubierto
        if digs_gp in digs_rd or digs_rd in digs_gp:
            casi_iguales.append((f'{digs_gp}~{digs_rd}', [d_gp], [d_rd]))

# 5) Diferencias de importe — clasificar por causa
diffs = []
for dni in ambos:
    gp = gp_por_dni[dni]
    rd = round_por_dni[dni]
    rd_imp = float(rd.get('importe_total') or 0)
    delta = round(rd_imp - gp['importe'], 2)
    if abs(delta) < 0.01: continue
    # Determinar causa
    rd_cuotas = set(rd.get('cuotas') or [])
    gp_cur = ' + '.join(gp['cursos'])
    rd_cur = ' + '.join(sorted(rd_cuotas))
    # Heurística:
    cuotas_diff = len(gp['cursos']) != len(rd_cuotas)
    # Modificaciones Round
    rd_mod_total = 0
    rd_notas_mod = []
    for d in (rd.get('cuotas_detalle') or []):
        for m in (d.get('modificaciones_struct') or []):
            rd_mod_total += float(m.get('delta') or 0)
            tag = f"{d.get('codigo')} {m.get('tipo')} {m.get('delta'):+.2f}€"
            if m.get('razon'): tag += f" ({m['razon']})"
            rd_notas_mod.append(tag)
    for m in (rd.get('modificaciones_globales_struct') or []):
        rd_mod_total += float(m.get('delta') or 0)
        tag = f"global {m.get('tipo')} {m.get('delta'):+.2f}€"
        if m.get('razon'): tag += f" ({m['razon']})"
        rd_notas_mod.append(tag)
    # Descuentos Round
    rd_desc_total = 0
    rd_notas_desc = []
    for d in (rd.get('cuotas_detalle') or []):
        for x in (d.get('descuentos_struct') or []):
            rd_desc_total += float(x.get('ahorro') or 0)
            rd_notas_desc.append(f"{d.get('codigo')} {x.get('codigo')} −{x.get('ahorro'):.2f}€")
    # Clasificar causa
    if cuotas_diff:
        causa = 'CUOTAS DISTINTAS' if len(gp['cursos']) > len(rd_cuotas) \
            else 'CUOTAS DISTINTAS (Round tiene más)'
        explicacion = (f'GP factura {len(gp["cursos"])} cuota(s) ({gp_cur}); '
                       f'Round tiene {len(rd_cuotas)} sub(s) ({rd_cur}). '
                       f'Falta alta de sub en Round o sobran en GP.')
    elif rd_mod_total != 0:
        causa = 'MODIFICACIÓN aplicada en Round (no en GP)'
        explicacion = f'Modificación Round: {"; ".join(rd_notas_mod)}'
    elif rd_desc_total > 0 and abs(delta + rd_desc_total) < 0.5:
        causa = 'DESCUENTO aplicado en Round (no en GP)'
        explicacion = f'Descuento Round: {"; ".join(rd_notas_desc)}'
    elif gp['importe'] == 0:
        causa = 'GP no factura (¿beca/promo?)'
        explicacion = ('GP marca importe 0; Round emite normal. Si es '
                       'beca/promo, marcar en Round con importe 0 justificado.')
    else:
        causa = 'OTRO (revisar)'
        explicacion = (f'GP={gp["importe"]:.2f}, Round={rd_imp:.2f}. '
                       f'Descuento Round: {", ".join(rd_notas_desc) or "—"}. '
                       f'Modif Round: {", ".join(rd_notas_mod) or "—"}.')
    diffs.append({
        'dni': dni, 'gp_nombre': gp['nombre'],
        'rd_nombre': rd.get('nombre'), 'gp_imp': gp['importe'],
        'rd_imp': rd_imp, 'delta': delta,
        'gp_cur': gp_cur, 'rd_cur': rd_cur,
        'causa': causa, 'explicacion': explicacion,
    })

# Ordenar por causa, luego por |delta| descendente
orden_causa = {
    'CUOTAS DISTINTAS': 1,
    'CUOTAS DISTINTAS (Round tiene más)': 2,
    'MODIFICACIÓN aplicada en Round (no en GP)': 3,
    'DESCUENTO aplicado en Round (no en GP)': 4,
    'GP no factura (¿beca/promo?)': 5,
    'OTRO (revisar)': 9,
}
diffs.sort(key=lambda d: (orden_causa.get(d['causa'], 99), -abs(d['delta'])))

# ─── Generar Excel ───────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

border = Border(*[Side(style='thin', color='CCCCCC')] * 4)
F_HDR = Font(bold=True, color='FFFFFF', size=11, name='Arial')
F_BOLD = Font(bold=True, name='Arial')
FILL_FIJO = PatternFill('solid', fgColor='2DD4A8')
FILL_DIFF = PatternFill('solid', fgColor='DC2626')
FILL_SOLOGP = PatternFill('solid', fgColor='F59E0B')
FILL_SOLORD = PatternFill('solid', fgColor='3B82F6')
FILL_CASI   = PatternFill('solid', fgColor='8B5CF6')
FILL_GROUP  = PatternFill('solid', fgColor='F3F4F6')

# Hoja 1 — RESUMEN
ws = wb.create_sheet('RESUMEN')
ws.cell(1, 1, 'Comparativa Round (junio 2026) vs GestPlus').font = Font(bold=True, size=14, name='Arial')
ws.cell(3, 1, 'Clientes GestPlus (únicos por DNI)').font = F_BOLD
ws.cell(3, 2, len(gp_por_dni))
ws.cell(4, 1, 'Recibos GestPlus totales').font = F_BOLD
ws.cell(4, 2, gp_ws.max_row - 1)
ws.cell(5, 1, 'Coherentes Round (a emitir)').font = F_BOLD
ws.cell(5, 2, len(coherentes))
ws.cell(7, 1, 'Cruzan por DNI').font = F_BOLD
ws.cell(7, 2, len(ambos))
ws.cell(8, 1, '  · Cuadran exactos').font = F_BOLD
ws.cell(8, 2, len(ambos) - len(diffs))
ws.cell(9, 1, '  · Con diferencias').font = F_BOLD
ws.cell(9, 2, len(diffs))
ws.cell(11, 1, 'Solo en GestPlus').font = F_BOLD
ws.cell(11, 2, len(solo_gp))
ws.cell(12, 1, 'Solo en Round').font = F_BOLD
ws.cell(12, 2, len(solo_round))
ws.cell(13, 1, '  · DNIs casi iguales (probable mismo cliente)').font = F_BOLD
ws.cell(13, 2, len(casi_iguales))

# Diferencias por causa
from collections import Counter
causas = Counter(d['causa'] for d in diffs)
r = 16
ws.cell(r, 1, 'DIFERENCIAS POR CAUSA').font = F_HDR
ws.cell(r, 1).fill = FILL_DIFF
ws.cell(r, 2, '# Casos').font = F_HDR; ws.cell(r, 2).fill = FILL_DIFF
ws.cell(r, 3, 'Δ neto €').font = F_HDR; ws.cell(r, 3).fill = FILL_DIFF
for col in (1, 2, 3): ws.cell(r, col).alignment = Alignment(horizontal='center')
r += 1
for causa, _ in sorted(causas.items(), key=lambda x: orden_causa.get(x[0], 99)):
    grupo = [d for d in diffs if d['causa'] == causa]
    ws.cell(r, 1, causa)
    ws.cell(r, 2, len(grupo))
    ws.cell(r, 3, round(sum(d['delta'] for d in grupo), 2)).number_format = '#,##0.00 €'
    r += 1

for col, w in zip('ABC', (60, 12, 14)):
    ws.column_dimensions[col].width = w

# Hoja 2 — DIFERENCIAS (ordenadas por causa)
ws = wb.create_sheet('Diferencias')
heads = [('Causa', 50), ('DNI', 12), ('Nombre GP', 30), ('Nombre Round', 30),
         ('Importe GP €', 13), ('Importe Round €', 14), ('Δ €', 10),
         ('Cuotas GP', 32), ('Cuotas Round', 22), ('Explicación', 70)]
for col, (h, w) in enumerate(heads, 1):
    c = ws.cell(1, col, h); c.font = F_HDR; c.fill = FILL_DIFF; c.border = border
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = 'A2'

# Render con separadores por causa
row = 2
prev_causa = None
for d in diffs:
    if d['causa'] != prev_causa:
        # Header de grupo
        gc = ws.cell(row, 1, f'━━━  {d["causa"]}  ━━━')
        gc.font = Font(bold=True, color='1F2937', name='Arial')
        gc.fill = FILL_GROUP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(heads))
        row += 1
        prev_causa = d['causa']
    vals = [d['causa'], d['dni'], d['gp_nombre'], d['rd_nombre'],
            d['gp_imp'], d['rd_imp'], d['delta'],
            d['gp_cur'], d['rd_cur'], d['explicacion']]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row, j, v)
        cell.border = border
        cell.font = Font(name='Arial', size=10)
        if j in (5, 6, 7):
            cell.number_format = '#,##0.00 €'
            cell.alignment = Alignment(horizontal='right')
        if j == 7 and d['delta'] < 0:
            cell.font = Font(name='Arial', size=10, color='C00000', bold=True)
        if j == 7 and d['delta'] > 0:
            cell.font = Font(name='Arial', size=10, color='006400', bold=True)
        if j == 10:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    row += 1
ws.auto_filter.ref = f'A1:{get_column_letter(len(heads))}{row-1}'

# Hoja 3 — Solo GP
ws = wb.create_sheet('Solo GP')
heads = [('DNI', 12), ('Nombre', 35), ('Cuotas GP', 30), ('Importe €', 12),
         ('Modif €', 10), ('Cod. desc', 10)]
for col, (h, w) in enumerate(heads, 1):
    c = ws.cell(1, col, h); c.font = F_HDR; c.fill = FILL_SOLOGP; c.border = border
    c.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
for i, dni in enumerate(sorted(solo_gp), 2):
    gp = gp_por_dni[dni]
    vals = [dni, gp['nombre'], ' + '.join(gp['cursos']), gp['importe'],
            gp.get('modif', 0), ', '.join(gp['descs'])]
    for j, v in enumerate(vals, 1):
        c = ws.cell(i, j, v); c.border = border; c.font = Font(name='Arial', size=10)
        if j in (4, 5): c.number_format = '#,##0.00 €'
ws.auto_filter.ref = ws.dimensions

# Hoja 4 — Solo Round
ws = wb.create_sheet('Solo Round')
heads = [('DNI', 12), ('Nombre', 35), ('Cuotas Round', 22),
         ('Importe €', 12), ('Forma pago', 12)]
for col, (h, w) in enumerate(heads, 1):
    c = ws.cell(1, col, h); c.font = F_HDR; c.fill = FILL_SOLORD; c.border = border
    c.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
for i, dni in enumerate(sorted(solo_round), 2):
    rd = round_por_dni[dni]
    vals = [dni, rd.get('nombre'), ' + '.join(rd.get('cuotas') or []),
            float(rd.get('importe_total') or 0), rd.get('forma_pago')]
    for j, v in enumerate(vals, 1):
        c = ws.cell(i, j, v); c.border = border; c.font = Font(name='Arial', size=10)
        if j == 4: c.number_format = '#,##0.00 €'
ws.auto_filter.ref = ws.dimensions

# Hoja 5 — DNIs casi iguales
ws = wb.create_sheet('DNIs casi iguales')
ws.cell(1, 1, 'Probable mismo cliente con DNI mal escrito en una de las BDs').font = Font(italic=True, name='Arial', size=10)
heads = [('DNI GP', 14), ('Nombre GP', 30), ('Imp GP €', 11),
         ('DNI Round', 14), ('Nombre Round', 30), ('Imp Round €', 11),
         ('Sugerencia', 30)]
for col, (h, w) in enumerate(heads, 1):
    c = ws.cell(3, col, h); c.font = F_HDR; c.fill = FILL_CASI; c.border = border
    c.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(col)].width = w
i = 4
for digs, gps, rds in casi_iguales:
    for d_gp in gps:
        for d_rd in rds:
            gp = gp_por_dni.get(d_gp, {})
            rd = round_por_dni.get(d_rd, {})
            vals = [d_gp, gp.get('nombre',''), gp.get('importe',0),
                    d_rd, rd.get('nombre',''), float(rd.get('importe_total',0) or 0),
                    'Unificar DNI en la BD que esté mal']
            for j, v in enumerate(vals, 1):
                c = ws.cell(i, j, v); c.border = border
                c.font = Font(name='Arial', size=10)
                if j in (3, 6): c.number_format = '#,##0.00 €'
            i += 1
ws.freeze_panes = 'A4'

wb.save(OUT)
print(f'OK: {OUT}')
print(f'  Diferencias: {len(diffs)}')
print(f'  Solo GP: {len(solo_gp)}')
print(f'  Solo Round: {len(solo_round)}')
print(f'  Casi iguales: {len(casi_iguales)}')
