"""Lee las anotaciones manuales del usuario en comparativa_user_edit.xlsx."""
from openpyxl import load_workbook
from collections import Counter

wb = load_workbook('comparativa_user_edit.xlsx', data_only=True)
ws = wb.active

hits = []
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 7).value  # col 7 = "Alta en GestPlus" donde el usuario marcó tags
    if v is None: continue
    s = str(v).strip()
    if not s: continue
    if len(s) <= 10 and s.count('/') == 2: continue  # fecha dd/mm/yyyy
    cliente = ws.cell(r, 1).value
    dni = ws.cell(r, 3).value
    email = ws.cell(r, 4).value
    origen = ws.cell(r, 11).value
    hits.append({'r': r, 'marca': s, 'cliente': cliente, 'dni': dni,
                 'email': email, 'origen': origen})

print(f'Total anotaciones: {len(hits)}')
print()
print('Valores marcados:', Counter(h['marca'].lower() for h in hits).most_common())
print()
for h in hits:
    cli = (h['cliente'] or '')[:35]
    dni = (h['dni'] or '')[:11]
    print(f"  r{h['r']:>4d}  marca={h['marca']:<14s}  origen={str(h['origen'])[:8]:<8s}  cliente={cli:<35s}  dni={dni}")
