"""Genera un Excel con la comparativa cliente a cliente:

Columnas:
  - Cliente (nombre apellidos)
  - DNI
  - Email
  - Alta en GestPlus  (fecha o vacío)
  - Alta en NoofitPro (Sí/No)
  - Baja en GestPlus  (fecha o vacío)
  - Archivado en NoofitPro (Sí/No)
  - Origen           (gestplus / noofit / ambos)
  - Acción sugerida

Uso:
  python build_excel_compare.py [gestplus_dump.json] [noofit_clientes_dump.json] [salida.xlsx]
"""
import os, sys, json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def normaliza_dni(d):
    return (d or '').strip().upper().replace(' ', '').replace('-', '')


def normaliza_email(e):
    return (e or '').strip().lower()


def fmt_fecha(s):
    if not s: return ''
    try:
        d = datetime.fromisoformat(str(s).replace('Z', ''))
        return d.strftime('%d/%m/%Y')
    except Exception:
        return str(s)[:10]


def build_filas(gp_path, nf_path):
    with open(gp_path, 'r', encoding='utf-8') as f:
        gp = json.load(f)
    with open(nf_path, 'r', encoding='utf-8') as f:
        nf = json.load(f)

    todos_gp = gp.get('altas', []) + gp.get('bajas_recientes_12m', [])
    # Bajas descartadas (>12m) las añadimos también para visibilidad — solo cabecera mínima
    bajas_old = gp.get('bajas_descartadas', []) or []
    noofit = nf.get('clientes', [])

    # Indexar NoofitPro por DNI y email
    nf_by_dni = {}
    nf_by_email = {}
    for c in noofit:
        d = normaliza_dni(c.get('dni') or c.get('nif'))
        if d: nf_by_dni[d] = c
        e = normaliza_email(c.get('email'))
        if e: nf_by_email[e] = c

    filas = []
    matched_noofit_ids = set()

    def find_in_noofit(g):
        for d in (normaliza_dni(g.get('dni')), normaliza_dni(g.get('dniContr'))):
            if d and d in nf_by_dni: return nf_by_dni[d]
        e = normaliza_email(g.get('email'))
        if e and e in nf_by_email: return nf_by_email[e]
        return None

    # GestPlus alta + baja reciente
    for g in todos_gp:
        n = find_in_noofit(g)
        if n: matched_noofit_ids.add(n.get('id'))
        es_alta_gp = g.get('estado') == 1
        en_noofit = n is not None
        archivado_noofit = bool(n) and (n.get('enabled') is False or n.get('enabled') == 0)
        alta_noofit = en_noofit and not archivado_noofit
        baja_gp = (not es_alta_gp) and bool(g.get('fechaBaja'))

        # Acción sugerida
        if es_alta_gp and not en_noofit:
            accion = '➕ CREAR en NoofitPro'
        elif es_alta_gp and archivado_noofit:
            accion = '🔓 REACTIVAR en NoofitPro'
        elif baja_gp and alta_noofit:
            accion = '📦 ARCHIVAR en NoofitPro'
        elif baja_gp and archivado_noofit:
            accion = '✅ OK (archivado en ambos)'
        elif es_alta_gp and alta_noofit:
            accion = '✅ OK (activo en ambos)'
        elif baja_gp and not en_noofit:
            accion = '➖ Baja antigua (no estaba en Noofit)'
        else:
            accion = '?'

        filas.append({
            'cliente': f"{g.get('nombre','') or ''} {g.get('apellidos','') or ''}".strip(),
            'dni': g.get('dni') or '',
            'email': g.get('email') or '',
            'codigo_gp': g.get('codigo'),
            'noofit_id': (n.get('id') if n else ''),
            'alta_gestplus': fmt_fecha(g.get('fechaAlta')) if es_alta_gp else '',
            'alta_noofit': 'Sí' if alta_noofit else 'No',
            'baja_gestplus': fmt_fecha(g.get('fechaBaja')) if baja_gp else '',
            'archivado_noofit': 'Sí' if archivado_noofit else 'No',
            'origen': 'ambos' if en_noofit else 'gestplus',
            'accion': accion,
        })

    # Huérfanos: NoofitPro que no aparecen en GestPlus alta/baja_reciente
    # Considerar también bajas antiguas para no marcar como huérfanos los que están en GP pero antiguos
    gp_dnis = set()
    gp_emails = set()
    for g in todos_gp + bajas_old:
        for k in ('dni', 'dniContr'):
            d = normaliza_dni(g.get(k))
            if d: gp_dnis.add(d)
        e = normaliza_email(g.get('email'))
        if e: gp_emails.add(e)
    for c in noofit:
        if c.get('id') in matched_noofit_ids: continue
        d = normaliza_dni(c.get('dni') or c.get('nif'))
        e = normaliza_email(c.get('email'))
        if (d and d in gp_dnis) or (e and e in gp_emails): continue
        archivado = c.get('enabled') is False or c.get('enabled') == 0
        nombre = f"{c.get('nombre') or c.get('name') or ''} {c.get('apellidos') or c.get('surname') or ''}".strip()
        filas.append({
            'cliente': nombre or '(sin nombre)',
            'dni': c.get('dni') or c.get('nif') or '',
            'email': c.get('email') or '',
            'codigo_gp': '',
            'noofit_id': c.get('id'),
            'alta_gestplus': '',
            'alta_noofit': 'No' if archivado else 'Sí',
            'baja_gestplus': '',
            'archivado_noofit': 'Sí' if archivado else 'No',
            'origen': 'noofit',
            'accion': '🔍 Huérfano (revisar)',
        })

    # Orden: ambos primero, luego gestplus, luego noofit; dentro alfabético
    orden = {'ambos': 0, 'gestplus': 1, 'noofit': 2}
    filas.sort(key=lambda r: (orden.get(r['origen'], 9), r['cliente'].upper()))
    return filas


def write_excel(filas, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Comparativa'

    headers = [
        ('Cliente', 32),
        ('DNI', 12),
        ('Email', 30),
        ('Cód. GestPlus', 12),
        ('ID NoofitPro', 11),
        ('Alta en GestPlus', 14),
        ('Alta en NoofitPro', 14),
        ('Baja en GestPlus', 14),
        ('Archivado en NoofitPro', 16),
        ('Origen', 10),
        ('Acción sugerida', 32),
    ]
    keys = ['cliente', 'dni', 'email', 'codigo_gp', 'noofit_id',
            'alta_gestplus', 'alta_noofit', 'baja_gestplus', 'archivado_noofit',
            'origen', 'accion']

    # Estilo cabecera
    head_fill = PatternFill('solid', fgColor='2DD4A8')
    head_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(*[Side(style='thin', color='CCCCCC')]*4)

    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # Filas
    fill_si = PatternFill('solid', fgColor='D1FADF')   # verde claro
    fill_no = PatternFill('solid', fgColor='FFE4E4')   # rojo claro
    fill_warn = PatternFill('solid', fgColor='FFF7CC') # amarillo
    for i, f in enumerate(filas, 2):
        for j, k in enumerate(keys, 1):
            v = f.get(k, '')
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = border
            if k in ('alta_noofit', 'archivado_noofit'):
                if v == 'Sí':
                    cell.fill = fill_si if k == 'alta_noofit' else fill_warn
                elif v == 'No':
                    cell.fill = fill_no if k == 'alta_noofit' else fill_si
            elif k == 'accion':
                if 'CREAR' in v:    cell.fill = fill_no
                elif 'ARCHIVAR' in v: cell.fill = fill_warn
                elif 'OK' in v:     cell.fill = fill_si
                elif 'Huérfano' in v: cell.fill = fill_warn

    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)
    return out_path


def main():
    gp = sys.argv[1] if len(sys.argv) > 1 else 'gestplus_dump_' + datetime.now().date().isoformat() + '.json'
    nf = sys.argv[2] if len(sys.argv) > 2 else 'noofit_clientes_dump.json'
    out = sys.argv[3] if len(sys.argv) > 3 else 'comparativa_gestplus_noofit.xlsx'
    if not os.path.exists(gp): sys.exit(f'No existe {gp}')
    if not os.path.exists(nf): sys.exit(f'No existe {nf}')
    filas = build_filas(gp, nf)
    write_excel(filas, out)
    print(f'✅ Excel generado: {out}')
    print(f'   Total filas: {len(filas)}')
    # Resumen por accion
    from collections import Counter
    cnt = Counter(f['accion'] for f in filas)
    for k, v in cnt.most_common():
        print(f'   {v:5d}  {k}')


if __name__ == '__main__':
    main()
