"""Genera un Excel con todos los clientes activos del manager + su categoría
actual, listo para que el manager ponga la categoría nueva en bloque y
después se importe con `apply_clientes_categorias.py`.

Salida: /tmp/clientes_categorias.xlsx con columnas:
  - id_noofit
  - DNI
  - Nombre
  - Apellidos
  - Email
  - Teléfono
  - Estado NoofitPro
  - Categoría actual
  - Nueva categoría (vacía — RELLENAR)

NOTA importante: usa el dump local NoofitPro existente
(/opt/round_config_api/noofit_clientes_dump.json). Si quieres datos
frescos, primero ejecuta `noofit_extract_compare.py`.
"""
import json, os, sys
sys.path.insert(0, '/opt/round_config_api')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.db import get_conn

ID_MANAGER = os.getenv('ID_MANAGER', '17675')
NF_DUMP = os.getenv('NF_DUMP', '/opt/round_config_api/noofit_clientes_dump.json')
OUT = os.getenv('OUT', '/tmp/clientes_categorias.xlsx')

# 1) Cargar clientes NoofitPro
with open(NF_DUMP, 'r', encoding='utf-8') as f:
    nf = json.load(f)
clientes = nf.get('clientes', [])
activos = [c for c in clientes if c.get('enabled') is not False and c.get('enabled') != 0]
print(f'NoofitPro dump: {len(clientes)} totales | {len(activos)} activos')

# 2) Cargar categorías del manager
with get_conn() as conn, conn.cursor() as cur:
    cur.execute("""
        SELECT id, nombre, color, activa
          FROM categoria WHERE id_manager=%s ORDER BY activa DESC, nombre
    """, (str(ID_MANAGER),))
    cats = cur.fetchall()
    cur.execute("""
        SELECT cliente_idnoofit, c.nombre AS cat_nombre
          FROM cliente_categoria cc
          JOIN categoria c ON c.id = cc.categoria_id
         WHERE cc.id_manager = %s
    """, (str(ID_MANAGER),))
    asignaciones = {row['cliente_idnoofit']: row['cat_nombre'] for row in cur.fetchall()}

cat_nombres = [c['nombre'] for c in cats if c['activa']]
print(f'Categorías activas del manager: {", ".join(cat_nombres) or "(ninguna)"}')
print(f'Asignaciones existentes: {len(asignaciones)}')

# 3) Construir Excel
wb = Workbook()
ws = wb.active
ws.title = 'Clientes activos'

# Cabecera
HEADERS = [
    ('id_noofit',         11),
    ('DNI',               14),
    ('Nombre',            22),
    ('Apellidos',         24),
    ('Email',             32),
    ('Teléfono',          14),
    ('Estado NF',         10),
    ('Categoría actual',  18),
    ('Nueva categoría',   22),  # ← columna a rellenar por el usuario
]
fill_h = PatternFill('solid', fgColor='2DD4A8')
font_h = Font(bold=True, color='FFFFFF', size=11)
border = Border(*[Side(style='thin', color='CCCCCC')]*4)

for col, (h, w) in enumerate(HEADERS, 1):
    c = ws.cell(1, col, h); c.fill = fill_h; c.font = font_h
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

# Resaltar la columna "Nueva categoría" (col 9)
fill_new = PatternFill('solid', fgColor='FFF7CC')
font_new = Font(bold=True, color='8B4513', size=11)

# Filas
def fullkey(c):
    return f"{c.get('name') or c.get('nombre','')} {c.get('surname') or c.get('apellidos','')}"
activos_sorted = sorted(activos, key=lambda c: fullkey(c).upper())

for i, c in enumerate(activos_sorted, start=2):
    cid = c.get('id')
    cat_actual = asignaciones.get(str(cid)) or asignaciones.get(cid) or ''
    vals = [
        cid,
        c.get('dni') or c.get('nif') or '',
        c.get('name') or c.get('nombre') or '',
        c.get('surname') or c.get('apellidos') or '',
        c.get('email') or '',
        c.get('cellPhone') or c.get('telefono') or '',
        'Activo' if c.get('enabled') is not False else 'Archivado',
        cat_actual,
        '',   # Nueva categoría — vacío
    ]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(i, j, v); cell.border = border
        if j == 9:
            cell.fill = fill_new

# Validación de dato (lista desplegable) en la columna "Nueva categoría"
if cat_nombres:
    formula = '"' + ','.join([''] + cat_nombres + ['(quitar categoría)']) + '"'
    dv = DataValidation(type='list', formula1=formula, allow_blank=True,
                        showErrorMessage=True,
                        errorTitle='Categoría inválida',
                        error='Elige una de la lista o déjalo vacío.')
    dv.add(f'I2:I{len(activos_sorted) + 1}')
    ws.add_data_validation(dv)

# Hoja "Instrucciones"
ws2 = wb.create_sheet('Instrucciones', 0)
ws2.column_dimensions['A'].width = 90
lines = [
    ('CÓMO USAR ESTE EXCEL', True),
    ('', False),
    ('1. Ve a la hoja "Clientes activos".', False),
    ('2. En la columna "Nueva categoría" (ámbar) elige la categoría desde el desplegable.', False),
    ('   Categorías disponibles: ' + ', '.join(cat_nombres or ['—']), False),
    ('   Para QUITAR la categoría a un cliente, elige "(quitar categoría)".', False),
    ('   Si no quieres cambiar nada para un cliente, déjalo vacío.', False),
    ('3. Guarda el archivo.', False),
    ('4. Súbelo de vuelta — el sistema aplicará los cambios solo a las filas con "Nueva categoría" rellena.', False),
    ('', False),
    ('Nota: el id_noofit (col A) NO se debe modificar — es la clave para localizar al cliente.', True),
]
for i, (text, bold) in enumerate(lines, 1):
    c = ws2.cell(i, 1, text)
    if bold: c.font = Font(bold=True, size=12)

wb.save(OUT)
print(f'\n→ {OUT}')
print(f'  {len(activos_sorted)} clientes activos exportados')
