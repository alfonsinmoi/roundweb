"""Aplica los cambios de categoría desde el Excel rellenado.

Lee /tmp/clientes_categorias_in.xlsx (o el path que pases) y para cada fila
con la columna "Nueva categoría" rellena:
  - "(quitar categoría)" → DELETE en cliente_categoria
  - cualquier otro valor → UPSERT con la categoria_id correspondiente

Modo dry-run: muestra qué haría. Para aplicar de verdad: CONFIRM=1.

Uso:
  IN_XLSX=/tmp/clientes_categorias_in.xlsx \\
  ID_MANAGER=17675 \\
  CONFIRM=1 \\
  /opt/round_config_api/venv/bin/python3 apply_clientes_categorias.py
"""
import os, sys
sys.path.insert(0, '/opt/round_config_api')
from openpyxl import load_workbook
from app.db import get_conn

ID_MANAGER = os.getenv('ID_MANAGER', '17675')
IN_XLSX = os.getenv('IN_XLSX', '/tmp/clientes_categorias_in.xlsx')
CONFIRM = os.getenv('CONFIRM') == '1'

# 1) Cargar mapping nombre → categoria_id
with get_conn() as conn, conn.cursor() as cur:
    cur.execute("SELECT id, nombre FROM categoria WHERE id_manager=%s",
                (str(ID_MANAGER),))
    cat_by_name = {row['nombre'].lower(): row['id'] for row in cur.fetchall()}
print(f'Categorías encontradas: {", ".join(cat_by_name)}')

# 2) Leer Excel
wb = load_workbook(IN_XLSX, data_only=True)
ws = wb['Clientes activos'] if 'Clientes activos' in wb.sheetnames else wb.active

# Identificar columnas por cabecera
headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
required = ['id_noofit', 'Categoría actual', 'Nueva categoría']
for r in required:
    if r not in headers:
        sys.exit(f'Columna "{r}" no encontrada en el Excel')
COL_ID = headers['id_noofit']
COL_ACTUAL = headers['Categoría actual']
COL_NEW = headers['Nueva categoría']
COL_NOMBRE = headers.get('Nombre')
COL_APE = headers.get('Apellidos')

# 3) Iterar filas y construir cambios
to_set = []        # (id_noofit, categoria_id, nombre_label)
to_unset = []      # id_noofit
errors = []

for r in range(2, ws.max_row + 1):
    cid = ws.cell(r, COL_ID).value
    nueva = ws.cell(r, COL_NEW).value
    if not cid: continue
    if not nueva or not str(nueva).strip(): continue
    nueva = str(nueva).strip()
    nombre_cliente = f"{ws.cell(r, COL_NOMBRE).value or ''} {ws.cell(r, COL_APE).value or ''}".strip() if COL_NOMBRE else str(cid)

    if nueva.lower() in ('(quitar categoría)', '(quitar categoria)', 'quitar', 'quitar categoría', 'quitar categoria'):
        to_unset.append((cid, nombre_cliente))
        continue
    cat_id = cat_by_name.get(nueva.lower())
    if not cat_id:
        errors.append(f'fila {r}: categoría "{nueva}" no existe (cliente {nombre_cliente})')
        continue
    to_set.append((cid, cat_id, nueva, nombre_cliente))

print(f'\nResumen:')
print(f'  ✅ Asignar/cambiar: {len(to_set)}')
print(f'  ✖  Quitar:         {len(to_unset)}')
print(f'  ⚠  Errores:        {len(errors)}')
for e in errors: print(f'     · {e}')

if not to_set and not to_unset:
    print('\nNada que hacer.')
    sys.exit(0)

# Mostrar preview
print('\n--- PREVIEW (max 30 filas) ---')
for cid, cat_id, cat_nombre, nombre in to_set[:30]:
    print(f'  SET   {cid:>10}  {nombre[:30]:30s} → {cat_nombre}')
for cid, nombre in to_unset[:10]:
    print(f'  UNSET {cid:>10}  {nombre[:30]:30s}')

if not CONFIRM:
    print('\n[INFO] Modo dry-run. Para aplicar de verdad, lanza con CONFIRM=1.')
    sys.exit(0)

# 4) Aplicar
print('\n=== APLICANDO ===')
with get_conn() as conn, conn.cursor() as cur:
    for cid, cat_id, cat_nombre, nombre in to_set:
        cur.execute("""
            INSERT INTO cliente_categoria (id_manager, cliente_idnoofit, categoria_id)
            VALUES (%s, %s, %s)
            ON CONFLICT (id_manager, cliente_idnoofit) DO UPDATE
              SET categoria_id = EXCLUDED.categoria_id,
                  updated_at = CURRENT_TIMESTAMP
        """, (str(ID_MANAGER), str(cid), cat_id))
    for cid, nombre in to_unset:
        cur.execute("""
            DELETE FROM cliente_categoria
             WHERE id_manager=%s AND cliente_idnoofit=%s
        """, (str(ID_MANAGER), str(cid)))

print(f'✅ Aplicados: {len(to_set)} asignaciones + {len(to_unset)} quitadas.')
