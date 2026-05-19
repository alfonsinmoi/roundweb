"""Preview v2: filtrado a recibos 2026, con detección de periodicidad por
cliente y consistencia. Genera Excel listo para importar (excluyendo
anomalías bloqueantes).

Periodicidad por recibo (desde fechaDesde a fechaHasta):
  - 25-35 días   → mensual
  - 80-100 días  → trimestral
  - 150-200 días → semestral
  - 300-400 días → anual
  - resto        → otro (bono / cargo puntual)

Anomalías bloqueantes (NO se importan):
  - importe_invalido (importe inicial <= 0)
  - importe_inconsistente (final > inicial)
  - sin_fechas
  - sin_match_nf (cliente no localizable en NoofitPro)

Anomalías informativas (SÍ se importan, marcadas):
  - cobrado_sin_banco → cobro en caja/efectivo
  - devuelto         → impagado tras devolución bancaria
  - periodicidad_inconsistente → cliente con mezcla de periodicidades

Filtro temporal: fechaDesde dentro del año 2026 (cualquier mes/trimestre).
"""
import json, sys, re, unicodedata
from datetime import datetime, date
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GP = '/opt/round_config_api/gestplus_dump_2026-05-08.json'
NF = '/opt/round_config_api/noofit_clientes_dump.json'
OUT = '/tmp/preview_recibos_2026_v2.xlsx'
TARGET_YEAR = 2026


# ─── Helpers ─────────────────────────────────────────────────────────────────
def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s or '')) if unicodedata.category(c) != 'Mn')
def norm_dni(d):
    s = strip_accents(str(d or '')).upper()
    return re.sub(r'[^A-Z0-9]', '', s) if s else ''
def norm_email(e):
    return (e or '').strip().lower()


def parse_dt(s):
    if not s: return None
    try:
        return datetime.fromisoformat(str(s).replace('Z', '')).date()
    except Exception:
        try: return datetime.strptime(str(s)[:10], '%Y-%m-%d').date()
        except Exception: return None


def detectar_periodicidad(fd, fh):
    """Devuelve ('mensual'|'trimestral'|'semestral'|'anual'|'otro', días)."""
    if not fd or not fh: return ('desconocida', None)
    dias = (fh - fd).days + 1   # +1 para incluir día último
    if 25 <= dias <= 35: return ('mensual', dias)
    if 80 <= dias <= 100: return ('trimestral', dias)
    if 150 <= dias <= 200: return ('semestral', dias)
    if 300 <= dias <= 400: return ('anual', dias)
    return ('otro', dias)


# ─── Cargar dumps ────────────────────────────────────────────────────────────
gp = json.load(open(GP, 'r', encoding='utf-8'))
nf = json.load(open(NF, 'r', encoding='utf-8'))
nf_clientes = nf.get('clientes', [])

nf_by_dni = {}
nf_by_email = {}
for c in nf_clientes:
    d = norm_dni(c.get('dni') or c.get('nif'))
    if d and len(d) >= 7: nf_by_dni.setdefault(d, c)
    e = norm_email(c.get('email'))
    if e and '@' in e: nf_by_email.setdefault(e, c)

gp_clients_all = (gp.get('altas') or []) + (gp.get('bajas_recientes_12m') or [])

# ─── Recoger recibos 2026 ────────────────────────────────────────────────────
recibos = []
for c in gp_clients_all:
    for r in c.get('_recibos', []):
        fd = parse_dt(r.get('fechaDesde'))
        fh = parse_dt(r.get('fechaHasta'))
        # Filtro temporal: fechaDesde DEBE estar en 2026
        if not fd or fd.year != TARGET_YEAR:
            continue
        recibos.append({**r,
            '_codcli': c.get('codigo'),
            '_dni': c.get('dni'),
            '_dni_contr': c.get('dniContr'),
            '_nombre': f"{c.get('nombre','')} {c.get('apellidos','')}".strip(),
            '_email': c.get('email'),
            '_iban': c.get('_iban'),
            '_estado_cliente': 'Alta' if c.get('estado') == 1 else 'Baja',
            '_fechaBajaCli': c.get('fechaBaja'),
            '_fd': fd, '_fh': fh,
        })
print(f'Recibos con fechaDesde en {TARGET_YEAR}: {len(recibos)}')

# ─── Cruce con NoofitPro ─────────────────────────────────────────────────────
def find_in_nf(r):
    for k in ('_dni', '_dni_contr'):
        d = norm_dni(r.get(k))
        if d and len(d) >= 7 and d in nf_by_dni:
            return nf_by_dni[d], 'dni'
    e = norm_email(r.get('_email'))
    if e and e in nf_by_email:
        return nf_by_email[e], 'email'
    return None, None

for r in recibos:
    nf_match, method = find_in_nf(r)
    r['_nf_id'] = nf_match.get('id') if nf_match else None
    r['_nf_match_by'] = method
    r['_nf_name'] = f"{nf_match.get('name','')} {nf_match.get('surname','')}".strip() if nf_match else ''
    p, dias = detectar_periodicidad(r['_fd'], r['_fh'])
    r['_periodicidad'] = p
    r['_dias'] = dias
    r['_metodo_pago'] = ('banco' if (r.get('importeBanco') or 0) > 0
                         else 'caja' if r.get('cobrado') == 1
                         else 'pendiente')


# ─── Detectar anomalías bloqueantes y consistencia por cliente ───────────────
BLOCK_TYPES = {'importe_invalido', 'importe_inconsistente', 'sin_fechas', 'sin_match_nf'}

def anomalias_de(r):
    out = []
    if not r.get('_nf_id'):
        out.append(('sin_match_nf', 'Cliente no localizado en NoofitPro'))
    if not r.get('_fd') or not r.get('_fh'):
        out.append(('sin_fechas', f'fd={r.get("fechaDesde")} fh={r.get("fechaHasta")}'))
    ii = r.get('importeInicial') or 0
    if not isinstance(ii, (int, float)) or ii <= 0:
        out.append(('importe_invalido', f'inicial={ii}'))
    f = r.get('importeFinal') or 0
    if f > ii * 1.01:
        out.append(('importe_inconsistente', f'final={f} > inicial={ii}'))
    return out

for r in recibos:
    r['_anom'] = anomalias_de(r)
    r['_anom_block'] = any(a[0] in BLOCK_TYPES for a in r['_anom'])

# Agrupar por cliente y detectar periodicidad dominante / mezcla
recibos_por_cli = defaultdict(list)
for r in recibos:
    recibos_por_cli[r['_codcli']].append(r)

cliente_periodicidad_summary = {}
for cod, lst in recibos_por_cli.items():
    vals = [r['_periodicidad'] for r in lst if r['_periodicidad'] != 'desconocida']
    cnt = Counter(vals)
    if not cnt:
        cliente_periodicidad_summary[cod] = ('desconocida', cnt, False)
        continue
    dominant, n_dom = cnt.most_common(1)[0]
    inconsistente = (len(cnt) > 1)
    cliente_periodicidad_summary[cod] = (dominant, cnt, inconsistente)


# ─── Construir filas import-ready (excluye anomalías bloqueantes) ────────────
import_ready = [r for r in recibos if not r['_anom_block']]
print(f'Import-ready: {len(import_ready)} (descartados {len(recibos) - len(import_ready)} por anomalía bloqueante)')

# ─── Excel ───────────────────────────────────────────────────────────────────
wb = Workbook(); wb.remove(wb.active)
border = Border(*[Side(style='thin', color='CCCCCC')]*4)

def hsheet(name, headers, color='2DD4A8'):
    ws = wb.create_sheet(name)
    fill_h = PatternFill('solid', fgColor=color)
    font_h = Font(bold=True, color='FFFFFF', size=11)
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.fill = fill_h; c.font = font_h
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
    return ws


# RESUMEN
ws = wb.create_sheet('RESUMEN', 0)
for col, w in zip('ABC', [38, 18, 60]):
    ws.column_dimensions[col].width = w
ws.cell(1, 1, f'PREVIEW v2 — Recibos {TARGET_YEAR} → Odoo').font = Font(bold=True, size=14)

def w_kv(row, k, v, note=''):
    ws.cell(row, 1, k).font = Font(bold=True)
    ws.cell(row, 2, v)
    if note: ws.cell(row, 3, note).font = Font(italic=True, color='888888')

# Stats
n_total = len(recibos)
n_block = sum(1 for r in recibos if r['_anom_block'])
n_ready = len(import_ready)
n_pagados = sum(1 for r in import_ready if r.get('cobrado') == 1)
n_imp = n_ready - n_pagados
n_caja = sum(1 for r in import_ready if r['_metodo_pago'] == 'caja')
n_banco = sum(1 for r in import_ready if r['_metodo_pago'] == 'banco')
n_pend_pago = sum(1 for r in import_ready if r['_metodo_pago'] == 'pendiente')
total_emitido = sum(float(r.get('importeFinal') or 0) for r in import_ready)
total_cobrado_banco = sum(float(r.get('importeBanco') or 0) for r in import_ready)
total_cobrado_caja = sum(float(r.get('importeFinal') or 0) for r in import_ready
                          if r.get('cobrado') == 1 and (r.get('importeBanco') or 0) == 0)

w_kv(3, 'Recibos con fechaDesde 2026',     n_total)
w_kv(4, '  Anomalías BLOQUEANTES',          n_block, 'no se importan')
w_kv(5, '  IMPORT-READY',                   n_ready, '✓ se importarán')
w_kv(7, 'De los import-ready:', '')
ws.cell(7, 1).font = Font(bold=True, color='1A9A7A')
w_kv(8, '  Pagados (cobrado=1)',            n_pagados)
w_kv(9, '    via banco',                    n_banco, f'{total_cobrado_banco:.2f} €')
w_kv(10, '   via caja/efectivo',            n_caja, f'{total_cobrado_caja:.2f} €')
w_kv(11, '  Impagados / pendientes',        n_imp)
w_kv(13, 'TOTALES (€)', '')
ws.cell(13, 1).font = Font(bold=True)
w_kv(14, '  Total emitido',                 f'{total_emitido:.2f}')
w_kv(15, '  Total cobrado',                 f'{total_cobrado_banco + total_cobrado_caja:.2f}')
w_kv(16, '  Total pendiente',               f'{total_emitido - total_cobrado_banco - total_cobrado_caja:.2f}')

w_kv(18, 'PERIODICIDADES detectadas', '')
ws.cell(18, 1).font = Font(bold=True)
periods = Counter(r['_periodicidad'] for r in import_ready)
row = 19
for p, n in periods.most_common():
    w_kv(row, f'  {p}', n)
    row += 1

w_kv(row + 1, 'CLIENTES con MEZCLA de periodicidades', '')
ws.cell(row + 1, 1).font = Font(bold=True, color='C00000')
n_mezcla = sum(1 for cod, (_, _, mix) in cliente_periodicidad_summary.items() if mix)
w_kv(row + 2, '  Total clientes inconsistentes', n_mezcla, '⚠️ ver hoja PERIODICIDAD_CLIENTE')


# IMPORT_READY (los que se importarán)
ws = hsheet('IMPORT_READY', [
    ('numRec', 11), ('Cód cli', 11), ('Cliente', 30), ('DNI', 14), ('Email', 26),
    ('NF id', 11), ('Curso', 14),
    ('Pagado', 8), ('Método', 11), ('Periodicidad', 13),
    ('Inicial', 9), ('Final', 9), ('Banco', 9),
    ('Desde', 11), ('Hasta', 11),
    ('Devolución', 11), ('Recobro', 11),
    ('Match', 8),
], color='2DD4A8')
for i, r in enumerate(import_ready, 2):
    vals = [
        r.get('numRec'), r.get('_codcli'), r.get('_nombre'),
        r.get('_dni') or '', r.get('_email') or '',
        r.get('_nf_id') or '', r.get('codcur') or '',
        'Sí' if r.get('cobrado') == 1 else 'No',
        r['_metodo_pago'],
        r['_periodicidad'],
        r.get('importeInicial'), r.get('importeFinal'), r.get('importeBanco'),
        str(r['_fd']) if r['_fd'] else '',
        str(r['_fh']) if r['_fh'] else '',
        (r.get('fechDevolucion') or '')[:10] if r.get('fechDevolucion') else '',
        (r.get('fechRecobro') or '')[:10] if r.get('fechRecobro') else '',
        r.get('_nf_match_by') or '',
    ]
    for j, v in enumerate(vals, 1):
        ws.cell(i, j, v).border = border
ws.auto_filter.ref = ws.dimensions


# PERIODICIDAD_CLIENTE
ws = hsheet('PERIODICIDAD_CLIENTE', [
    ('Cód GP', 11), ('Cliente', 30), ('NF id', 11),
    ('Periodicidad dominante', 22),
    ('Mensual', 9), ('Trimestral', 11), ('Semestral', 10), ('Anual', 8), ('Otro', 8),
    ('¿Inconsistente?', 16),
], color='FBBF24')
filas = []
for cod, (dom, cnt, mix) in cliente_periodicidad_summary.items():
    primero = recibos_por_cli[cod][0]
    filas.append([
        cod, primero.get('_nombre'), primero.get('_nf_id') or '',
        dom,
        cnt.get('mensual', 0), cnt.get('trimestral', 0), cnt.get('semestral', 0),
        cnt.get('anual', 0), cnt.get('otro', 0),
        '⚠️ SÍ' if mix else 'No',
    ])
filas.sort(key=lambda x: (-1 if x[9].startswith('⚠️') else 0, x[1] or ''))
for i, vals in enumerate(filas, 2):
    for j, v in enumerate(vals, 1):
        c = ws.cell(i, j, v); c.border = border
        if j == 10 and v.startswith('⚠️'):
            c.fill = PatternFill('solid', fgColor='FFE4E4')
ws.auto_filter.ref = ws.dimensions


# DESCARTADOS_BLOCK
ws = hsheet('DESCARTADOS', [
    ('Razones', 36), ('numRec', 11), ('Cód cli', 11), ('Cliente', 30),
    ('DNI', 14), ('Email', 26), ('Inicial', 9), ('Final', 9), ('Banco', 9),
    ('Desde', 11), ('Hasta', 11),
], color='F87171')
for i, r in enumerate([r for r in recibos if r['_anom_block']], 2):
    razones = '; '.join(t for t, _ in r['_anom'] if t in BLOCK_TYPES)
    detalles = '; '.join(d for t, d in r['_anom'] if t in BLOCK_TYPES)
    vals = [
        f'{razones}: {detalles}',
        r.get('numRec'), r.get('_codcli'), r.get('_nombre'),
        r.get('_dni') or '', r.get('_email') or '',
        r.get('importeInicial'), r.get('importeFinal'), r.get('importeBanco'),
        str(r['_fd']) if r['_fd'] else '', str(r['_fh']) if r['_fh'] else '',
    ]
    for j, v in enumerate(vals, 1):
        ws.cell(i, j, v).border = border
ws.auto_filter.ref = ws.dimensions

wb.save(OUT)
print(f'\n→ {OUT}')
print(f'  RESUMEN | IMPORT_READY ({n_ready}) | PERIODICIDAD_CLIENTE ({len(filas)}) | DESCARTADOS ({n_block})')
print(f'\n  Pagados banco: {n_banco} ({total_cobrado_banco:.2f} €)')
print(f'  Pagados caja:  {n_caja} ({total_cobrado_caja:.2f} €)')
print(f'  Impagados:     {n_imp}')
print(f'  Periodicidad mezclada: {n_mezcla} clientes')
