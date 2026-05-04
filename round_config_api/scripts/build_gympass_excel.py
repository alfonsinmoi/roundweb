"""Genera Excel con SOLO los clientes Gympass.

Origen:
  - NoofitPro: clientes con gympassId no nulo/vacío
  - GestPlus: clientes con gympassId en su registro

Crucemos por DNI/email/nombre. Salida con columnas para verificar:
  - Cliente, gympassId NoofitPro, gympassId GestPlus
  - DNI, Email
  - Estado en cada sistema
  - Reservas NF, Recibos 2026 GP
"""
import os, sys, json, hashlib, time, unicodedata, re
from datetime import datetime
from collections import defaultdict
import requests, urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
VERIFY = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = 'https://pro.wiemspro.com/wiemspro'
APP_VERSION = '1.8.39'
APP_ID = '1'
EMAIL = os.getenv('NOOFIT_EMAIL', '')
PWD   = os.getenv('NOOFIT_PASS', '')
GP_DUMP = os.getenv('GP_DUMP', 'gestplus_dump_2026-05-02.json')
NF_DUMP = os.getenv('NF_DUMP', 'noofit_clientes_dump.json')
OUT = os.getenv('OUT', 'gympass_clientes.xlsx')


def banner(s): print(f'\n{"="*60}\n{s}\n{"="*60}')


def login():
    body = {'email': EMAIL, 'appVersion': APP_VERSION,
            'password': hashlib.md5(PWD.encode()).hexdigest().upper()}
    r = requests.post(f'{BASE}/account/loginEasy', json=body, verify=VERIFY, timeout=30)
    if r.status_code != 200: sys.exit('Login fallido')
    return r.headers.get('X-CustomToken'), r.headers.get('X-TRAINER_MANAGER', '')


def hdrs(t, m, extra=None):
    h = {'X-CustomToken':t,'X-TRAINER_MANAGER':m,'locale':'es',
         'appVersion':APP_VERSION,'appId':APP_ID,'Content-Type':'application/json'}
    if extra: h.update(extra)
    return h


def get_reservas(t, m, cid):
    try:
        r = requests.post(f'{BASE}/api/dispositivos/getReservasByUser',
                          json={'id':cid}, headers=hdrs(t, m, {'initialId':'0'}),
                          verify=VERIFY, timeout=30)
        d = r.json()
        if d.get('mensaje') != 'OK': return 0
        return len(d.get('clases') or d.get('reservas') or [])
    except Exception: return 0


def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')


def norm_dni(d):
    return re.sub(r'[^A-Z0-9]', '', strip_accents(str(d or '')).upper())


def norm_email(e): return (e or '').strip().lower()


def fullname_gp(g): return f"{g.get('nombre','') or ''} {g.get('apellidos','') or ''}".strip()


def fullname_nf(n): return f"{n.get('nombre') or n.get('name') or ''} {n.get('apellidos') or n.get('surname') or ''}".strip()


def fmt_fecha(s):
    if not s: return ''
    try: return datetime.fromisoformat(str(s).replace('Z','')).strftime('%d/%m/%Y')
    except Exception: return str(s)[:10]


def main():
    banner('1) Cargar dumps')
    with open(GP_DUMP, 'r', encoding='utf-8') as f: gp = json.load(f)
    with open(NF_DUMP, 'r', encoding='utf-8') as f: nf = json.load(f)

    nf_clientes = list({c['id']:c for c in (nf.get('clientes') or [])}.values())
    gp_full = gp.get('altas', []) + gp.get('bajas_recientes_12m', [])

    PATRON = re.compile(r'wellhub|gympass|wellpass', re.I)
    def es_gympass_nf(c):
        fields = ' '.join(str(c.get(k) or '') for k in ('alias','name','nombre','surname','apellidos','email'))
        return bool(PATRON.search(fields))
    def es_gympass_gp(c):
        fields = ' '.join(str(c.get(k) or '') for k in ('nombre','apellidos','email','notas'))
        return bool(PATRON.search(fields))

    nf_gympass = [c for c in nf_clientes if es_gympass_nf(c)]
    gp_gympass = [g for g in gp_full if es_gympass_gp(g)]
    print(f'  NoofitPro con marca Wellhub/Gympass: {len(nf_gympass)}')
    print(f'  GestPlus con marca Wellhub/Gympass:  {len(gp_gympass)}')

    # Indexar GestPlus por DNI/email
    gp_by_dni = {}
    gp_by_email = {}
    for g in gp_full:
        for k in ('dni','dniContr'):
            d = norm_dni(g.get(k))
            if d and len(d) >= 8: gp_by_dni.setdefault(d, g)
        e = norm_email(g.get('email'))
        if e and '@' in e: gp_by_email.setdefault(e, g)

    # Combinar: union de NF gympass + GP gympass
    filas = []
    matched_gp_codes = set()

    banner('2) Login NoofitPro y consulta reservas')
    token, manager = login()

    # NF con gympass → cruzar con GP por DNI / email
    for n in nf_gympass:
        d = norm_dni(n.get('dni') or n.get('nif'))
        e = norm_email(n.get('email'))
        g = None
        match_via = ''
        if d and len(d) >= 8 and d in gp_by_dni:
            g = gp_by_dni[d]; match_via = 'dni'
        elif e and e in gp_by_email:
            g = gp_by_email[e]; match_via = 'email'
        if g: matched_gp_codes.add(g['codigo'])

        n_res = get_reservas(token, manager, n['id'])

        archivado = n.get('enabled') is False or n.get('enabled') == 0
        recibos = len(g.get('_recibos') or []) if g else 0

        # estado GP
        if not g:
            estado_gp = '— no en GP —'
        elif g.get('estado') == 1:
            estado_gp = f"Alta {fmt_fecha(g.get('fechaAlta'))}"
        else:
            estado_gp = f"Baja {fmt_fecha(g.get('fechaBaja'))}"

        # acción sugerida
        if archivado:
            accion = '📦 Ya archivado en Noofit'
        elif n_res == 0 and not g:
            accion = '⚠️ Gympass NF sin GP y sin reservas → revisar'
        elif n_res == 0:
            accion = '⚠️ Gympass sin reservas → revisar'
        elif g and g.get('estado') == 0:
            accion = '⚠️ Activo en NF, baja en GP → revisar'
        else:
            accion = '✅ OK Gympass activo'

        filas.append({
            'cliente_nf': fullname_nf(n) or '(sin nombre)',
            'cliente_gp': fullname_gp(g) if g else '',
            'alias_nf': n.get('alias') or '',
            'notas_gp': (g.get('notas') or '')[:80] if g else '',
            'dni': n.get('dni') or n.get('nif') or '',
            'email': n.get('email') or '',
            'codigo_gp': g.get('codigo') if g else '',
            'noofit_id': n['id'],
            'archivado_nf': 'Sí' if archivado else 'No',
            'estado_gp': estado_gp,
            'reservas_nf': n_res,
            'recibos_2026': recibos,
            'match_via': match_via or 'sin match GP',
            'accion': accion,
        })
        time.sleep(0.05)

    # GP con gympass que NO se matchearon con un NF gympass
    for g in gp_gympass:
        if g.get('codigo') in matched_gp_codes: continue
        # Pero podría matchear con un NF que NO tiene gympassId — buscar en todo NF
        nf_dni = {norm_dni(c.get('dni') or c.get('nif')): c for c in nf_clientes if (c.get('dni') or c.get('nif'))}
        d1 = norm_dni(g.get('dni'))
        d2 = norm_dni(g.get('dniContr'))
        n = None
        for d in (d1, d2):
            if d and len(d) >= 8 and d in nf_dni:
                n = nf_dni[d]; break
        accion = '⚠️ GP Gympass sin Gympass en Noofit → sincronizar'
        n_res = get_reservas(token, manager, n['id']) if n else 0
        if n: time.sleep(0.05)
        filas.append({
            'cliente_nf': fullname_nf(n) if n else '(no en NF)',
            'cliente_gp': fullname_gp(g),
            'alias_nf': n.get('alias') if n else '',
            'notas_gp': (g.get('notas') or '')[:80],
            'dni': g.get('dni') or '',
            'email': g.get('email') or '',
            'codigo_gp': g['codigo'],
            'noofit_id': n['id'] if n else '',
            'archivado_nf': ('Sí' if (n and n.get('enabled') is False) else 'No') if n else '— no en NF —',
            'estado_gp': f"Alta {fmt_fecha(g.get('fechaAlta'))}" if g.get('estado') == 1 else f"Baja {fmt_fecha(g.get('fechaBaja'))}",
            'reservas_nf': n_res if n else '',
            'recibos_2026': len(g.get('_recibos') or []),
            'match_via': 'gp_only' if not n else 'dni',
            'accion': accion,
        })

    write_excel(filas)
    from collections import Counter
    print('\nResumen:')
    for k,v in Counter(f['accion'] for f in filas).most_common():
        print(f'  {v:4d}  {k}')


def write_excel(filas):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Gympass'
    headers = [
        ('Cliente NoofitPro', 28),
        ('Cliente GestPlus', 28),
        ('Alias NF (Wellhub)', 30),
        ('Notas GP', 32),
        ('DNI', 13),
        ('Email', 28),
        ('Cód. GestPlus', 12),
        ('ID NoofitPro', 11),
        ('Archivado NF', 11),
        ('Estado GestPlus', 22),
        ('Reservas NF', 11),
        ('Recibos 2026 GP', 13),
        ('Match', 10),
        ('Acción sugerida', 40),
    ]
    keys = ['cliente_nf','cliente_gp','alias_nf','notas_gp','dni','email',
            'codigo_gp','noofit_id','archivado_nf','estado_gp','reservas_nf',
            'recibos_2026','match_via','accion']
    fill_h = PatternFill('solid', fgColor='2DD4A8')
    font_h = Font(bold=True, color='FFFFFF', size=11)
    border = Border(*[Side(style='thin', color='CCCCCC')]*4)
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_h; c.font = font_h
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    f_si = PatternFill('solid', fgColor='D1FADF')
    f_warn = PatternFill('solid', fgColor='FFF7CC')
    f_no = PatternFill('solid', fgColor='FFE4E4')
    for i, f in enumerate(filas, 2):
        for j, k in enumerate(keys, 1):
            cell = ws.cell(row=i, column=j, value=f.get(k, ''))
            cell.border = border
            if k == 'accion':
                v = str(f.get(k,''))
                if 'OK' in v: cell.fill = f_si
                elif 'archivado' in v.lower() or 'revisar' in v.lower(): cell.fill = f_warn
                elif 'sincronizar' in v.lower(): cell.fill = f_no
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUT)
    print(f'  → {OUT}')


if __name__ == '__main__':
    main()
