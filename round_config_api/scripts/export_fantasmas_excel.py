"""Genera Excel con los 9 fantasmas que parecen clientes reales de Round Málaga.
Una hoja resumen por cliente + una hoja con TODAS sus reservas individuales."""
from datetime import date, timedelta
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.db import get_conn
from app import noofit_client as nc

OBJETIVOS = {
    1818777: 'fantasma',  # Chari Rodríguez
    1819182: 'fantasma',  # Javier López Pastrana
    1819838: 'fantasma',  # rafael-ga
    1819986: 'fantasma',  # Joaquín-Ál
    1817691: 'cross_17677',  # Cristobal
    1817692: 'cross_17677',  # Ignaciogp87
    1817751: 'cross_17677',  # Martaita19
    1817933: 'cross_17677',  # Maida
    1817351: 'cross_17679',  # elnik
}

hoy = date.today()
desde = (hoy - timedelta(days=365)).isoformat() + 'T00:00:00+02:00'
hasta = (hoy + timedelta(days=1)).isoformat() + 'T00:00:00+02:00'
reservas = nc.get_reservas_confirmadas(desde, hasta) or []
# Filtrar a los objetivos
por_cli = {cid: [] for cid in OBJETIVOS}
for r in reservas:
    try: cid = int(r['cliente_id'])
    except (TypeError, ValueError): continue
    if cid in OBJETIVOS:
        por_cli[cid].append(r)

# Lookup cross-manager nombres/emails/dni
with get_conn() as conn, conn.cursor() as cur:
    ph = ','.join(['%s']*len(OBJETIVOS))
    cur.execute(f"SELECT id, id_manager, raw_data->>'name' AS n, "
                f"raw_data->>'surname' AS s, raw_data->>'email' AS e, "
                f"raw_data->>'dni' AS d "
                f"FROM cliente_cache WHERE id IN ({ph})", list(OBJETIVOS))
    cross_data = {int(r['id']): r for r in cur.fetchall()}

wb = Workbook()
header_fill = PatternFill('solid', fgColor='1F7A5C')
header_font = Font(bold=True, color='FFFFFF')

# ─── Hoja 1: Resumen ─────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Resumen'
cols = ['Tipo', 'Cliente ID', 'Nombre (NoofitPro)', 'Nombre (cache cross-mgr)',
        'Email', 'DNI', '# Reservas', '1ª Reserva', 'Última Reserva',
        '# Actividades distintas', 'Actividades', 'En cache de manager']
ws.append(cols)
for cell in ws[1]:
    cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(vertical='center')

for cid, tipo in OBJETIVOS.items():
    rs = por_cli.get(cid, [])
    nombre_nf = next((r.get('cliente_nombre') for r in rs if r.get('cliente_nombre')), '')
    cd = cross_data.get(cid)
    nombre_cache = (f"{(cd['n'] or '').strip()} {(cd['s'] or '').strip()}".strip()
                    if cd else '')
    email = cd['e'] if cd else ''
    dni = cd['d'] if cd else ''
    en_cache = cd['id_manager'] if cd else ''
    fechas = sorted({r['fecha'] for r in rs if r.get('fecha')})
    acts = sorted({r['actividad_nombre'] for r in rs if r.get('actividad_nombre')})
    ws.append([
        tipo, cid, nombre_nf, nombre_cache, email, dni,
        len(rs),
        fechas[0] if fechas else '',
        fechas[-1] if fechas else '',
        len(acts),
        ', '.join(acts),
        en_cache,
    ])

# Anchos
widths = [14, 12, 24, 26, 28, 14, 11, 12, 12, 8, 70, 18]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64+i)].width = w
ws.freeze_panes = 'A2'

# ─── Hoja 2: Reservas detalladas ─────────────────────────────────────────────
ws2 = wb.create_sheet('Reservas detalladas')
cols2 = ['Cliente ID', 'Nombre', 'Fecha', 'Hora', 'Actividad', 'Sala ID', 'idTrainer']
ws2.append(cols2)
for cell in ws2[1]:
    cell.fill = header_fill; cell.font = header_font

for cid in OBJETIVOS:
    rs = sorted(por_cli.get(cid, []), key=lambda r: (r.get('fecha') or '', r.get('hora') or ''))
    nombre = next((r.get('cliente_nombre') for r in rs if r.get('cliente_nombre')), '')
    for r in rs:
        ws2.append([
            cid, nombre,
            r.get('fecha') or '',
            r.get('hora') or '',
            r.get('actividad_nombre') or '',
            r.get('sala_id') or '',
            str(r.get('id_trainer') or ''),
        ])

widths2 = [12, 24, 12, 8, 32, 10, 10]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[chr(64+i)].width = w
ws2.freeze_panes = 'A2'

# ─── Hoja 3: Análisis / contexto ─────────────────────────────────────────────
ws3 = wb.create_sheet('Análisis')
notas = [
    ['Reporte: 9 clientes reservando en clases de Round Málaga Centro (trainer 17675) sin estar en cache local'],
    ['Generado: ' + hoy.isoformat()],
    [''],
    ['Tipo "fantasma" → NO está en cache local de NINGÚN manager.'],
    ['Tipo "cross_17677" → está en cache del manager 17677 (pruebasnoofit) pero reserva habitualmente en Round Málaga real.'],
    ['Tipo "cross_17679" → está en cache del manager 17679 (pruebas) pero reserva en Round Málaga.'],
    [''],
    ['Causa raíz:'],
    ['  NoofitPro devuelve a roundmalagacentro@noofit.com solo los clientes que ESE login creó (310).'],
    ['  Estos 9 los creó otra cuenta (manager parent, otro trainer, o cuenta administrativa de NoofitPro),'],
    ['  pero reservan en clases del trainer 17675. Por eso no aparecen en getClienteSimple ni en nuestra cache.'],
    [''],
    ['Acción recomendada:'],
    ['  1. Verificar con NoofitPro/Round que estos 9 son clientes reales de Round Málaga Centro.'],
    ['  2. Pedir a NoofitPro que los reasocie/duplique bajo el trainer 17675 (roundmalagacentro).'],
    ['  3. El cron horario los traerá automáticamente a cache cuando NoofitPro los vincule.'],
]
for row in notas:
    ws3.append(row)
ws3.column_dimensions['A'].width = 110

import sys
out = sys.argv[1] if len(sys.argv) > 1 else '/tmp/fantasmas_round_malaga.xlsx'
wb.save(out)
print(f'OK: {out}  ({len(OBJETIVOS)} clientes, '
      f'{sum(len(v) for v in por_cli.values())} reservas)')
