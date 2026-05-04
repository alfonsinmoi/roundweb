"""Comparativa GestPlus ↔ NoofitPro v2.

Mejoras vs v1:
  - Matching por DNI / email / nombre tokens (apellidos sub-string).
  - Columna "Reservas NoofitPro" para huérfanos (count via API).
  - Columna "Recibos 2026 GP" para clientes solo en GestPlus.
  - Lee las anotaciones manuales del usuario (col 7 'Alta en GestPlus' del
    Excel anterior): "no", "trabajador", o un código GP (ej. "6174").

Uso:
  NOOFIT_EMAIL=... NOOFIT_PASS=... \
  python build_excel_v2.py
"""
import os, sys, json, hashlib, time, unicodedata, re
from datetime import datetime
from collections import defaultdict
import requests, urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
VERIFY = False

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = 'https://pro.wiemspro.com/wiemspro'
APP_VERSION = '1.8.39'
APP_ID = '1'
EMAIL = os.getenv('NOOFIT_EMAIL', '')
PWD   = os.getenv('NOOFIT_PASS', '')
GP_DUMP = os.getenv('GP_DUMP', 'gestplus_dump_2026-05-02.json')
NF_DUMP = os.getenv('NF_DUMP', 'noofit_clientes_dump.json')
USER_EDIT = os.getenv('USER_EDIT', 'comparativa_user_edit.xlsx')
OUT = os.getenv('OUT', 'comparativa_v2.xlsx')


def banner(s): print(f'\n{"="*60}\n{s}\n{"="*60}')


def login():
    body = {'email': EMAIL, 'appVersion': APP_VERSION,
            'password': hashlib.md5(PWD.encode()).hexdigest().upper()}
    r = requests.post(f'{BASE}/account/loginEasy', json=body, verify=VERIFY, timeout=30)
    if r.status_code != 200: sys.exit('Login fallido: ' + r.text[:200])
    return r.headers.get('X-CustomToken'), r.headers.get('X-TRAINER_MANAGER', '')


def hdrs(token, manager, extra=None):
    h = {'X-CustomToken':token,'X-TRAINER_MANAGER':manager,'locale':'es',
         'appVersion':APP_VERSION,'appId':APP_ID,'Content-Type':'application/json'}
    if extra: h.update(extra)
    return h


def get_reservas_count(token, manager, id_cliente):
    try:
        r = requests.post(f'{BASE}/api/dispositivos/getReservasByUser',
                          json={'id': id_cliente},
                          headers=hdrs(token, manager, {'initialId':'0'}),
                          verify=VERIFY, timeout=30)
        if r.status_code != 200: return 0
        d = r.json()
        if d.get('mensaje') != 'OK': return 0
        return len(d.get('clases') or d.get('reservas') or [])
    except Exception:
        return 0


# ── Normalización ──────────────────────────────────────────────────────────
STOPWORDS = {'DE','DA','DEL','LA','LAS','LOS','EL','MC','MAC','VAN','VON',
             'SAN','SANTO','SANTA','DI','DOS','DAS'}

def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def norm_text(s):
    s = strip_accents(str(s or '')).upper()
    return re.sub(r'[^A-Z0-9 ]+', ' ', s).strip()

def name_tokens(s):
    return [t for t in norm_text(s).split() if t and t not in STOPWORDS and len(t) > 1]

def norm_dni(d):
    return re.sub(r'[^A-Z0-9]', '', strip_accents(str(d or '')).upper())

def norm_email(e):
    return (e or '').strip().lower()


# ── Matching ───────────────────────────────────────────────────────────────
def fullname_gp(g):
    return f"{g.get('nombre','') or ''} {g.get('apellidos','') or ''}".strip()

def fullname_nf(n):
    return f"{n.get('nombre') or n.get('name') or ''} {n.get('apellidos') or n.get('surname') or ''}".strip()


def build_match(gp_clientes, nf_clientes):
    """Devuelve dict gp_codigo → noofit_cliente (o None)."""
    # Indices NoofitPro
    nf_by_dni = {}
    nf_by_email = {}
    nf_by_tokens = []  # lista de (set_tokens, nf)
    for n in nf_clientes:
        d = norm_dni(n.get('dni') or n.get('nif'))
        if d and len(d) >= 8: nf_by_dni.setdefault(d, n)
        e = norm_email(n.get('email'))
        if e and '@' in e: nf_by_email.setdefault(e, n)
        toks = set(name_tokens(fullname_nf(n)))
        if len(toks) >= 2:
            nf_by_tokens.append((toks, n))

    matched = {}
    methods_count = defaultdict(int)
    nf_used_ids = set()

    def use_match(gp_codigo, n, method):
        # Permitir re-uso si nombre+token (mismo cliente noofit)
        matched[gp_codigo] = n
        methods_count[method] += 1
        nf_used_ids.add(n['id'])

    for g in gp_clientes:
        cod = g.get('codigo')
        # 1) DNI
        for k in ('dni', 'dniContr'):
            d = norm_dni(g.get(k))
            if d and len(d) >= 8 and d in nf_by_dni:
                use_match(cod, nf_by_dni[d], 'dni')
                break
        if cod in matched: continue
        # 2) Email
        e = norm_email(g.get('email'))
        if e and '@' in e and e in nf_by_email:
            use_match(cod, nf_by_email[e], 'email')
            continue
        # 3) Tokens nombre — Jaccard >= 0.6 con al menos 2 tokens compartidos
        gp_toks = set(name_tokens(fullname_gp(g)))
        if len(gp_toks) < 2: continue
        best_sim, best_n = 0, None
        for toks, n in nf_by_tokens:
            inter = gp_toks & toks
            if len(inter) < 2: continue
            uni = gp_toks | toks
            sim = len(inter) / len(uni)
            # Bonus si nombre/apellido GP es subconjunto de NF (o viceversa)
            if gp_toks.issubset(toks) or toks.issubset(gp_toks):
                sim = max(sim, 0.85)
            if sim > best_sim:
                best_sim = sim; best_n = n
        if best_sim >= 0.6 and best_n is not None:
            use_match(cod, best_n, 'name')

    print('  matchings:', dict(methods_count), 'total matched:', len(matched))
    return matched, nf_used_ids


def fmt_fecha(s):
    if not s: return ''
    try:
        return datetime.fromisoformat(str(s).replace('Z','')).strftime('%d/%m/%Y')
    except Exception:
        return str(s)[:10]


# ── Lectura de marcas del usuario en Excel previo ──────────────────────────
def leer_marcas_usuario(path):
    """Devuelve dict id_noofit → marca y dict gp_codigo → marca."""
    if not os.path.exists(path):
        print(f'  (sin archivo previo {path}, sigo sin marcas)')
        return {}, {}
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    marcas_nf = {}
    marcas_gp = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 7).value
        if v is None: continue
        s = str(v).strip()
        if not s or (len(s) <= 10 and s.count('/') == 2): continue
        nf_id = ws.cell(r, 6).value
        gp_cod = ws.cell(r, 5).value
        if nf_id: marcas_nf[int(nf_id)] = s
        if gp_cod: marcas_gp[str(gp_cod).strip()] = s
    print(f'  marcas usuario: {len(marcas_nf)} por nf_id, {len(marcas_gp)} por gp_cod')
    return marcas_nf, marcas_gp


# ── Construcción de filas ──────────────────────────────────────────────────
def es_empleado(g):
    return g.get('empleado') == 1


def main():
    if not EMAIL or not PWD: sys.exit('ERROR: NOOFIT_EMAIL y NOOFIT_PASS')

    banner('1) Cargar dumps')
    with open(GP_DUMP, 'r', encoding='utf-8') as f: gp = json.load(f)
    with open(NF_DUMP, 'r', encoding='utf-8') as f: nf = json.load(f)
    nf_clientes = nf.get('clientes') or []
    # Deduplicar literal
    seen = {}
    for c in nf_clientes:
        if c['id'] not in seen: seen[c['id']] = c
    nf_clientes = list(seen.values())
    # Excluir clientes Gympass/Wellhub (ya gestionados aparte)
    PATRON_GYM = re.compile(r'wellhub|gympass|wellpass', re.I)
    def es_gympass_nf(c):
        if c.get('gympassId'): return True
        fields = ' '.join(str(c.get(k) or '') for k in ('alias','name','nombre','surname','apellidos','email'))
        return bool(PATRON_GYM.search(fields))
    def es_gympass_gp(g):
        fields = ' '.join(str(g.get(k) or '') for k in ('nombre','apellidos','email','notas'))
        return bool(PATRON_GYM.search(fields))
    n_nf_pre = len(nf_clientes)
    nf_clientes = [c for c in nf_clientes if not es_gympass_nf(c)]
    excluidos_gym_nf = n_nf_pre - len(nf_clientes)
    gp_altas = [g for g in gp['altas'] if not es_gympass_gp(g)]
    gp_bajas = [g for g in gp['bajas_recientes_12m'] if not es_gympass_gp(g)]
    excluidos_gym_gp = (len(gp['altas']) - len(gp_altas)) + (len(gp['bajas_recientes_12m']) - len(gp_bajas))
    gp = {**gp, 'altas': gp_altas, 'bajas_recientes_12m': gp_bajas}
    print(f'  Gympass excluidos: {excluidos_gym_nf} NF + {excluidos_gym_gp} GP')
    print(f'  GP altas: {len(gp["altas"])}, GP bajas12m: {len(gp["bajas_recientes_12m"])}, NF: {len(nf_clientes)}')

    banner('2) Marcas usuario')
    marcas_nf, marcas_gp = leer_marcas_usuario(USER_EDIT)

    # Inyectar matches manuales (gp código que el usuario añadió a un huérfano)
    # Si el usuario puso un código GP (numérico) en la marca de un nf_id, asociar.
    forzar_gp_para_nf = {}
    for nf_id, marca in marcas_nf.items():
        if marca.isdigit() or (marca.startswith('0') and marca[1:].isdigit()):
            cod = marca.zfill(8)  # 6174 → 00006174
            forzar_gp_para_nf[nf_id] = cod

    banner('3) Matching')
    todos_gp = gp['altas'] + gp['bajas_recientes_12m']
    matched_by_cod, nf_used_ids = build_match(todos_gp, nf_clientes)

    # Aplicar overrides manuales
    for nf_id, gp_cod in forzar_gp_para_nf.items():
        # Buscar gp por codigo (en altas + bajas + bajas_descartadas)
        candidatos = todos_gp + gp.get('bajas_descartadas', [])
        gp_match = next((c for c in todos_gp if c.get('codigo') == gp_cod), None)
        if gp_match is None:
            # Buscar en bajas descartadas (más antiguas) — lookup en dump completo
            print(f'  ⚠️ código GP {gp_cod} (de marca usuario) no está en altas/bajas12m')
            continue
        nf_obj = next((n for n in nf_clientes if n['id'] == nf_id), None)
        if nf_obj is None: continue
        matched_by_cod[gp_cod] = nf_obj
        nf_used_ids.add(nf_id)
        print(f'  ✅ Override manual: GP {gp_cod} ({fullname_gp(gp_match)}) ↔ NF {nf_id}')

    banner('4) Reservas para huérfanos NF')
    token, manager = login()
    huerfanos_ids = [n['id'] for n in nf_clientes if n['id'] not in nf_used_ids]
    print(f'  huérfanos a consultar: {len(huerfanos_ids)}')
    reservas = {}
    for i, cid in enumerate(huerfanos_ids, 1):
        reservas[cid] = get_reservas_count(token, manager, cid)
        if i % 10 == 0: print(f'    {i}/{len(huerfanos_ids)}')
        time.sleep(0.05)

    banner('5) Construir filas')
    filas = []

    # GP cruzado con NF
    for g in todos_gp:
        cod = g.get('codigo')
        es_alta_gp = g.get('estado') == 1
        n = matched_by_cod.get(cod)
        en_nf = n is not None
        archivado_nf = bool(n) and (n.get('enabled') is False or n.get('enabled') == 0)
        alta_nf = en_nf and not archivado_nf
        baja_gp = (not es_alta_gp) and bool(g.get('fechaBaja'))

        # Recibos 2026 del cliente GP
        recibos_2026 = len(g.get('_recibos') or [])

        # Marca usuario
        marca = ''
        if n and n['id'] in marcas_nf: marca = marcas_nf[n['id']]
        if cod in marcas_gp: marca = marcas_gp[cod]

        # Acción sugerida
        if marca.lower() == 'no':
            accion = '🚫 Ignorar (marcado por usuario)'
        elif marca.lower() == 'trabajador':
            accion = '👷 Trabajador (traspasar sin cuota)'
        elif es_empleado(g):
            accion = '👷 Empleado GP (traspasar sin cuota)'
        elif es_alta_gp and not en_nf:
            accion = '➕ CREAR en NoofitPro'
        elif es_alta_gp and archivado_nf:
            accion = '🔓 REACTIVAR en NoofitPro'
        elif baja_gp and alta_nf:
            accion = '📦 ARCHIVAR en NoofitPro'
        elif baja_gp and archivado_nf:
            accion = '✅ OK (archivado en ambos)'
        elif es_alta_gp and alta_nf:
            accion = '✅ OK (activo en ambos)'
        elif baja_gp and not en_nf:
            accion = '➖ Baja GP (no estaba en Noofit)'
        else:
            accion = '?'

        filas.append({
            'cliente': fullname_gp(g),
            'gestplus_marca': marca,
            'dni': g.get('dni') or '',
            'email': g.get('email') or '',
            'codigo_gp': cod,
            'noofit_id': (n.get('id') if n else ''),
            'alta_gestplus': fmt_fecha(g.get('fechaAlta')) if es_alta_gp else '',
            'alta_noofit': 'Sí' if alta_nf else 'No',
            'baja_gestplus': fmt_fecha(g.get('fechaBaja')) if baja_gp else '',
            'archivado_noofit': 'Sí' if archivado_nf else 'No',
            'reservas_nf': '',
            'recibos_2026': recibos_2026,
            'origen': 'ambos' if en_nf else 'gestplus',
            'accion': accion,
            'matched_via': '',  # podríamos rellenar pero ya tenemos contadores arriba
        })

    def fecha_alta_nf(n):
        """Devuelve la fecha de alta del cliente NoofitPro (lo que esté disponible)."""
        for k in ('dtCreated', 'editionDate', 'dtEditionDate'):
            v = n.get(k)
            if v and not str(v).startswith('-') and 'T' not in str(v)[:1]:
                # quitar parte hora si tiene
                s = str(v)[:10]
                if '-' in s and len(s) == 10: return s
                if isinstance(v, (int, float)) and v > 0:
                    try: return datetime.fromtimestamp(v/1000).strftime('%Y-%m-%d')
                    except Exception: pass
        return ''

    # Huérfanos NoofitPro
    for n in nf_clientes:
        if n['id'] in nf_used_ids: continue
        archivado = n.get('enabled') is False or n.get('enabled') == 0
        nombre = fullname_nf(n) or '(sin nombre)'
        marca = marcas_nf.get(n['id'], '')
        n_res = reservas.get(n['id'], 0)
        f_alta_nf = fecha_alta_nf(n)

        if marca.lower() == 'no':
            accion = '🚫 Ignorar (marcado por usuario)'
        elif marca.lower() == 'trabajador':
            accion = '👷 Trabajador (mantener sin cuota)'
        elif n_res > 0:
            accion = f'⚠️ NF activo con {n_res} reservas — revisar'
        else:
            accion = '🔍 Huérfano sin reservas — archivar?'

        filas.append({
            'cliente': nombre,
            'gestplus_marca': marca,
            'dni': n.get('dni') or n.get('nif') or '',
            'email': n.get('email') or '',
            'codigo_gp': '',
            'noofit_id': n['id'],
            'alta_gestplus': '',
            'alta_noofit': f_alta_nf or ('No' if archivado else 'Sí'),
            'baja_gestplus': '',
            'archivado_noofit': 'Sí' if archivado else 'No',
            'reservas_nf': n_res,
            'recibos_2026': '',
            'origen': 'noofit',
            'accion': accion,
            'matched_via': '',
        })

    # Filtro: solo dejar las acciones que requieren intervención
    KEEP_ACTIONS = ['CREAR', 'ARCHIVAR', 'Huérfano']
    filas = [f for f in filas if any(k in f['accion'] for k in KEEP_ACTIONS)]
    print(f'  Filas tras filtrar (solo CREAR/ARCHIVAR/Huérfano): {len(filas)}')

    # Orden
    orden = {'ambos':0,'gestplus':1,'noofit':2}
    filas.sort(key=lambda f: (orden.get(f['origen'],9), f['cliente'].upper()))

    banner('6) Excel')
    write_excel(filas)
    # Resumen
    from collections import Counter
    cnt = Counter(f['accion'] for f in filas)
    print('Resumen acciones:')
    for k, v in cnt.most_common():
        print(f'  {v:5d}  {k}')


def write_excel(filas):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Comparativa v2'
    headers = [
        ('Cliente', 32),
        ('Marca usuario', 14),
        ('DNI', 14),
        ('Email', 28),
        ('Cód. GestPlus', 12),
        ('ID NoofitPro', 11),
        ('Alta en GestPlus', 14),
        ('Alta en NoofitPro', 13),
        ('Baja en GestPlus', 14),
        ('Archivado NoofitPro', 14),
        ('Reservas NF', 11),
        ('Recibos 2026 GP', 13),
        ('Origen', 9),
        ('Acción sugerida', 36),
    ]
    keys = ['cliente','gestplus_marca','dni','email','codigo_gp','noofit_id',
            'alta_gestplus','alta_noofit','baja_gestplus','archivado_noofit',
            'reservas_nf','recibos_2026','origen','accion']
    fill_h = PatternFill('solid', fgColor='2DD4A8')
    font_h = Font(bold=True, color='FFFFFF', size=11)
    border = Border(*[Side(style='thin', color='CCCCCC')]*4)
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_h; c.font = font_h
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    f_si = PatternFill('solid', fgColor='D1FADF')
    f_no = PatternFill('solid', fgColor='FFE4E4')
    f_warn = PatternFill('solid', fgColor='FFF7CC')
    f_ignore = PatternFill('solid', fgColor='E5E7EB')
    f_worker = PatternFill('solid', fgColor='DBEAFE')

    for i, f in enumerate(filas, 2):
        for j, k in enumerate(keys, 1):
            v = f.get(k, '')
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = border
            if k == 'accion':
                if 'Ignorar' in v: cell.fill = f_ignore
                elif 'Trabajador' in v or 'Empleado' in v: cell.fill = f_worker
                elif 'CREAR' in v: cell.fill = f_no
                elif 'ARCHIVAR' in v: cell.fill = f_warn
                elif 'OK' in v: cell.fill = f_si
                elif 'Huérfano' in v: cell.fill = f_warn
                elif 'reservas' in v.lower() and 'revisar' in v.lower(): cell.fill = f_warn
            elif k == 'gestplus_marca' and v:
                cell.fill = f_worker if v.lower() == 'trabajador' else (f_ignore if v.lower() == 'no' else f_si)

    ws.auto_filter.ref = ws.dimensions
    wb.save(OUT)
    print(f'  → {OUT}')


if __name__ == '__main__':
    main()
