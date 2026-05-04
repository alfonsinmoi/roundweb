"""Detecta clientes DUPLICADOS activos en NoofitPro y consulta reservas
de cada uno para sugerir cuál archivar.

Criterios de duplicado (en orden):
  1. DNI igual (normalizado)
  2. Email igual
  3. Nombre+Apellidos+fechaNacimiento iguales

Para cada grupo duplicado, llama a getReservasByUser para contar reservas
y propone archivar el ID con menos reservas (o el más antiguo si empatan).

Uso:
  NOOFIT_EMAIL=... NOOFIT_PASS=... \
  python find_duplicates_noofit.py [noofit_clientes_dump.json] [salida.xlsx]
"""
import os, sys, json, hashlib, time
from datetime import datetime
from collections import defaultdict
import requests, urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
VERIFY = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = 'https://pro.wiemspro.com/wiemspro'
APP_VERSION = '1.8.39'
APP_ID = '1'

EMAIL = os.getenv('NOOFIT_EMAIL', '')
PWD = os.getenv('NOOFIT_PASS', '')
if not EMAIL or not PWD:
    sys.exit('ERROR: define NOOFIT_EMAIL y NOOFIT_PASS')


def banner(s): print(f'\n{"="*60}\n{s}\n{"="*60}')


def login():
    body = {'email': EMAIL, 'appVersion': APP_VERSION,
            'password': hashlib.md5(PWD.encode()).hexdigest().upper()}
    r = requests.post(f'{BASE}/account/loginEasy', json=body, verify=VERIFY, timeout=30)
    if r.status_code != 200: sys.exit('Login fallido: ' + r.text[:200])
    return r.headers.get('X-CustomToken'), r.headers.get('X-TRAINER_MANAGER', '')


def get_reservas(token, manager, id_cliente):
    headers = {
        'X-CustomToken': token,
        'X-TRAINER_MANAGER': manager,
        'locale': 'es', 'appVersion': APP_VERSION, 'appId': APP_ID,
        'Content-Type': 'application/json',
        'initialId': '0',
    }
    r = requests.post(f'{BASE}/api/dispositivos/getReservasByUser',
                      json={'id': id_cliente}, headers=headers,
                      verify=VERIFY, timeout=30)
    if r.status_code != 200: return None
    try:
        d = r.json()
        if d.get('mensaje') != 'OK': return None
        return d.get('clases') or d.get('reservas') or []
    except Exception:
        return None


def normaliza_dni(d):
    return (d or '').strip().upper().replace(' ', '').replace('-', '').replace('.', '')


def normaliza_email(e):
    return (e or '').strip().lower()


def normaliza_nombre(n):
    return (n or '').strip().upper().replace('  ', ' ')


def fullname(c):
    n = c.get('nombre') or c.get('name') or ''
    a = c.get('apellidos') or c.get('surname') or ''
    return f'{n} {a}'.strip()


def encontrar_duplicados(clientes):
    """Devuelve lista de grupos. Cada grupo es lista de clientes."""
    activos = [c for c in clientes if c.get('enabled') is not False and c.get('enabled') != 0]
    print(f'  activos: {len(activos)} de {len(clientes)}')

    # Indices
    by_dni = defaultdict(list)
    by_email = defaultdict(list)
    by_nombre = defaultdict(list)

    for c in activos:
        d = normaliza_dni(c.get('dni') or c.get('nif'))
        e = normaliza_email(c.get('email'))
        nm = normaliza_nombre(fullname(c))
        bd = (c.get('birthdate') or '')[:10]
        if d and len(d) >= 8: by_dni[d].append(c)
        if e and '@' in e:    by_email[e].append(c)
        if nm and bd:         by_nombre[f'{nm}|{bd}'].append(c)

    # Construir grupos: union-find por id
    parent = {c['id']: c['id'] for c in activos}
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb: parent[ra] = rb

    for groups in (by_dni.values(), by_email.values(), by_nombre.values()):
        for g in groups:
            if len(g) < 2: continue
            for c in g[1:]:
                union(g[0]['id'], c['id'])

    # Reagrupar por root
    grupos_dict = defaultdict(list)
    for c in activos:
        grupos_dict[find(c['id'])].append(c)
    grupos = [g for g in grupos_dict.values() if len(g) >= 2]
    print(f'  grupos duplicados detectados: {len(grupos)}')
    return grupos


def write_excel(grupos, reservas_count, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Duplicados'

    headers = [
        ('Grupo #', 8),
        ('Cliente', 32),
        ('DNI', 14),
        ('Email', 28),
        ('ID NoofitPro', 12),
        ('Birthdate', 12),
        ('Última actividad', 12),
        ('Bonos disponibles', 12),
        ('Nº reservas', 11),
        ('Acción sugerida', 28),
    ]
    fill_h = PatternFill('solid', fgColor='2DD4A8')
    font_h = Font(bold=True, color='FFFFFF', size=11)
    border = Border(*[Side(style='thin', color='CCCCCC')]*4)
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_h; c.font = font_h
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    fill_keep = PatternFill('solid', fgColor='D1FADF')
    fill_drop = PatternFill('solid', fgColor='FFE4E4')
    fill_mark = PatternFill('solid', fgColor='FFF7CC')

    row = 2
    for gi, g in enumerate(grupos, 1):
        # Decidir qué cliente es "mejor" (más reservas, sino más reciente)
        decorated = []
        for c in g:
            n = reservas_count.get(c['id'], 0)
            decorated.append((n, c.get('id'), c))
        decorated.sort(reverse=True)  # más reservas primero
        keep_id = decorated[0][2]['id']
        max_res = decorated[0][0]

        for n_res, _id, c in decorated:
            accion = ''
            fill = None
            if c['id'] == keep_id:
                accion = ('🔒 MANTENER (más usado)' if max_res > 0 else '⚠️ Sin reservas — revisar manualmente')
                fill = fill_keep if max_res > 0 else fill_mark
            else:
                if n_res == 0:
                    accion = '🗑️ ARCHIVAR (sin reservas)'
                else:
                    accion = f'⚠️ ARCHIVAR? tiene {n_res} reservas'
                fill = fill_drop if n_res == 0 else fill_mark

            ws.cell(row=row, column=1, value=gi)
            ws.cell(row=row, column=2, value=fullname(c) or '(sin nombre)')
            ws.cell(row=row, column=3, value=c.get('dni') or c.get('nif') or '')
            ws.cell(row=row, column=4, value=c.get('email') or '')
            ws.cell(row=row, column=5, value=c.get('id'))
            ws.cell(row=row, column=6, value=(c.get('birthdate') or '')[:10])
            # última actividad
            ult = c.get('lastSession') or c.get('updatedAt') or c.get('createdAt') or ''
            ws.cell(row=row, column=7, value=str(ult)[:10])
            ws.cell(row=row, column=8, value=c.get('bonosDisponibles', ''))
            ws.cell(row=row, column=9, value=n_res)
            ws.cell(row=row, column=10, value=accion)
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if fill: cell.fill = fill
            row += 1

    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)
    print(f'  → {out_path}')


def main():
    nf_path = sys.argv[1] if len(sys.argv) > 1 else 'noofit_clientes_dump.json'
    out = sys.argv[2] if len(sys.argv) > 2 else 'duplicados_noofit.xlsx'
    if not os.path.exists(nf_path): sys.exit(f'No existe {nf_path}')

    banner('1) Login NoofitPro')
    token, manager = login()
    print(f'  token: {(token or "")[:12]}... manager: {manager}')

    banner('2) Cargar clientes')
    with open(nf_path, 'r', encoding='utf-8') as f:
        clientes = (json.load(f).get('clientes')) or []
    print(f'  total: {len(clientes)}')

    banner('3) Detectar duplicados')
    grupos = encontrar_duplicados(clientes)

    banner('4) Consultar reservas de cada duplicado')
    todos_ids = sorted({c['id'] for g in grupos for c in g})
    print(f'  IDs a consultar: {len(todos_ids)}')
    reservas_count = {}
    for i, cid in enumerate(todos_ids, 1):
        rsv = get_reservas(token, manager, cid)
        reservas_count[cid] = len(rsv or [])
        if i % 10 == 0: print(f'   {i}/{len(todos_ids)}')
        time.sleep(0.05)

    banner('5) Generar Excel')
    write_excel(grupos, reservas_count, out)
    # Resumen
    total_dup = sum(len(g) for g in grupos)
    a_archivar_seguro = sum(
        1 for g in grupos for c in g
        if reservas_count.get(c['id'], 0) == 0
        and c['id'] != sorted(g, key=lambda x: -reservas_count.get(x['id'], 0))[0]['id']
    )
    print(f'\n  Grupos: {len(grupos)} · Clientes implicados: {total_dup}')
    print(f'  Candidatos a archivar (sin reservas y duplicado): {a_archivar_seguro}')


if __name__ == '__main__':
    main()
